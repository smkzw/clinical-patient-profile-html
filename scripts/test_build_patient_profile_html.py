from __future__ import annotations

import csv
import importlib.util
import os
import shutil
import subprocess
import tempfile
import unittest
from pathlib import Path

import openpyxl


REPO_ROOT = Path(__file__).resolve().parents[1]
SCRIPT_PATH = REPO_ROOT / "scripts" / "build_patient_profile_html.py"
SPEC = importlib.util.spec_from_file_location("build_patient_profile_html", SCRIPT_PATH)
MODULE = importlib.util.module_from_spec(SPEC)
assert SPEC and SPEC.loader
SPEC.loader.exec_module(MODULE)


def create_listing_workbook(path: Path) -> None:
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "SUBJ--受试者页"
    ws.append(["研究中心", "中心编号", "受试者", "受试者状态"])
    ws.append(["", "", "", ""])
    ws.append(["示例中心医院", "01", "S01001", "完成试验"])

    ws = wb.create_sheet("DM--人口统计学")
    ws.append(["研究中心", "受试者", "年龄（系统生成）", "性别"])
    ws.append(["", "", "", ""])
    ws.append(["示例中心医院", "S01001", 28, "女性"])

    ws = wb.create_sheet("RAND--入组随机页")
    ws.append(["研究中心", "受试者", "入组日期", "随机日期和时间（系统生成）", "受试者是否入组本研究？", "是否进行PK血样采集？", "未入组原因", "其他原因，请说明"])
    ws.append(["", "", "", "", "", "", "", ""])
    ws.append(["示例中心医院", "S01001", "2026-01-02", "2026-01-02", "是", "否", "", ""])

    ws = wb.create_sheet("IC--知情同意")
    ws.append(["研究中心", "受试者", "首次签署知情同意书日期"])
    ws.append(["", "", ""])
    ws.append(["示例中心医院", "S01001", "2026-01-01"])

    ws = wb.create_sheet("SV--访视日期")
    ws.append(["研究中心", "受试者", "访视OID", "访视名称", "访视日期", "计划最早访视日期（系统生成）", "计划最晚访视日期（系统生成）"])
    ws.append(["", "", "", "", "", "", ""])
    ws.append(["示例中心医院", "S01001", "SCR", "筛选期", "2026-01-01", "2026-01-01", "2026-01-01"])
    ws.append(["示例中心医院", "S01001", "D1", "基线", "2026-01-02", "2026-01-02", "2026-01-02"])
    ws.append(["示例中心医院", "S01001", "D15", "第15天", "2026-01-16", "2026-01-15", "2026-01-17"])

    ws = wb.create_sheet("AE--不良事件")
    ws.append(["研究中心", "受试者", "是否发生不良事件？", "是否为严重不良事件（SAE）？", "不良事件名称", "最早开始日期", "最严重程度（CTCAE V5.0）", "与研究用药的关系", "转归", "结束日期"])
    ws.append(["", "", "", "", "", "", "", "", "", ""])
    ws.append(["示例中心医院", "S01001", "否", "否", "", "", "", "", "", ""])

    ws = wb.create_sheet("VS--生命体征")
    ws.append(["研究中心", "受试者", "访视OID", "访视名称", "是否进行生命体征检查？", "检查日期", "临床意义评价", "异常有临床意义，请说明", "体温", "体温_UNIT", "脉搏", "脉搏_UNIT", "收缩压", "收缩压_UNIT", "舒张压", "舒张压_UNIT"])
    ws.append(["", "", "", "", "", "", "", "", "", "", "", "", "", "", "", ""])
    ws.append(["示例中心医院", "S01001", "D1", "基线", "是", "2026-01-02", "正常", "", 36.7, "°C", 72, "次/分", 118, "mmHg", 76, "mmHg"])
    ws.append(["示例中心医院", "S01001", "D15", "第15天", "是", "2026-01-16", "正常", "", 36.8, "°C", 70, "次/分", 120, "mmHg", 78, "mmHg"])

    ws = wb.create_sheet("EASI--湿疹面积及严重程度指数评分（EASI）")
    ws.append(["研究中心", "受试者", "访视OID", "访视名称", "评估日期", "EASI得分（4个部位的评分总和）"])
    ws.append(["", "", "", "", "", ""])
    ws.append(["示例中心医院", "S01001", "D1", "基线", "2026-01-02", 18])
    ws.append(["示例中心医院", "S01001", "D15", "第15天", "2026-01-16", 9])

    ws = wb.create_sheet("IGA--研究者整体评分（IGA）")
    ws.append(["研究中心", "受试者", "访视OID", "访视名称", "评估日期", "IGA得分"])
    ws.append(["", "", "", "", "", ""])
    ws.append(["示例中心医院", "S01001", "D1", "基线", "2026-01-02", "3分-中度"])
    ws.append(["示例中心医院", "S01001", "D15", "第15天", "2026-01-16", "1分-几乎清除"])

    ws = wb.create_sheet("NRS--瘙痒数值评定量表（Itch NRS）-平均值")
    ws.append(["研究中心", "受试者", "访视OID", "访视名称", "评估日期", "平均瘙痒程度"])
    ws.append(["", "", "", "", "", ""])
    ws.append(["示例中心医院", "S01001", "D1", "基线", "2026-01-02", 7])
    ws.append(["示例中心医院", "S01001", "D15", "第15天", "2026-01-16", 3])

    ws = wb.create_sheet("LBHEMA--实验室检查-血常规")
    ws.append(["研究中心", "受试者", "访视OID", "访视名称", "采样日期", "检查项目", "检查结果", "单位", "下限", "上限", "标准值", "临床意义评价", "异常有临床意义，请说明", "正常值范围标记"])
    ws.append(["", "", "", "", "", "", "", "", "", "", "", "", "", ""])
    ws.append(["示例中心医院", "S01001", "D1", "基线", "2026-01-02", "白细胞计数", 6.2, "10^9/L", 3.5, 9.5, "3.5-9.5", "正常", "", "正常"])
    ws.append(["示例中心医院", "S01001", "D15", "第15天", "2026-01-16", "白细胞计数", 5.8, "10^9/L", 3.5, 9.5, "3.5-9.5", "正常", "", "正常"])

    ws = wb.create_sheet("EG--12导联心电图")
    ws.append(["研究中心", "受试者", "访视OID", "访视名称", "是否进行12导联心电图检查？", "检查日期", "QTc间期", "单位", "临床意义评价", "若异常，请详述"])
    ws.append(["", "", "", "", "", "", "", "", "", ""])
    ws.append(["示例中心医院", "S01001", "D1", "基线", "是", "2026-01-02", 410, "ms", "正常", ""])
    ws.append(["示例中心医院", "S01001", "D15", "第15天", "是", "2026-01-16", 412, "ms", "正常", ""])

    wb.save(path)


def create_finding_workbook(path: Path) -> None:
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "示例中心-自查问题汇总"
    ws.append(["序号", "中心", "受试者编号", "发现日期", "分类", "一类分级", "访视", "自查问题描述", "问题状态（关闭/未关闭/无法整改，解释）", "PM审核后回复", "研究者确认回复", "自查问题整改措施"])
    ws.append([1, "示例中心医院", "S01001", "2026-01-20", "疗效评价", "低", "D15", "量表原始记录需复核。", "已整改", "已核对", "中心已确认", "补充原始记录索引"])
    wb.save(path)


def create_pd_definition_workbook(path: Path) -> None:
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "方案偏离分类"
    ws.append(["分类", "说明"])
    ws.append(["一般偏离", "示例"])
    wb.save(path)


def create_protocol_text(path: Path) -> None:
    path.write_text(
        "\n".join(
            [
                "研究方案",
                "主要终点：第15天 EASI 评分较基线变化。",
                "关键次要终点：第15天达到 EASI-75 的受试者比例。",
                "次要终点：第15天 IGA 评分变化。",
                "次要终点：第15天达到 IGA-TS 的受试者比例。",
                "研究流程：筛选期、基线（D1）、第15天（D15）进行疗效评估。",
                "实验室、生命体征和心电图在W0/D1进行；若筛选期访视发生在给药前7天内，可接受筛选期结果作为基线，D1无需重复。",
                "受试者如提前退出，应完成提前退出访视。",
            ]
        ),
        encoding="utf-8",
    )


class BuildPatientProfileHtmlTest(unittest.TestCase):
    def setUp(self) -> None:
        self.temp_root = Path(tempfile.mkdtemp(prefix="patient-profile-skill-", dir="/private/tmp"))
        self.project_dir = self.temp_root / "project"
        self.output_dir = self.temp_root / "output"
        self.project_dir.mkdir(parents=True, exist_ok=True)
        create_listing_workbook(self.project_dir / "demo_listing.xlsx")
        create_finding_workbook(self.project_dir / "demo_finding.xlsx")
        create_pd_definition_workbook(self.project_dir / "demo_pd_definition.xlsx")
        create_protocol_text(self.project_dir / "研究方案.txt")

    def tearDown(self) -> None:
        shutil.rmtree(self.temp_root, ignore_errors=True)

    def run_builder(self, *extra_args: str, include_finding: bool = True) -> None:
        cmd = [
            "python3",
            str(SCRIPT_PATH),
            "--work-dir",
            str(self.project_dir),
            "--listing-xlsx",
            str(self.project_dir / "demo_listing.xlsx"),
            "--pd-def-xlsx",
            str(self.project_dir / "demo_pd_definition.xlsx"),
            "--centers",
            "示例中心医院",
            "--subject-scope",
            "randomized",
            "--include-usv",
            "no",
            "--output-dir",
            str(self.output_dir),
            *extra_args,
        ]
        if include_finding:
            cmd.extend(["--finding-xlsx", str(self.project_dir / "demo_finding.xlsx")])
        subprocess.run(cmd, check=True)

    def count_rows(self, file_name: str) -> int:
        return self.count_rows_from_dir(self.output_dir, file_name)

    @staticmethod
    def count_rows_from_dir(directory: Path, file_name: str) -> int:
        with (directory / file_name).open(encoding="utf-8-sig", newline="") as handle:
            return max(sum(1 for _ in csv.reader(handle)) - 1, 0)

    def test_precheck_only_writes_reports(self) -> None:
        self.run_builder("--precheck-only")
        self.assertTrue((self.output_dir / "input_precheck.md").exists())
        self.assertTrue((self.output_dir / "input_precheck.json").exists())
        self.assertTrue((self.output_dir / "suggested_project_config.json").exists())
        self.assertTrue((self.output_dir / "protocol_endpoint_summary.md").exists())
        precheck = (self.output_dir / "input_precheck.md").read_text(encoding="utf-8")
        self.assertIn("可继续构建：是", precheck)
        self.assertIn("示例中心医院", precheck)
        protocol_summary = (self.output_dir / "protocol_endpoint_summary.md").read_text(encoding="utf-8")
        self.assertIn("EASI", protocol_summary)
        self.assertIn("IGA", protocol_summary)

    def test_full_build_writes_expected_outputs(self) -> None:
        self.run_builder()
        for name in [
            "patient_profile.html",
            "cleaned_subject_profile_dataset.csv",
            "efficacy_longitudinal_dataset.csv",
            "lab_longitudinal_dataset.csv",
            "vital_signs_longitudinal_dataset.csv",
            "finding_subject_level_dataset.csv",
        ]:
            self.assertTrue((self.output_dir / name).exists(), name)
        self.assertGreater(self.count_rows("cleaned_subject_profile_dataset.csv"), 0)
        self.assertGreater(self.count_rows("efficacy_longitudinal_dataset.csv"), 0)
        self.assertGreater(self.count_rows("lab_longitudinal_dataset.csv"), 0)
        html = (self.output_dir / "patient_profile.html").read_text(encoding="utf-8")
        self.assertIn("示例中心医院", html)
        self.assertIn("受试者Patient Profile", html)
        self.assertIn("疗效数据模块", html)
        efficacy_csv = (self.output_dir / "efficacy_longitudinal_dataset.csv").read_text(encoding="utf-8-sig")
        self.assertIn("EASI", efficacy_csv)
        self.assertIn("IGA", efficacy_csv)
        self.assertNotIn("NRS", efficacy_csv)

    def test_build_with_generated_config(self) -> None:
        self.run_builder("--precheck-only")
        config_path = self.output_dir / "suggested_project_config.json"
        second_dir = self.temp_root / "from-config"
        cmd = [
            "python3",
            str(SCRIPT_PATH),
            "--work-dir",
            str(self.project_dir),
            "--listing-xlsx",
            str(self.project_dir / "demo_listing.xlsx"),
            "--finding-xlsx",
            str(self.project_dir / "demo_finding.xlsx"),
            "--pd-def-xlsx",
            str(self.project_dir / "demo_pd_definition.xlsx"),
            "--subject-scope",
            "randomized",
            "--include-usv",
            "no",
            "--config-json",
            str(config_path),
            "--output-dir",
            str(second_dir),
        ]
        subprocess.run(cmd, check=True)
        self.assertTrue((second_dir / "patient_profile.html").exists())
        self.assertGreater(self.count_rows_from_dir(second_dir, "efficacy_longitudinal_dataset.csv"), 0)

    def test_precheck_blocks_when_protocol_missing(self) -> None:
        os.remove(self.project_dir / "研究方案.txt")
        self.run_builder("--precheck-only")
        precheck = (self.output_dir / "input_precheck.md").read_text(encoding="utf-8")
        self.assertIn("[阻断] 方案文件", precheck)

    def test_group_label_and_visit_name_are_generic(self) -> None:
        self.assertEqual(MODULE.summary_group_label("试验组"), "试验组")
        self.assertEqual(MODULE.summary_group_label("对照组"), "对照组")
        self.assertEqual(MODULE.summary_group_label("placebo"), "对照组")
        self.assertEqual(MODULE.simplify_visit_label("D29±1d", "双盲治疗期V5（D29±1d）"), "D29")
        self.assertEqual(MODULE.simplify_visit_label("SCR", "筛选/导入期V1（D-7~D-1）"), "SCR")

    def test_protocol_summary_detects_baseline_and_early_exit(self) -> None:
        summary = MODULE.detect_protocol_endpoint_summary([self.project_dir / "研究方案.txt"])
        self.assertEqual(summary["baseline_rules"]["实验室"]["status"], "confirmed")
        self.assertEqual(summary["baseline_rules"]["心电图"]["status"], "confirmed")
        self.assertEqual(summary["baseline_rules"]["生命体征"]["status"], "confirmed")
        self.assertTrue(summary["baseline_rules"]["has_early_exit"])

    def test_protocol_summary_extracts_response_rules_from_protocol_text(self) -> None:
        summary = MODULE.detect_protocol_endpoint_summary([self.project_dir / "研究方案.txt"])
        labels = {item["label"] for item in summary["response_rules"]}
        self.assertIn("EASI-75", labels)
        self.assertIn("IGA-TS", labels)

    def test_no_protocol_response_definition_keeps_response_blank(self) -> None:
        protocol_path = self.project_dir / "only_continuous.txt"
        protocol_path.write_text(
            "\n".join(
                [
                    "研究方案",
                    "主要终点：第15天 EASI 评分较基线变化。",
                    "次要终点：第15天 IGA 评分变化。",
                ]
            ),
            encoding="utf-8",
        )
        summary = MODULE.detect_protocol_endpoint_summary([protocol_path])
        self.assertEqual(summary["response_rules"], [])
        original_summary = MODULE.PROTOCOL_SUMMARY
        try:
            MODULE.PROTOCOL_SUMMARY = summary
            self.assertEqual(MODULE.numeric_response_label("EASI", 18.0, 9.0), "")
            self.assertEqual(MODULE.numeric_response_label("IGA", 3.0, 1.0), "")
        finally:
            MODULE.PROTOCOL_SUMMARY = original_summary

    def test_detect_efficacy_config_uses_protocol_metrics_not_preset_catalog(self) -> None:
        protocol_path = self.project_dir / "custom_metric.txt"
        protocol_path.write_text(
            "\n".join(
                [
                    "研究方案",
                    "主要终点：第15天 鼻部症状综合评分（NSS）较基线变化。",
                    "研究流程：D1和D15进行鼻部症状综合评分（NSS）评估。",
                ]
            ),
            encoding="utf-8",
        )
        listing_path = self.project_dir / "custom_listing.xlsx"
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "NSS--鼻部症状综合评分"
        ws.append(["研究中心", "受试者", "访视OID", "访视名称", "评估日期", "NSS总分"])
        ws.append(["示例中心医院", "S01001", "D1", "基线", "2026-01-02", 9])
        ws.append(["示例中心医院", "S01001", "D15", "第15天", "2026-01-16", 4])
        wb.save(listing_path)
        summary = MODULE.detect_protocol_endpoint_summary([protocol_path])
        configs = MODULE.detect_efficacy_config(listing_path, summary)
        self.assertEqual(summary["selected_metrics"][0]["metric"], "NSS")
        self.assertEqual(configs[0]["metric"], "NSS")
        self.assertEqual(configs[0]["sheet"], "NSS--鼻部症状综合评分")

    def test_build_without_finding_hides_finding_ui(self) -> None:
        os.remove(self.project_dir / "demo_finding.xlsx")
        self.run_builder(include_finding=False)
        html = (self.output_dir / "patient_profile.html").read_text(encoding="utf-8")
        self.assertNotIn("Finding筛选", html)
        self.assertIn("疗效、实验室与生命体征的中文交互式核查视图", html)
        self.assertIn('"has_findings": false', html)
        self.assertNotIn("是否进入ITT", html)
        self.assertNotIn("是否进入PKCS", html)


@unittest.skipUnless(os.environ.get("PATIENT_PROFILE_REAL_FIXTURE_DIR"), "Optional regression test requires PATIENT_PROFILE_REAL_FIXTURE_DIR")
class BuildPatientProfileHtmlRegressionTest(unittest.TestCase):
    def test_real_fixture_precheck(self) -> None:
        fixture_dir = Path(os.environ["PATIENT_PROFILE_REAL_FIXTURE_DIR"]).resolve()
        output_dir = Path(tempfile.mkdtemp(prefix="patient-profile-regression-", dir="/private/tmp"))
        try:
            cmd = [
                "python3",
                str(SCRIPT_PATH),
                "--work-dir",
                str(fixture_dir),
                "--output-dir",
                str(output_dir),
                "--precheck-only",
            ]
            subprocess.run(cmd, check=True)
            self.assertTrue((output_dir / "input_precheck.md").exists())
            self.assertTrue((output_dir / "suggested_project_config.json").exists())
        finally:
            shutil.rmtree(output_dir, ignore_errors=True)


if __name__ == "__main__":
    unittest.main()
