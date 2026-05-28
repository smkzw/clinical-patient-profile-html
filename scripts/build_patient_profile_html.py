from __future__ import annotations

import argparse
import csv
import json
import math
import re
import zipfile
from copy import deepcopy
from collections import Counter, defaultdict
from datetime import date, datetime
from pathlib import Path
from typing import Any

import openpyxl
import xlrd

try:
    from pypdf import PdfReader  # type: ignore
except Exception:  # pragma: no cover - optional dependency
    PdfReader = None


ROOT = Path(__file__).resolve().parents[1]
WORK = ROOT.parent
OUTPUT_DIR = ROOT / "output"

LISTING_XLSX = WORK / "__listing__.xlsx"
FINDING_XLSX = WORK / "__finding__.xlsx"
PD_DEF_XLSX: Path | None = None
SUBJECT_REPORT_XLS = WORK / "__subject_report__.xls"
REF_HTMLS: list[Path] = []

TARGET_CENTERS: list[str] = []
PROTOCOL_FILES: list[Path] = []
FINDING_SHEET_SPECS: list[dict[str, str]] = []
PROTOCOL_SUMMARY: dict[str, Any] = {"files": [], "endpoint_items": [], "selected_metrics": [], "response_rules": [], "visit_mentions": []}
MISSING_TEXT = "未在现有文件中确认"

REQUIRED_LISTING_SHEETS = [
    "SUBJ",
    "DM",
    "RAND",
    "IC",
    "AE",
    "VS",
]

FILE_PATTERNS = {
    "listing": [r"列表.*数据集", r"listing", r"受试者级"],
    "finding": [r"自查.*问题", r"稽查.*问题", r"finding"],
    "pd_def": [r"偏离分类", r"方案偏离"],
    "subject_report": [r"受试者报表", r"subject.*report"],
    "protocol": [r"研究方案", r"试验方案", r"protocol", r"勘误"],
    "reference_html": [r"subject_timeline.*\.html", r"patient_profile.*\.html"],
}

SUBJECT_ID_RE = re.compile(r"S\d{5}")

CORE_SHEET_CATALOG = {
    "SV": {"required": True, "patterns": [r"^SV--", r"访视日期"]},
    "SUBJ": {"required": True, "patterns": [r"^SUBJ--", r"受试者页"]},
    "DM": {"required": True, "patterns": [r"^DM--", r"人口统计"]},
    "RAND": {"required": True, "patterns": [r"^RAND--", r"入组随机", r"随机页"]},
    "IC": {"required": True, "patterns": [r"^IC--", r"知情同意"]},
    "DSEOT": {"required": False, "patterns": [r"^DSEOT--", r"治疗结束"]},
    "DSEOS": {"required": False, "patterns": [r"^DSEOS--", r"研究结束"]},
    "AE": {"required": True, "patterns": [r"^AE--", r"不良事件"]},
    "PC": {"required": False, "patterns": [r"^PC--", r"药代", r"PK"]},
    "EG": {"required": False, "patterns": [r"^EG--", r"心电图", r"12导联"]},
    "VS": {"required": True, "patterns": [r"^VS--", r"生命体征"]},
}


def parse_args() -> argparse.Namespace:
    parser = argparse.ArgumentParser(description="Build clinical patient profile HTML from listing/finding workbooks.")
    parser.add_argument("--work-dir", default=".", help="Directory containing input files.")
    parser.add_argument("--listing-xlsx", help="Path to the subject/listing workbook.")
    parser.add_argument("--finding-xlsx", help="Path to the finding workbook.")
    parser.add_argument("--pd-def-xlsx", help="Path to the protocol deviation definition workbook.")
    parser.add_argument("--subject-report-xls", help="Path to the subject report workbook.")
    parser.add_argument("--protocol-file", action="append", default=[], help="Optional protocol or amendment files.")
    parser.add_argument("--ref-html", action="append", default=[], help="Optional reference HTML files for style/group hints.")
    parser.add_argument("--config-json", help="Optional project config JSON generated from a previous precheck.")
    parser.add_argument("--output-dir", help="Directory for generated CSV/HTML outputs.")
    parser.add_argument("--centers", help="Semicolon-separated centers to include. Default: auto-detect all centers in listing.")
    parser.add_argument("--precheck-only", action="store_true", help="Only scan inputs and write a precheck report.")
    return parser.parse_args()


def find_candidate_files(work_dir: Path, pattern_group: str, suffixes: tuple[str, ...]) -> list[Path]:
    patterns = FILE_PATTERNS[pattern_group]
    candidates: list[Path] = []
    for path in sorted(work_dir.iterdir()):
        if not path.is_file():
            continue
        if path.suffix.lower() not in suffixes:
            continue
        if any(re.search(pattern, path.name, re.I) for pattern in patterns):
            candidates.append(path)
    return candidates


def pick_first_existing(paths: list[Path]) -> Path | None:
    for path in paths:
        if path.exists():
            return path
    return None


def score_sheet_name(sheet_name: str, patterns: list[str]) -> int:
    score = 0
    for pattern in patterns:
        if re.search(pattern, sheet_name, re.I):
            score += 1
    return score


def choose_best_sheet(sheet_names: list[str], patterns: list[str]) -> str:
    ranked = sorted(
        ((score_sheet_name(name, patterns), name) for name in sheet_names),
        key=lambda item: (-item[0], item[1]),
    )
    if ranked and ranked[0][0] > 0:
        return ranked[0][1]
    return ""


def choose_candidate_column(headers: list[str], candidates: list[str]) -> str:
    for candidate in candidates:
        if candidate in headers:
            return candidate
    normalized_headers = {re.sub(r"[\s_\-（）()]+", "", header).lower(): header for header in headers}
    for candidate in candidates:
        key = re.sub(r"[\s_\-（）()]+", "", candidate).lower()
        if key in normalized_headers:
            return normalized_headers[key]
    return ""


def detect_sheet_aliases(path: Path) -> dict[str, str]:
    wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
    try:
        sheet_names = wb.sheetnames[:]
    finally:
        wb.close()
    aliases: dict[str, str] = {}
    for alias, spec in CORE_SHEET_CATALOG.items():
        aliases[alias] = choose_best_sheet(sheet_names, spec["patterns"])
    return aliases


def detect_efficacy_config(path: Path) -> list[dict[str, Any]]:
    wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
    try:
        sheet_names = wb.sheetnames[:]
    finally:
        wb.close()
    configs: list[dict[str, Any]] = []
    for spec in DEFAULT_EFFICACY_CONFIG:
        sheet_name = choose_best_sheet(sheet_names, spec.get("sheet_patterns", []))
        if not sheet_name:
            continue
        sheet = read_listing_sheet(path, sheet_name)
        headers = sheet["headers"]
        chosen_value_col = choose_candidate_column(headers, spec.get("value_candidates", [spec["value_col"]]))
        chosen_date_col = choose_candidate_column(headers, spec.get("date_candidates", [spec["date_col"]]))
        value_col = chosen_value_col or spec["value_col"]
        date_col = chosen_date_col or spec["date_col"]
        resolved = deepcopy(spec)
        resolved["sheet"] = sheet_name
        resolved["value_col"] = value_col
        resolved["date_col"] = date_col
        resolved["value_col_confirmed"] = bool(chosen_value_col)
        resolved["date_col_confirmed"] = bool(chosen_date_col)
        configs.append(resolved)
    return configs


def select_efficacy_config_by_protocol(
    detected_efficacy: list[dict[str, Any]],
    protocol_summary: dict[str, Any],
) -> list[dict[str, Any]]:
    selected_metrics = {item["metric"]: item for item in protocol_summary.get("selected_metrics", [])}
    if not selected_metrics:
        return []
    selected_configs: list[dict[str, Any]] = []
    for config in detected_efficacy:
        selected = selected_metrics.get(config["metric"])
        if not selected:
            continue
        resolved = deepcopy(config)
        resolved["endpoint_role"] = selected.get("endpoint_role", "未分类")
        resolved["variable_type"] = selected.get("variable_type", "continuous")
        selected_configs.append(resolved)
    return selected_configs


def detect_lab_config(path: Path) -> list[dict[str, str]]:
    wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
    try:
        sheet_names = wb.sheetnames[:]
    finally:
        wb.close()
    configs: list[dict[str, str]] = []
    for spec in DEFAULT_LAB_CONFIG:
        sheet_name = choose_best_sheet(sheet_names, spec.get("sheet_patterns", []))
        if not sheet_name:
            continue
        configs.append({"sheet": sheet_name, "lab_category": spec["lab_category"]})
    return configs


def normalize_protocol_text(text: str) -> str:
    cleaned = text.replace("\u3000", " ")
    cleaned = re.sub(r"[ \t]+", " ", cleaned)
    cleaned = re.sub(r"\n{2,}", "\n", cleaned)
    return cleaned.strip()


def read_protocol_file(path: Path) -> str:
    suffix = path.suffix.lower()
    if suffix == ".txt":
        return normalize_protocol_text(path.read_text(encoding="utf-8", errors="ignore"))
    if suffix == ".docx":
        with zipfile.ZipFile(path) as zf:
            xml = zf.read("word/document.xml").decode("utf-8", errors="ignore")
        text = re.sub(r"</w:p>", "\n", xml)
        text = re.sub(r"<[^>]+>", "", text)
        return normalize_protocol_text(text)
    if suffix == ".pdf" and PdfReader is not None:
        reader = PdfReader(str(path))
        text = "\n".join(page.extract_text() or "" for page in reader.pages)
        return normalize_protocol_text(text)
    return ""


def classify_endpoint_role(line: str) -> str:
    text = clean_text(line)
    if any(token in text for token in ["主要终点", "主要疗效终点", "primary endpoint", "primary efficacy"]):
        return "主要终点"
    if any(token in text for token in ["关键次要终点", "key secondary", "key secondary endpoint"]):
        return "关键次要终点"
    if any(token in text for token in ["次要终点", "次要疗效终点", "secondary endpoint", "secondary efficacy"]):
        return "次要终点"
    if any(token in text for token in ["探索性终点", "exploratory endpoint"]):
        return "探索性终点"
    return "未分类"


def classify_variable_type(endpoint_text: str) -> str:
    text = clean_text(endpoint_text).lower()
    if any(token in text for token in ["应答", "达标", "比例", "率", "proportion", "respond", "responder", "达到", "0/1", "75", "90", "50"]):
        return "binary"
    if any(token in text for token in ["变化", "change", "平均", "均值", "中位数", "评分", "score", "总分", "百分比变化", "reduction"]):
        return "continuous"
    return "需确认"


def metric_matches_protocol(metric: str, line: str) -> bool:
    text = clean_text(line)
    if metric in text:
        return True
    metric_upper = metric.upper()
    if metric_upper and metric_upper in text.upper():
        return True
    metric_alias_map = {
        "EASI": ["湿疹面积及严重程度指数"],
        "IGA": ["研究者整体评分", "vIGA-AD"],
        "BSA": ["受累体表面积"],
        "SCORAD": ["特应性皮炎评分"],
        "DLQI": ["皮肤病生活质量指数"],
        "CDLQI": ["儿童皮肤病生活质量指数"],
        "NRS": ["Itch NRS", "瘙痒数值评定量表", "Pruritus NRS"],
        "QSI-PROMIS睡眠相关影响8a": ["PROMIS", "睡眠相关影响8a"],
        "QSI-PROMIS睡眠困扰8b": ["PROMIS", "睡眠困扰8b"],
    }
    return any(alias.lower() in text.lower() for alias in metric_alias_map.get(metric, []))


def detect_protocol_endpoint_summary(protocol_files: list[Path], detected_efficacy: list[dict[str, Any]]) -> dict[str, Any]:
    summary = {"files": [str(path) for path in protocol_files], "endpoint_items": [], "selected_metrics": [], "response_rules": [], "visit_mentions": []}
    if not protocol_files:
        return summary

    lines: list[str] = []
    for path in protocol_files:
        text = read_protocol_file(path)
        if not text:
            continue
        lines.extend(line.strip() for line in text.splitlines() if clean_text(line))
        summary["visit_mentions"].extend(sorted(set(re.findall(r"\b(?:D\d+|W\d+|Week\s*\d+|筛选期|基线|治疗结束)\b", text, flags=re.I))))

    detected_metrics: dict[str, dict[str, Any]] = {}
    response_rules: dict[str, dict[str, Any]] = {}
    for line in lines:
        if not any(keyword in line for keyword in ["终点", "endpoint", "研究流程", "访视", "评估", "疗效"]):
            continue
        endpoint_role = classify_endpoint_role(line)
        for item in detected_efficacy:
            metric = item["metric"]
            if not metric_matches_protocol(metric, line):
                continue
            variable_type = classify_variable_type(line)
            existing = detected_metrics.get(metric)
            if existing is None or existing["endpoint_role"] == "未分类":
                detected_metrics[metric] = {
                    "metric": metric,
                    "endpoint_role": endpoint_role,
                    "variable_type": variable_type if variable_type != "需确认" else item.get("default_variable_type", "continuous"),
                    "source_excerpt": line[:220],
                }
            summary["endpoint_items"].append(
                {
                    "metric": metric,
                    "endpoint_role": endpoint_role,
                    "variable_type": variable_type,
                    "source_excerpt": line[:220],
                }
            )
        for rule in RESPONSE_RULE_CATALOG:
            if any(re.search(pattern, line, re.I) for pattern in rule["patterns"]):
                response_rules[rule["rule_key"]] = {
                    "rule_key": rule["rule_key"],
                    "metric": rule["metric"],
                    "label": rule["label"],
                    "endpoint_role": classify_endpoint_role(line),
                    "variable_type": "binary",
                    "source_excerpt": line[:220],
                }

    summary["selected_metrics"] = sorted(detected_metrics.values(), key=lambda item: efficacy_sort_key(item["metric"]))
    summary["response_rules"] = sorted(response_rules.values(), key=lambda item: (efficacy_sort_key(item["metric"]), item["label"]))
    return summary


def build_suggested_project_config() -> dict[str, Any]:
    return {
        "project_name": WORK.name,
        "html_title": "受试者Patient Profile",
        "target_centers": TARGET_CENTERS[:],
        "sheet_aliases": SHEET_ALIASES.copy(),
        "protocol_summary": deepcopy(PROTOCOL_SUMMARY),
        "efficacy_config": deepcopy(EFFICACY_CONFIG),
        "lab_config": deepcopy(LAB_CONFIG),
        "finding_sheet_specs": deepcopy(FINDING_SHEET_SPECS),
        "protocol_files": [str(path) for path in PROTOCOL_FILES],
        "reference_htmls": [str(path) for path in REF_HTMLS],
    }


def apply_project_config(config: dict[str, Any]) -> None:
    global TARGET_CENTERS, SHEET_ALIASES, EFFICACY_CONFIG, LAB_CONFIG, FINDING_SHEET_SPECS, HTML_TITLE, PROTOCOL_SUMMARY
    if config.get("target_centers"):
        TARGET_CENTERS = [clean_text(item) for item in config.get("target_centers", []) if clean_text(item)]
    if config.get("sheet_aliases"):
        SHEET_ALIASES = {alias: clean_text(name) for alias, name in config["sheet_aliases"].items()}
    if config.get("protocol_summary"):
        PROTOCOL_SUMMARY = deepcopy(config["protocol_summary"])
    if config.get("efficacy_config"):
        EFFICACY_CONFIG = deepcopy(config["efficacy_config"])
    if config.get("lab_config"):
        LAB_CONFIG = deepcopy(config["lab_config"])
    if config.get("finding_sheet_specs"):
        FINDING_SHEET_SPECS = deepcopy(config["finding_sheet_specs"])
    if config.get("html_title"):
        HTML_TITLE = clean_text(config["html_title"]) or HTML_TITLE


def infer_centers_from_listing(path: Path) -> list[str]:
    if not path.exists():
        return []
    aliases = detect_sheet_aliases(path)
    for alias in ["SUBJ", "SV", "DM"]:
        sheet_name = aliases.get(alias, "")
        if not sheet_name:
            continue
        try:
            sheet = read_listing_sheet(path, sheet_name)
        except Exception:
            continue
        centers = sorted(
            {
                standardize_center_name(row.get("研究中心") or row.get("中心名称") or row.get("中心"))
                for row in sheet["records"]
                if clean_text(row.get("研究中心") or row.get("中心名称") or row.get("中心"))
            }
        )
        if centers:
            return centers
    return []


def detect_finding_sheet_specs(path: Path, candidate_centers: list[str]) -> list[dict[str, str]]:
    if not path.exists():
        return []
    wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
    try:
        specs: list[dict[str, str]] = []
        center_fallback = candidate_centers[:]
        for sheet_name in wb.sheetnames:
            if not any(token in sheet_name for token in ["问题", "稽查", "自查", "finding", "Finding"]):
                continue
            try:
                sheet = read_generic_sheet(path, sheet_name)
            except Exception:
                continue
            headers = sheet["headers"]
            subject_col = ""
            for candidate in ["受试者筛选编号", "受试者编号", "受试者", "SUBJID", "Subject ID", "1"]:
                if candidate in headers:
                    subject_col = candidate
                    break
            center_guess = ""
            for center in center_fallback:
                if center and (center[:2] in sheet_name or center[:1] in sheet_name):
                    center_guess = center
                    break
            if not center_guess:
                for row in sheet["records"][:20]:
                    center_guess = standardize_center_name(row.get("中心") or row.get("研究中心") or row.get("中心名称"))
                    if center_guess:
                        break
            specs.append({"sheet_name": sheet_name, "center": center_guess or MISSING_TEXT, "subject_col": subject_col or MISSING_TEXT})
        return specs
    finally:
        wb.close()


def configure_runtime(args: argparse.Namespace) -> dict[str, Any]:
    global WORK, OUTPUT_DIR, LISTING_XLSX, FINDING_XLSX, PD_DEF_XLSX, SUBJECT_REPORT_XLS, REF_HTMLS
    global TARGET_CENTERS, PROTOCOL_FILES, FINDING_SHEET_SPECS, SHEET_ALIASES, EFFICACY_CONFIG, LAB_CONFIG, HTML_TITLE, PROTOCOL_SUMMARY

    WORK = Path(args.work_dir).resolve()
    OUTPUT_DIR = Path(args.output_dir).resolve() if args.output_dir else (WORK / "patient_profile_output")

    listing = Path(args.listing_xlsx).resolve() if args.listing_xlsx else pick_first_existing(find_candidate_files(WORK, "listing", (".xlsx", ".xlsm", ".xls")))
    finding = Path(args.finding_xlsx).resolve() if args.finding_xlsx else pick_first_existing(find_candidate_files(WORK, "finding", (".xlsx", ".xlsm")))
    pd_def = Path(args.pd_def_xlsx).resolve() if args.pd_def_xlsx else pick_first_existing(find_candidate_files(WORK, "pd_def", (".xlsx", ".xlsm", ".xls")))
    subject_report = Path(args.subject_report_xls).resolve() if args.subject_report_xls else pick_first_existing(find_candidate_files(WORK, "subject_report", (".xls", ".xlsx")))
    protocol_files = [Path(p).resolve() for p in args.protocol_file] if args.protocol_file else find_candidate_files(WORK, "protocol", (".pdf", ".doc", ".docx", ".txt"))
    ref_htmls = [Path(p).resolve() for p in args.ref_html] if args.ref_html else find_candidate_files(WORK, "reference_html", (".html",))

    LISTING_XLSX = listing or (WORK / "__missing_listing__.xlsx")
    FINDING_XLSX = finding or (WORK / "__missing_finding__.xlsx")
    PD_DEF_XLSX = pd_def
    SUBJECT_REPORT_XLS = subject_report or (WORK / "__missing_subject_report__.xls")
    PROTOCOL_FILES = protocol_files
    REF_HTMLS = ref_htmls

    SHEET_ALIASES = detect_sheet_aliases(LISTING_XLSX) if LISTING_XLSX.exists() else {alias: "" for alias in CORE_SHEET_CATALOG}
    detected_efficacy = detect_efficacy_config(LISTING_XLSX) if LISTING_XLSX.exists() else []
    PROTOCOL_SUMMARY = detect_protocol_endpoint_summary(PROTOCOL_FILES, detected_efficacy)
    EFFICACY_CONFIG = select_efficacy_config_by_protocol(detected_efficacy, PROTOCOL_SUMMARY)
    LAB_CONFIG = detect_lab_config(LISTING_XLSX) if LISTING_XLSX.exists() else []
    HTML_TITLE = "受试者Patient Profile"

    if args.centers:
        TARGET_CENTERS = [clean_text(item) for item in re.split(r"[;；]", args.centers) if clean_text(item)]
    else:
        TARGET_CENTERS = infer_centers_from_listing(LISTING_XLSX) if LISTING_XLSX.exists() else []

    FINDING_SHEET_SPECS = detect_finding_sheet_specs(FINDING_XLSX, TARGET_CENTERS) if FINDING_XLSX.exists() else []
    if args.config_json:
        config = json.loads(Path(args.config_json).read_text(encoding="utf-8"))
        apply_project_config(config)
    return {
        "work_dir": str(WORK),
        "listing_xlsx": str(LISTING_XLSX) if LISTING_XLSX else "",
        "finding_xlsx": str(FINDING_XLSX) if FINDING_XLSX else "",
        "pd_def_xlsx": str(PD_DEF_XLSX) if PD_DEF_XLSX else "",
        "subject_report_xls": str(SUBJECT_REPORT_XLS) if SUBJECT_REPORT_XLS else "",
        "protocol_files": [str(path) for path in PROTOCOL_FILES],
        "ref_htmls": [str(path) for path in REF_HTMLS],
        "target_centers": TARGET_CENTERS[:],
        "centers_from_user": bool(args.centers),
        "finding_sheet_specs": FINDING_SHEET_SPECS[:],
        "sheet_aliases": SHEET_ALIASES.copy(),
        "protocol_summary": deepcopy(PROTOCOL_SUMMARY),
        "efficacy_config": deepcopy(EFFICACY_CONFIG),
        "lab_config": deepcopy(LAB_CONFIG),
        "output_dir": str(OUTPUT_DIR),
    }

FIELD_ALIASES = {
    "研究中心": "中心",
    "中心名称": "中心",
    "site": "中心",
    "sitenm": "中心",
    "siteid": "中心编号",
    "中心编号": "中心编号",
    "受试者": "受试者编号",
    "受试者编号": "受试者编号",
    "subjid": "受试者编号",
    "subjectid": "受试者编号",
    "subject id": "受试者编号",
    "subjectidnumber": "受试者编号",
    "受试者状态": "受试者状态",
    "访视名称": "访视名称",
    "visit": "访视名称",
    "访视oid": "访视编号",
    "vistoid": "访视编号",
    "访视号": "访视序号",
    "评估日期": "日期",
    "采样日期": "日期",
    "检查日期": "日期",
    "访视日期": "日期",
    "入组日期": "日期",
    "知情同意日期": "日期",
    "结果值": "结果值",
    "检查结果": "结果值",
    "原始结果": "结果值",
    "lbtest": "检验项目",
    "检查项目": "检验项目",
    "单位": "单位",
    "下限": "正常值下限",
    "上限": "正常值上限",
    "标准值": "正常范围原始文本",
    "发现日期": "发现日期",
    "分类": "问题分类",
    "一类分级": "问题严重程度",
    "自查问题描述": "原始问题描述",
    "问题状态（关闭/未关闭/无法整改，解释）": "问题状态",
    "问题状态（关闭/未关闭/无法整改，解释），cra审核后回复": "问题状态",
}

DISPLAY_FIELD_KEYWORDS = (
    "中心",
    "受试者",
    "访视",
    "日期",
    "结果",
    "评分",
    "检验",
    "问题",
    "严重程度",
    "分类",
    "知情",
    "随机",
    "年龄",
    "性别",
    "下限",
    "上限",
    "备注",
)

DEFAULT_EFFICACY_CONFIG = [
    {
        "sheet": "EASI--湿疹面积及严重程度指数评分（EASI）",
        "sheet_patterns": [r"\bEASI\b", r"湿疹面积及严重程度指数"],
        "metric": "EASI",
        "group": "疗效",
        "value_col": "EASI得分（4个部位的评分总和）",
        "value_candidates": ["EASI得分（4个部位的评分总和）", "EASI得分", "总分", "结果值"],
        "date_col": "评估日期",
        "date_candidates": ["评估日期", "日期", "访视日期"],
        "detail_fields": ["部位", "该部位BSA结果（系统生成）", "皮损累及面积评分", "红斑", "水肿/丘疹", "表皮脱落", "苔藓样变"],
        "aggregate": "total_with_details",
    },
    {
        "sheet": "BSA--受累体表面积（BSA）",
        "sheet_patterns": [r"\bBSA\b", r"受累体表面积"],
        "metric": "BSA",
        "group": "疗效",
        "value_col": "BSA结果（4个部位的评分总和）",
        "value_candidates": ["BSA结果（4个部位的评分总和）", "BSA结果", "总分", "结果值"],
        "date_col": "评估日期",
        "date_candidates": ["评估日期", "日期", "访视日期"],
        "detail_fields": ["部位", "BSA结果"],
        "aggregate": "total_with_details",
    },
    {
        "sheet": "SR--特应性皮炎评分（SCORAD）-最后分值",
        "sheet_patterns": [r"SCORAD", r"最后分值"],
        "metric": "SCORAD",
        "group": "疗效",
        "value_col": "最后分值",
        "value_candidates": ["最后分值", "总分", "结果值"],
        "date_col": "评估日期",
        "date_candidates": ["评估日期", "日期", "访视日期"],
        "detail_fields": [],
        "aggregate": "single",
    },
    {
        "sheet": "IGA--研究者整体评分（IGA）",
        "sheet_patterns": [r"\bIGA\b", r"研究者整体评分"],
        "metric": "IGA",
        "group": "疗效",
        "value_col": "IGA得分",
        "value_candidates": ["IGA得分", "结果值", "评分"],
        "date_col": "评估日期",
        "date_candidates": ["评估日期", "日期", "访视日期"],
        "detail_fields": ["是否达到IGA-TS？（系统生成）"],
        "aggregate": "single",
    },
    {
        "sheet": "DLQI--皮肤病生活质量指数（DLQI）",
        "sheet_patterns": [r"\bDLQI\b", r"皮肤病生活质量指数"],
        "metric": "DLQI",
        "group": "PRO",
        "value_col": "合计评分（系统生成）",
        "value_candidates": ["合计评分（系统生成）", "合计评分", "总分", "结果值"],
        "date_col": "评估日期",
        "date_candidates": ["评估日期", "日期", "访视日期"],
        "detail_fields": [],
        "aggregate": "single",
    },
    {
        "sheet": "CDLQI--儿童皮肤病生活质量指数（CDLQI）",
        "sheet_patterns": [r"\bCDLQI\b", r"儿童皮肤病生活质量指数"],
        "metric": "CDLQI",
        "group": "PRO",
        "value_col": "合计评分（系统生成）",
        "value_candidates": ["合计评分（系统生成）", "合计评分", "总分", "结果值"],
        "date_col": "评估日期",
        "date_candidates": ["评估日期", "日期", "访视日期"],
        "detail_fields": [],
        "aggregate": "single",
    },
    {
        "sheet": "NRS--瘙痒数值评定量表（Itch NRS）-平均值",
        "sheet_patterns": [r"Itch NRS", r"\bNRS\b", r"瘙痒数值评定量表"],
        "metric": "NRS",
        "group": "PRO",
        "value_col": "平均瘙痒程度",
        "value_candidates": ["平均瘙痒程度", "平均值", "总分", "结果值"],
        "date_col": "评估日期",
        "date_candidates": ["评估日期", "日期", "访视日期"],
        "detail_fields": ["开始日期", "结束日期"],
        "aggregate": "single",
    },
    {
        "sheet": "QSI--PROMIS简表-睡眠相关影响8a-平均值",
        "sheet_patterns": [r"PROMIS.*8a", r"睡眠相关影响8a"],
        "metric": "QSI-PROMIS睡眠相关影响8a",
        "group": "PRO",
        "value_col": "平均评分",
        "value_candidates": ["平均评分", "平均值", "总分", "结果值"],
        "date_col": "评估日期",
        "date_candidates": ["评估日期", "日期", "访视日期"],
        "detail_fields": ["开始日期", "结束日期"],
        "aggregate": "single",
    },
    {
        "sheet": "QSD--PROMIS简表-睡眠困扰8b-平均值",
        "sheet_patterns": [r"PROMIS.*8b", r"睡眠困扰8b"],
        "metric": "QSI-PROMIS睡眠困扰8b",
        "group": "PRO",
        "value_col": "平均评分",
        "value_candidates": ["平均评分", "平均值", "总分", "结果值"],
        "date_col": "评估日期",
        "date_candidates": ["评估日期", "日期", "访视日期"],
        "detail_fields": ["开始日期", "结束日期"],
        "aggregate": "single",
    },
]

RESPONSE_RULE_CATALOG = [
    {
        "rule_key": "EASI75",
        "metric": "EASI",
        "label": "EASI-75",
        "patterns": [r"EASI[\s\-]?75", r"EASI.*75%"],
        "variable_type": "binary",
    },
    {
        "rule_key": "EASI90",
        "metric": "EASI",
        "label": "EASI-90",
        "patterns": [r"EASI[\s\-]?90", r"EASI.*90%"],
        "variable_type": "binary",
    },
    {
        "rule_key": "IGA_TS",
        "metric": "IGA",
        "label": "IGA-TS",
        "patterns": [r"IGA[\s\-]?TS", r"IGA.*0.?1.*改善.?2", r"vIGA.?AD 0.?1"],
        "variable_type": "binary",
    },
    {
        "rule_key": "IGA_01",
        "metric": "IGA",
        "label": "IGA 0/1分",
        "patterns": [r"IGA.*0.?1", r"IGA 0/1", r"vIGA.?AD 0.?1"],
        "variable_type": "binary",
    },
    {
        "rule_key": "NRS4",
        "metric": "NRS",
        "label": "Itch NRS改善≥4分",
        "patterns": [r"NRS.*改善.?≥?4", r"itch nrs.*4", r"pruritus nrs.*4"],
        "variable_type": "binary",
    },
    {
        "rule_key": "PROMIS8A_6",
        "metric": "QSI-PROMIS睡眠相关影响8a",
        "label": "PROMIS-8a临床意义改善",
        "patterns": [r"PROMIS.*8a.*改善.?≥?6", r"睡眠相关影响8a.*改善.?≥?6", r"PROMIS.*8a"],
        "variable_type": "binary",
    },
    {
        "rule_key": "PROMIS8B_6",
        "metric": "QSI-PROMIS睡眠困扰8b",
        "label": "PROMIS-8b临床意义改善",
        "patterns": [r"PROMIS.*8b.*改善.?≥?6", r"睡眠困扰8b.*改善.?≥?6", r"PROMIS.*8b"],
        "variable_type": "binary",
    },
]

ENDPOINT_METRIC_CATALOG = [
    {
        "metric": spec["metric"],
        "aliases": [spec["metric"], *spec.get("sheet_patterns", [])],
        "default_variable_type": "continuous",
        "config": spec,
    }
    for spec in DEFAULT_EFFICACY_CONFIG
]

DEFAULT_LAB_CONFIG = [
    {"sheet": "LBHEMA--实验室检查-血常规", "sheet_patterns": [r"^LBHEMA--", r"血常规"], "lab_category": "血常规"},
    {"sheet": "LBCHEM--实验室检查-血生化", "sheet_patterns": [r"^LBCHEM--", r"血生化"], "lab_category": "血生化"},
    {"sheet": "LBURIN--实验室检查-尿常规", "sheet_patterns": [r"^LBURIN--", r"尿常规"], "lab_category": "尿常规"},
]

EFFICACY_CONFIG = deepcopy(DEFAULT_EFFICACY_CONFIG)
LAB_CONFIG = deepcopy(DEFAULT_LAB_CONFIG)
SHEET_ALIASES = {alias: "" for alias in CORE_SHEET_CATALOG}
HTML_TITLE = "受试者Patient Profile"

LAB_CATEGORY_ORDER = {
    "血常规": 1,
    "肝功能相关血液指标": 2,
    "肾功能相关血液指标": 3,
    "血脂相关指标": 4,
    "其他血生化指标": 5,
    "心电图": 6,
    "尿常规": 7,
}

LAB_METRIC_ORDER = {
    "WBC": 1, "中性粒细胞计数（NEU#）": 2, "淋巴细胞计数（LYM#）": 3, "RBC": 4, "HGB": 5, "PLT": 6,
    "嗜酸性粒细胞计数（EO#）": 7, "单核细胞计数（MO#）": 8, "嗜碱性粒细胞计数（BA#）": 9, "HCT": 10,
    "MCV": 11, "MCH": 12, "MCHC": 13, "MPV": 14, "RDW-CV": 15, "RDW-SD": 16,
    "ALT": 101, "AST": 102, "ALP": 103, "GGT": 104, "TBIL": 105, "DBIL": 106, "IBIL": 107, "ALB": 108, "TP": 109,
    "CREA": 201, "尿素": 202, "UA": 203, "eGFR": 204,
    "TG": 301, "TC": 302, "HDL": 303, "LDL": 304,
    "GLU": 401, "CK": 402, "LDH": 403, "QTc间期": 500,
}

EFFICACY_METRIC_ORDER = {
    "IGA": 1,
    "EASI": 2,
    "BSA": 3,
    "DLQI": 4,
    "CDLQI": 5,
    "SCORAD": 6,
    "NRS": 7,
}

VITAL_METRIC_ORDER = {
    "体温": 1,
    "心率": 2,
    "收缩压": 3,
    "舒张压": 4,
}

def clean_text(value: Any) -> str:
    if value is None:
        return ""
    if isinstance(value, datetime):
        return value.strftime("%Y-%m-%d %H:%M:%S")
    if isinstance(value, date):
        return value.strftime("%Y-%m-%d")
    text = str(value).replace("\u3000", " ").replace("\r", " ").replace("\n", " ").strip()
    return re.sub(r"\s+", " ", text)


def parse_date(value: Any) -> str:
    if value is None or value == "":
        return ""
    if isinstance(value, datetime):
        return value.strftime("%Y-%m-%d")
    if isinstance(value, date):
        return value.strftime("%Y-%m-%d")
    text = clean_text(value)
    if not text:
        return ""
    text = text.replace("/", "-").replace(".", "-")
    if re.fullmatch(r"\d{4}-\d{1,2}-\d{1,2}", text):
        try:
            return datetime.strptime(text, "%Y-%m-%d").strftime("%Y-%m-%d")
        except ValueError:
            parts = text.split("-")
            return f"{int(parts[0]):04d}-{int(parts[1]):02d}-{int(parts[2]):02d}"
    return text


def parse_datetime_text(value: Any) -> str:
    if value is None or value == "":
        return ""
    if isinstance(value, datetime):
        return value.strftime("%Y-%m-%d %H:%M:%S")
    if isinstance(value, date):
        return value.strftime("%Y-%m-%d")
    return clean_text(value)


def to_float(value: Any) -> float | None:
    text = clean_text(value)
    if not text or text.upper() in {"NA", "N/A", "ND"}:
        return None
    text = text.replace(",", "")
    try:
        return float(text)
    except ValueError:
        return None


def parse_iga_score(value: Any) -> float | None:
    text = clean_text(value)
    if not text:
        return None
    match = re.search(r"(\d+(?:\.\d+)?)\s*分", text)
    if match:
        return float(match.group(1))
    return to_float(text)


def efficacy_sort_key(metric_name: str) -> tuple[int, str]:
    name = clean_text(metric_name)
    return (EFFICACY_METRIC_ORDER.get(name, 999), name)


def vital_sort_key(metric_name: str) -> tuple[int, str]:
    name = clean_text(metric_name)
    return (VITAL_METRIC_ORDER.get(name, 999), name)


def format_number(value: float | None, digits: int = 2) -> str:
    if value is None:
        return ""
    return f"{value:.{digits}f}"


def compute_mean(values: list[float]) -> float | None:
    return sum(values) / len(values) if values else None


def compute_median(values: list[float]) -> float | None:
    if not values:
        return None
    ordered = sorted(values)
    mid = len(ordered) // 2
    if len(ordered) % 2:
        return ordered[mid]
    return (ordered[mid - 1] + ordered[mid]) / 2


def compute_sample_sd(values: list[float]) -> float | None:
    if len(values) < 2:
        return 0.0 if values else None
    mean_value = compute_mean(values)
    assert mean_value is not None
    variance = sum((value - mean_value) ** 2 for value in values) / (len(values) - 1)
    return math.sqrt(variance)


def aggregate_continuous_by_visit(
    rows: list[dict[str, Any]],
    center: str,
    metric_key: str,
    metric_value: str,
    value_key: str = "数值结果",
) -> list[dict[str, Any]]:
    filtered = [
        row for row in rows
        if clean_text(row.get("中心")) == clean_text(center)
        and clean_text(row.get(metric_key)) == clean_text(metric_value)
        and isinstance(row.get(value_key), (int, float))
        and not is_unplanned_visit(row.get("访视编号"), row.get("访视名称"))
    ]
    grouped: defaultdict[tuple[str, str], list[float]] = defaultdict(list)
    source_row: dict[tuple[str, str], dict[str, Any]] = {}
    date_by_group: defaultdict[tuple[str, str], list[str]] = defaultdict(list)
    for row in filtered:
        key = (clean_text(row.get("访视编号")), clean_text(row.get("访视名称")))
        grouped[key].append(float(row[value_key]))
        source_row.setdefault(key, row)
        if clean_text(row.get("日期")):
            date_by_group[key].append(clean_text(row.get("日期")))
    summary_rows: list[dict[str, Any]] = []
    for key, values in grouped.items():
        row = source_row[key]
        unique_dates = sorted(set(date_by_group.get(key, [])))
        if not unique_dates:
            date_label = MISSING_TEXT
        elif len(unique_dates) == 1:
            date_label = unique_dates[0]
        else:
            date_label = f"{unique_dates[0]} ~ {unique_dates[-1]}"
        summary_rows.append(
            {
                "中心": center,
                "指标名称": clean_text(row.get(metric_key)),
                "访视编号": key[0],
                "访视名称": key[1],
                "日期": date_label,
                "例次": len(values),
                "均值": format_number(compute_mean(values)),
                "中位数": format_number(compute_median(values)),
                "标准差": format_number(compute_sample_sd(values)),
                "最小值": format_number(min(values)),
                "最大值": format_number(max(values)),
                "均值数值": compute_mean(values),
                "标准差数值": compute_sample_sd(values),
            }
        )
    return sorted(summary_rows, key=lambda row: (row["日期"] == MISSING_TEXT, row["日期"], row["访视编号"]))


def aggregate_binary_by_visit(
    rows: list[dict[str, Any]],
    center: str,
    metric_key: str,
    metric_value: str,
    response_name: str,
    matcher: Any,
) -> list[dict[str, Any]]:
    filtered = [
        row for row in rows
        if clean_text(row.get("中心")) == clean_text(center)
        and clean_text(row.get(metric_key)) == clean_text(metric_value)
        and clean_text(row.get("关键应答判定")) not in {"", MISSING_TEXT}
        and not is_unplanned_visit(row.get("访视编号"), row.get("访视名称"))
    ]
    grouped: defaultdict[tuple[str, str], list[dict[str, Any]]] = defaultdict(list)
    date_by_group: defaultdict[tuple[str, str], list[str]] = defaultdict(list)
    for row in filtered:
        key = (clean_text(row.get("访视编号")), clean_text(row.get("访视名称")))
        grouped[key].append(row)
        if clean_text(row.get("日期")):
            date_by_group[key].append(clean_text(row.get("日期")))
    summary_rows: list[dict[str, Any]] = []
    for key, visit_rows in grouped.items():
        total = len(visit_rows)
        responders = sum(1 for row in visit_rows if matcher(row))
        proportion = (responders / total * 100) if total else None
        unique_dates = sorted(set(date_by_group.get(key, [])))
        if not unique_dates:
            date_label = MISSING_TEXT
        elif len(unique_dates) == 1:
            date_label = unique_dates[0]
        else:
            date_label = f"{unique_dates[0]} ~ {unique_dates[-1]}"
        summary_rows.append(
            {
                "中心": center,
                "指标名称": clean_text(metric_value),
                "应答名称": response_name,
                "访视编号": key[0],
                "访视名称": key[1],
                "日期": date_label,
                "应答例数": responders,
                "例次": total,
                "应答比例": format_number(proportion) + "%" if proportion is not None else "",
                "应答比例数值": proportion,
            }
        )
    return sorted(summary_rows, key=lambda row: (row["日期"] == MISSING_TEXT, row["日期"], row["访视编号"]))


def aggregate_continuous_by_visit_group(
    rows: list[dict[str, Any]],
    center: str,
    metric_key: str,
    metric_value: str,
    subject_group_map: dict[str, str],
    denominator_by_group: dict[str, int],
    value_key: str = "数值结果",
) -> list[dict[str, Any]]:
    filtered = [
        row for row in rows
        if clean_text(row.get("中心")) == clean_text(center)
        and clean_text(row.get(metric_key)) == clean_text(metric_value)
        and isinstance(row.get(value_key), (int, float))
        and not is_excluded_center_summary_visit(row.get("访视编号"), row.get("访视名称"))
    ]
    grouped: defaultdict[tuple[str, str], dict[str, list[float]]] = defaultdict(lambda: {"试验组": [], "对照组": []})
    date_by_group: defaultdict[tuple[str, str], list[str]] = defaultdict(list)
    for row in filtered:
        group = summary_group_label(subject_group_map.get(clean_text(row.get("受试者键")), ""))
        if not group:
            continue
        key = (clean_text(row.get("访视编号")), clean_text(row.get("访视名称")))
        grouped[key][group].append(float(row[value_key]))
        if clean_text(row.get("日期")):
            date_by_group[key].append(clean_text(row.get("日期")))
    summary_rows: list[dict[str, Any]] = []
    for key, values_by_group in grouped.items():
        unique_dates = sorted(set(date_by_group.get(key, [])))
        if not unique_dates:
            date_label = MISSING_TEXT
        elif len(unique_dates) == 1:
            date_label = unique_dates[0]
        else:
            date_label = f"{unique_dates[0]} ~ {unique_dates[-1]}"
        row: dict[str, Any] = {
            "中心": center,
            "指标名称": clean_text(metric_value),
            "访视编号": key[0],
            "访视名称": key[1],
            "日期": date_label,
        }
        for group in ["试验组", "对照组"]:
            values = values_by_group[group]
            observed = len(values)
            missing = max(denominator_by_group.get(group, 0) - observed, 0)
            prefix = f"{group}"
            row[f"{prefix}例次（缺失）"] = f"{observed} ({missing})"
            row[f"{prefix}均值"] = format_number(compute_mean(values)) if values else ""
            row[f"{prefix}中位数"] = format_number(compute_median(values)) if values else ""
            row[f"{prefix}标准差"] = format_number(compute_sample_sd(values)) if values else ""
            row[f"{prefix}最小值"] = format_number(min(values)) if values else ""
            row[f"{prefix}最大值"] = format_number(max(values)) if values else ""
            row[f"{prefix}均值数值"] = compute_mean(values)
            row[f"{prefix}标准差数值"] = compute_sample_sd(values)
        summary_rows.append(row)
    return sorted(summary_rows, key=lambda row: (row["日期"] == MISSING_TEXT, row["日期"], row["访视编号"]))


def aggregate_binary_by_visit_group(
    rows: list[dict[str, Any]],
    center: str,
    metric_key: str,
    metric_value: str,
    response_name: str,
    matcher: Any,
    subject_group_map: dict[str, str],
    denominator_by_group: dict[str, int],
) -> list[dict[str, Any]]:
    filtered = [
        row for row in rows
        if clean_text(row.get("中心")) == clean_text(center)
        and clean_text(row.get(metric_key)) == clean_text(metric_value)
        and clean_text(row.get("关键应答判定")) not in {"", MISSING_TEXT}
        and not is_excluded_center_summary_visit(row.get("访视编号"), row.get("访视名称"))
    ]
    grouped: defaultdict[tuple[str, str], dict[str, list[dict[str, Any]]]] = defaultdict(lambda: {"试验组": [], "对照组": []})
    date_by_group: defaultdict[tuple[str, str], list[str]] = defaultdict(list)
    for row in filtered:
        group = summary_group_label(subject_group_map.get(clean_text(row.get("受试者键")), ""))
        if not group:
            continue
        key = (clean_text(row.get("访视编号")), clean_text(row.get("访视名称")))
        grouped[key][group].append(row)
        if clean_text(row.get("日期")):
            date_by_group[key].append(clean_text(row.get("日期")))
    summary_rows: list[dict[str, Any]] = []
    for key, rows_by_group in grouped.items():
        unique_dates = sorted(set(date_by_group.get(key, [])))
        if not unique_dates:
            date_label = MISSING_TEXT
        elif len(unique_dates) == 1:
            date_label = unique_dates[0]
        else:
            date_label = f"{unique_dates[0]} ~ {unique_dates[-1]}"
        row: dict[str, Any] = {
            "中心": center,
            "指标名称": clean_text(metric_value),
            "应答名称": response_name,
            "访视编号": key[0],
            "访视名称": key[1],
            "日期": date_label,
        }
        for group in ["试验组", "对照组"]:
            visit_rows = rows_by_group[group]
            observed = len(visit_rows)
            missing = max(denominator_by_group.get(group, 0) - observed, 0)
            responders = sum(1 for item in visit_rows if matcher(item))
            proportion = responders / observed * 100 if observed else None
            prefix = f"{group}"
            row[f"{prefix}例次（缺失）"] = f"{observed} ({missing})"
            row[f"{prefix}应答例数"] = responders
            row[f"{prefix}应答比例"] = format_number(proportion) + "%" if proportion is not None else ""
            row[f"{prefix}应答比例数值"] = proportion
        summary_rows.append(row)
    return sorted(summary_rows, key=lambda row: (row["日期"] == MISSING_TEXT, row["日期"], row["访视编号"]))


def response_label_matches(response_label: Any, response_name: str) -> bool:
    label = clean_text(response_label)
    if not label or label == MISSING_TEXT:
        return False
    if response_name == "EASI-75":
        return "EASI-75" in label
    if response_name == "EASI-90":
        return "EASI-90" in label
    if response_name == "IGA-TS":
        return "IGA-TS" in label
    if response_name == "IGA 0/1分":
        return "IGA 0/1分" in label
    if response_name == "Itch NRS改善≥4分":
        return label == "Itch NRS改善≥4分"
    if response_name in {"PROMIS-8a临床意义改善", "PROMIS-8b临床意义改善"}:
        return label == response_name
    return False


def build_center_profile_summary(subject_rows: list[dict[str, Any]], center: str) -> dict[str, Any]:
    filtered = [row for row in subject_rows if clean_text(row.get("中心")) == clean_text(center)]
    ages = [to_float(row.get("年龄")) for row in filtered if to_float(row.get("年龄")) is not None]
    screened_fail = [row for row in filtered if clean_text(row.get("受试者状态")) == "筛选失败"]
    treated = [row for row in filtered if clean_text(row.get("受试者状态")) != "筛选失败"]
    ended = [
        row for row in treated
        if row.get("是否完成研究") == "是" or clean_text(row.get("受试者状态")) in {"完成试验", "提前终止", "提前退出"}
    ]
    treated_ages = [to_float(row.get("年龄")) for row in treated if to_float(row.get("年龄")) is not None]
    sex_counter = Counter(clean_text(row.get("性别")) or MISSING_TEXT for row in filtered)
    age_group_counter = Counter(clean_text(row.get("年龄组")) or MISSING_TEXT for row in filtered)
    treated_sex_counter = Counter(clean_text(row.get("性别")) or MISSING_TEXT for row in treated)
    treated_age_group_counter = Counter(clean_text(row.get("年龄组")) or MISSING_TEXT for row in treated)
    total = len(filtered)
    treated_total = len(treated)

    def ratio_text(counter: Counter[str], denominator: int) -> str:
        parts = []
        for key, count in counter.items():
            pct = count / denominator * 100 if denominator else 0
            parts.append(f"{key} {count}/{denominator} ({pct:.1f}%)")
        return "；".join(parts)

    return {
        "中心": center,
        "受试者总数": total,
        "筛选失败例数": len(screened_fail),
        "进入治疗例数": len(treated),
        "研究结束例数": len(ended),
        "年龄均值": format_number(compute_mean([value for value in ages if value is not None])),
        "性别分布": ratio_text(sex_counter, total),
        "年龄组分布": ratio_text(age_group_counter, total),
        "剔除筛败后受试者数": treated_total,
        "年龄均值_剔除筛败后": format_number(compute_mean([value for value in treated_ages if value is not None])),
        "性别分布_剔除筛败后": ratio_text(treated_sex_counter, treated_total),
        "年龄组分布_剔除筛败后": ratio_text(treated_age_group_counter, treated_total),
    }


def is_unplanned_visit(visit_id: Any, visit_name: Any) -> bool:
    visit_code = clean_text(visit_id).upper()
    visit_label = clean_text(visit_name)
    return visit_code == "UNS" or "计划外" in visit_label or "非计划" in visit_label


def is_excluded_center_summary_visit(visit_id: Any, visit_name: Any) -> bool:
    visit_code = clean_text(visit_id).upper()
    visit_label = clean_text(visit_name)
    return is_unplanned_visit(visit_id, visit_name) or visit_code == "EOT" or "治疗结束" in visit_label


def summary_group_label(group_name: Any) -> str:
    text = clean_text(group_name)
    if "芦可替尼" in text:
        return "试验组"
    if "安慰剂" in text:
        return "对照组"
    return ""


def should_assign_response_label(visit_id: str, visit_name: str, metric: str) -> bool:
    metric_name = clean_text(metric)
    if not selected_response_rules_for_metric(metric_name):
        return False
    visit_id_text = clean_text(visit_id).upper()
    visit_name_text = clean_text(visit_name)
    if visit_id_text == "SCR" or "筛选期" in visit_name_text:
        return False
    return True


def efficacy_table_headers(metric: str) -> list[str]:
    headers = ["访视编号", "访视名称", "日期", "原始值", "较基线变化", "较基线百分比变化"]
    if selected_response_rules_for_metric(clean_text(metric)):
        headers.append("关键应答判定")
    headers.extend(["问题标记", "关联Finding编号", "数据来源sheet"])
    return headers


def make_unique_headers(headers: list[str]) -> list[str]:
    seen = Counter()
    out: list[str] = []
    for idx, raw in enumerate(headers):
        name = clean_text(raw) or f"col_{idx+1}"
        seen[name] += 1
        if seen[name] > 1:
            name = f"{name}_{seen[name]}"
        out.append(name)
    return out


def unique(values: list[Any]) -> list[Any]:
    seen = []
    for value in values:
        if value not in seen:
            seen.append(value)
    return seen


def normalize_study_group(raw_group: Any) -> str:
    text = clean_text(raw_group)
    if "安慰剂" in text:
        return "安慰剂组"
    if "磷酸芦可替尼" in text:
        return "磷酸芦可替尼组"
    return text or MISSING_TEXT


def infer_age_group_from_report(raw_value: Any) -> str:
    text = clean_text(raw_value)
    if any(token in text for token in ["12周岁≤年龄<18", "12<=年龄<18", "12<=年龄<18周岁"]):
        return "青少年"
    if any(token in text for token in ["年龄≥18周岁", ">=18周岁"]):
        return "成人"
    return MISSING_TEXT


def subject_should_show_efficacy(subject: dict[str, Any]) -> bool:
    return clean_text(subject.get("受试者状态")) != "筛选失败"


def subject_should_show_labs(subject: dict[str, Any]) -> bool:
    return clean_text(subject.get("受试者状态")) != "筛选失败"


def subject_status_tags(subject: dict[str, Any]) -> list[str]:
    tags = [clean_text(subject.get("随机组别"))]
    if clean_text(subject.get("年龄组")) == "青少年":
        tags.append("青少年受试者")
    if clean_text(subject.get("受试者状态")) == "筛选失败":
        tags.append("筛选失败")
    if clean_text(subject.get("是否存在AE/SAE/SUSAR")) not in {"", "否", MISSING_TEXT}:
        tags.append(clean_text(subject.get("是否存在AE/SAE/SUSAR")))
    if clean_text(subject.get("是否存在核查Finding")) == "是":
        tags.append("有Finding")
    return [tag for tag in tags if tag]


def should_include_lab_test(test_name: Any, lab_category: str) -> bool:
    text = clean_text(test_name)
    if "BUN" in text or "尿素氮" in text:
        return False
    return True


def infer_lab_group(metric_name: str, lab_category: str) -> str:
    if lab_category == "血常规":
        return "血常规"
    if lab_category == "尿常规":
        return "尿常规"
    if metric_name in {"ALT", "AST", "ALP", "GGT", "TBIL", "DBIL", "IBIL", "ALB", "TP"}:
        return "肝功能相关血液指标"
    if metric_name in {"CREA", "尿素", "UA", "eGFR"}:
        return "肾功能相关血液指标"
    if metric_name in {"TG", "TC", "HDL", "LDL"}:
        return "血脂相关指标"
    if metric_name == "QTc间期":
        return "心电图"
    return "其他血生化指标"


def lab_sort_key(row: dict[str, Any]) -> tuple[Any, ...]:
    group = row.get("实验室显示分组") or infer_lab_group(clean_text(row.get("检验项目标准化名称")), clean_text(row.get("实验室类别")))
    metric = clean_text(row.get("检验项目标准化名称"))
    return (
        clean_text(row.get("中心")),
        clean_text(row.get("受试者编号")),
        LAB_CATEGORY_ORDER.get(group, 99),
        LAB_METRIC_ORDER.get(metric, 999),
        metric,
        clean_text(row.get("日期")),
        clean_text(row.get("访视编号")),
    )


def vital_sort_row_key(row: dict[str, Any]) -> tuple[Any, ...]:
    metric = clean_text(row.get("生命体征指标"))
    return (
        clean_text(row.get("中心")),
        clean_text(row.get("受试者编号")),
        VITAL_METRIC_ORDER.get(metric, 999),
        metric,
        clean_text(row.get("日期")),
        clean_text(row.get("访视编号")),
    )


def vital_reference_range(metric_name: str) -> tuple[float | None, float | None]:
    metric = clean_text(metric_name)
    if metric == "体温":
        return (None, 37.2)
    if metric == "心率":
        return (60.0, 100.0)
    if metric == "收缩压":
        return (None, 140.0)
    if metric == "舒张压":
        return (None, 90.0)
    return (None, None)


def finding_matches_vital(finding: dict[str, Any], vital_row: dict[str, Any]) -> bool:
    if clean_text(finding.get("问题分类")) not in {"实验室检查", "AE/SAE/SUSAR"}:
        return False
    text = " ".join(
        clean_text(finding.get(k))
        for k in ["原始问题描述", "涉及访视", "涉及数据点或文件", "问题分类"]
    )
    metric_tokens = [
        clean_text(vital_row.get("生命体征指标")),
        clean_text(vital_row.get("原始指标名称")),
        "生命体征",
    ]
    visit_tokens = [clean_text(vital_row.get("访视编号")), clean_text(vital_row.get("访视名称"))]
    metric_match = any(token and token in text for token in metric_tokens if len(token) >= 2)
    visit_match = any(token and token in text for token in visit_tokens)
    return metric_match and visit_match


def finding_matches_efficacy(finding: dict[str, Any], efficacy_row: dict[str, Any]) -> bool:
    if clean_text(finding.get("问题分类")) not in {"疗效评价", "量表一致性", "eCOA", "日记卡"}:
        return False
    text = " ".join(
        clean_text(finding.get(k))
        for k in ["原始问题描述", "涉及访视", "涉及数据点或文件", "问题分类"]
    )
    visit_tokens = [clean_text(efficacy_row.get("访视编号")), clean_text(efficacy_row.get("访视名称"))]
    metric_tokens = [clean_text(efficacy_row.get("指标名称"))]
    visit_match = any(token and token in text for token in visit_tokens)
    metric_match = any(token and token in text for token in metric_tokens)
    return metric_match or (visit_match and any(word in text for word in ["量表", "评分", "eCOA", "日记", "疗效"]))


def finding_matches_lab(finding: dict[str, Any], lab_row: dict[str, Any]) -> bool:
    if clean_text(finding.get("问题分类")) not in {"实验室检查", "AE/SAE/SUSAR"}:
        return False
    text = " ".join(
        clean_text(finding.get(k))
        for k in ["原始问题描述", "涉及访视", "涉及数据点或文件", "问题分类"]
    )
    metric_tokens = [
        clean_text(lab_row.get("检验项目标准化名称")),
        clean_text(lab_row.get("原始检验项目名称")),
        clean_text(lab_row.get("实验室类别")),
    ]
    visit_tokens = [clean_text(lab_row.get("访视编号")), clean_text(lab_row.get("访视名称"))]
    metric_match = any(token and token in text for token in metric_tokens if len(token) >= 2)
    visit_match = any(token and token in text for token in visit_tokens)
    return metric_match and visit_match


def normalize_field_name(field_name: str) -> str:
    raw = clean_text(field_name)
    if not raw:
        return "需人工确认"
    key = re.sub(r"[\s_\-（）()]+", "", raw).lower()
    for alias, normalized in FIELD_ALIASES.items():
        alias_key = re.sub(r"[\s_\-（）()]+", "", alias).lower()
        if key == alias_key:
            return normalized
    if "受试者" in raw:
        return "受试者编号"
    if raw.endswith("日期") or raw.endswith("时间"):
        return "日期"
    if "访视" in raw and "OID" in raw.upper():
        return "访视编号"
    if "访视" in raw:
        return "访视名称"
    if "检查项目" in raw or raw == "LBTEST":
        return "检验项目"
    if "结果" in raw or raw in {"LBORRES", "LBSTRES"}:
        return "结果值"
    if raw in {"LBNRIND"}:
        return "正常值范围标记"
    return raw


def infer_data_type(field_name: str) -> str:
    name = normalize_field_name(field_name)
    raw = clean_text(field_name)
    if "日期" in name or "时间" in raw:
        return "日期/时间"
    if any(x in raw for x in ["得分", "评分", "下限", "上限", "年龄", "次数", "BSA", "EASI", "SCORAD"]):
        return "数值/文本混合"
    if "是否" in raw:
        return "布尔/枚举"
    return "文本"


def should_display_field(field_name: str) -> bool:
    raw = clean_text(field_name)
    return any(k in raw for k in DISPLAY_FIELD_KEYWORDS)


def needs_manual_review(field_name: str) -> bool:
    normalized = normalize_field_name(field_name)
    return normalized == "需人工确认" or normalized == clean_text(field_name)


def infer_sheet_category(sheet_name: str) -> str:
    name = clean_text(sheet_name)
    if any(x in name for x in ["EASI", "BSA", "SCORAD", "DLQI", "CDLQI", "NRS", "PROMIS", "IGA"]):
        return "疗效"
    if "实验室检查" in name or name.startswith("LB"):
        return "实验室"
    if "AE" in name or "不良事件" in name:
        return "AE"
    if "知情同意" in name or name.startswith("IC"):
        return "受试者基本信息"
    if "访视日期" in name or name.startswith("SV"):
        return "访视信息"
    if "受试者页" in name or "人口统计学" in name or "入组随机" in name:
        return "受试者基本信息"
    if "问题汇总" in name:
        return "Finding"
    if "偏离" in name:
        return "方案偏离"
    return "其他"


def metric_priority(metric_name: str, source_sheet: str) -> int:
    if metric_name == "SCORAD总分" or metric_name == "SCORAD":
        preferred_sheet = next((cfg["sheet"] for cfg in EFFICACY_CONFIG if cfg.get("metric") == "SCORAD"), "")
        if source_sheet == preferred_sheet:
            return 0
        return 10
    return 5


def classify_lab_value_for_plot(metric_name: str, result_value: Any, unit: Any) -> dict[str, Any]:
    text = clean_text(result_value)
    numeric = to_float(text)
    urine_like = any(k in clean_text(metric_name) for k in ["尿", "潜血", "蛋白", "酮体", "白细胞", "红细胞", "胆原", "胆红素"])
    qualitative_tokens = {"阴性", "阳性", "弱阳性", "未查", "稻黄色", "清亮", "浑浊", "NA", "1+", "2+", "3+", "4+", "+", "++", "+++", "++++"}
    if numeric is None and (urine_like or text in qualitative_tokens or "+" in text):
        return {"plot_mode": "categorical", "numeric_value": None, "display_value": text}
    return {"plot_mode": "numeric", "numeric_value": numeric, "display_value": text}


def build_subject_key(row: dict[str, Any]) -> str:
    return f"{clean_text(row.get('中心'))}|{clean_text(row.get('受试者编号'))}"


def standardize_center_name(name: Any) -> str:
    text = clean_text(name)
    if text.startswith("(03)"):
        text = text[4:]
    if text.startswith("(13)"):
        text = text[4:]
    return text


def extract_subject_id(*values: Any) -> str:
    for value in values:
        text = clean_text(value)
        if SUBJECT_ID_RE.fullmatch(text):
            return text
        match = SUBJECT_ID_RE.search(text)
        if match:
            return match.group(0)
    return ""


def read_sheet_matrix(path: Path, sheet_name: str, max_rows: int | None = None) -> list[list[Any]]:
    wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
    try:
        ws = wb[sheet_name]
        rows: list[list[Any]] = []
        for idx, row in enumerate(ws.iter_rows(values_only=True), start=1):
            rows.append(list(row))
            if max_rows is not None and idx >= max_rows:
                break
        return rows
    finally:
        wb.close()


def read_listing_sheet(path: Path, sheet_name: str) -> dict[str, Any]:
    rows = read_sheet_matrix(path, sheet_name)
    headers = make_unique_headers([clean_text(v) for v in (rows[0] if rows else [])])
    code_headers = make_unique_headers([clean_text(v) for v in (rows[1] if len(rows) > 1 else [])])
    records: list[dict[str, Any]] = []
    for raw_row in rows[2:]:
        if not any(v not in (None, "") for v in raw_row):
            continue
        row = {headers[idx]: raw_row[idx] if idx < len(raw_row) else "" for idx in range(len(headers))}
        records.append(row)
    return {"sheet_name": sheet_name, "headers": headers, "code_headers": code_headers, "records": records}


def resolve_sheet_alias(alias: str, optional: bool = False) -> str:
    sheet_name = clean_text(SHEET_ALIASES.get(alias))
    if sheet_name:
        return sheet_name
    if optional:
        return ""
    raise KeyError(f"Missing required sheet alias: {alias}")


def read_listing_alias(alias: str, optional: bool = False) -> dict[str, Any]:
    sheet_name = resolve_sheet_alias(alias, optional=optional)
    if not sheet_name:
        return {"sheet_name": "", "headers": [], "code_headers": [], "records": []}
    return read_listing_sheet(LISTING_XLSX, sheet_name)


def detect_header_row(rows: list[list[Any]]) -> int:
    for idx, row in enumerate(rows[:10]):
        joined = " | ".join(clean_text(v) for v in row if clean_text(v))
        if "序号" in joined and ("自查问题描述" in joined or "受试者筛选编号" in joined or "访视" in joined):
            return idx
    return 0


def read_generic_sheet(path: Path, sheet_name: str) -> dict[str, Any]:
    rows = read_sheet_matrix(path, sheet_name)
    header_idx = detect_header_row(rows)
    headers = make_unique_headers([clean_text(v) for v in (rows[header_idx] if rows else [])])
    records: list[dict[str, Any]] = []
    for raw_row in rows[header_idx + 1 :]:
        if not any(v not in (None, "") for v in raw_row):
            continue
        row = {headers[idx]: raw_row[idx] if idx < len(raw_row) else "" for idx in range(len(headers))}
        records.append(row)
    return {"sheet_name": sheet_name, "headers": headers, "records": records}


def load_subject_report_index() -> dict[str, dict[str, str]]:
    if not SUBJECT_REPORT_XLS.exists():
        return {}
    book = xlrd.open_workbook(str(SUBJECT_REPORT_XLS), formatting_info=False)
    sheet = book.sheet_by_index(0)
    headers = [clean_text(v) for v in sheet.row_values(4)]
    index: dict[str, dict[str, str]] = {}
    for row_idx in range(5, sheet.nrows):
        values = sheet.row_values(row_idx)
        if not values:
            continue
        subject = clean_text(values[0])
        if not SUBJECT_ID_RE.fullmatch(subject):
            continue
        row = {headers[i]: clean_text(values[i]) if i < len(values) else "" for i in range(len(headers))}
        center = standardize_center_name(row.get("中心名称"))
        if center not in TARGET_CENTERS:
            continue
        index[subject] = {
            "中心": center,
            "性别": row.get("性别", ""),
            "年龄组": infer_age_group_from_report(row.get("分层(IGA、年龄)")),
            "随机组别": normalize_study_group(row.get("研究分组")),
            "随机时间": row.get("随机时间", ""),
            "随机号": row.get("随机号", ""),
            "受试者状态_报表": row.get("受试者状态", ""),
        }
    return index


def scan_workbook_structure(path: Path, listing_mode: bool) -> list[dict[str, Any]]:
    wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
    try:
        sheet_specs = []
        for sheet_name in wb.sheetnames:
            ws = wb[sheet_name]
            preview = [list(row) for row in ws.iter_rows(min_row=1, max_row=6, values_only=True)]
            if listing_mode:
                headers = make_unique_headers([clean_text(v) for v in (preview[0] if preview else [])])
            else:
                header_idx = detect_header_row(preview)
                headers = make_unique_headers([clean_text(v) for v in (preview[header_idx] if preview else [])])
            sheet_specs.append({"file_name": path.name, "sheet_name": sheet_name, "headers": headers})
        return sheet_specs
    finally:
        wb.close()


def build_precheck_summary(runtime: dict[str, Any]) -> dict[str, Any]:
    issues: list[dict[str, str]] = []
    workbook_inventory: dict[str, list[dict[str, Any]]] = {}
    suggested_config = build_suggested_project_config()

    if not LISTING_XLSX.exists():
        issues.append({"级别": "阻断", "项目": "主listing", "说明": "未识别到主listing文件，无法继续。", "建议": "请提供受试者级listing或明确文件路径。"})
    else:
        listing_specs = scan_workbook_structure(LISTING_XLSX, listing_mode=True)
        workbook_inventory["listing"] = listing_specs
        for alias in REQUIRED_LISTING_SHEETS:
            if not runtime.get("sheet_aliases", {}).get(alias):
                issues.append({"级别": "阻断", "项目": "主listing", "说明": f"缺少关键sheet映射：{alias}", "建议": "请补充对应sheet，或明确替代sheet名称后再继续。"})
        if not runtime.get("efficacy_config"):
            issues.append({"级别": "阻断", "项目": "疗效终点识别", "说明": "未能基于方案主次要终点和研究流程识别出可纳入的疗效指标。", "建议": "请补充方案文件，或确认本轮需要纳入的疗效指标及其对应sheet。"})
        if not runtime.get("lab_config"):
            issues.append({"级别": "阻断", "项目": "实验室sheet", "说明": "未自动识别到任何可用实验室sheet。", "建议": "请确认实验室sheet名称，或说明该项目不纳入实验室profile。"})
        for item in runtime.get("efficacy_config", []):
            if not item.get("value_col_confirmed"):
                issues.append({"级别": "需确认", "项目": "疗效字段映射", "说明": f"{item['metric']} 的数值列未能自动确认，当前拟用 {item['value_col']}。", "建议": "请确认结果值列名，或决定本轮不纳入该指标。"})
            if not item.get("date_col_confirmed"):
                issues.append({"级别": "需确认", "项目": "疗效字段映射", "说明": f"{item['metric']} 的日期列未能自动确认，当前拟用 {item['date_col']}。", "建议": "请确认日期列名，或允许仅按访视顺序展示。"})
            if item.get("variable_type") == "需确认":
                issues.append({"级别": "需确认", "项目": "疗效变量类型", "说明": f"{item['metric']} 在方案终点描述中未能判断为连续型还是二分类。", "建议": "请确认该指标的变量类型及展示方式。"})

    if not FINDING_XLSX.exists():
        issues.append({"级别": "阻断", "项目": "Finding台账", "说明": "未识别到Finding台账文件，无法生成固定展示模块。", "建议": "请提供自查/稽查/Finding台账。"})
    else:
        finding_specs = scan_workbook_structure(FINDING_XLSX, listing_mode=False)
        workbook_inventory["finding"] = finding_specs
        if not FINDING_SHEET_SPECS:
            issues.append({"级别": "阻断", "项目": "Finding台账", "说明": "已找到Finding工作簿，但无法自动定位Finding sheet或受试者字段。", "建议": "请指定Finding sheet、中心对应关系和受试者字段。"})
        else:
            for spec in FINDING_SHEET_SPECS:
                if spec["center"] == MISSING_TEXT or spec["subject_col"] == MISSING_TEXT:
                    issues.append({"级别": "需确认", "项目": "Finding台账", "说明": f"Finding sheet {spec['sheet_name']} 的中心或受试者字段未能自动确认。", "建议": "请确认该sheet对应中心及受试者字段，或决定排除该sheet。"})

    if PD_DEF_XLSX and PD_DEF_XLSX.exists():
        workbook_inventory["pd_definition"] = scan_workbook_structure(PD_DEF_XLSX, listing_mode=False)
    else:
        issues.append({"级别": "提示", "项目": "方案偏离分类表", "说明": "未提供方案偏离分类表。", "建议": "如需区分重大/一般方案偏离，建议补充。"})

    if not SUBJECT_REPORT_XLS.exists():
        issues.append({"级别": "提示", "项目": "受试者报表", "说明": "未提供受试者报表，性别/年龄组/组别将仅依赖主listing。", "建议": "如需交叉核对DM与随机分组，建议补充受试者报表。"})

    if not PROTOCOL_FILES:
        issues.append({"级别": "阻断", "项目": "方案文件", "说明": "未识别到研究方案或勘误，无法基于主次要终点和研究流程确定疗效指标范围。", "建议": "请补充方案文件；如本轮允许手工指定疗效指标，请明确说明。"})
    elif not runtime.get("protocol_summary", {}).get("selected_metrics"):
        issues.append({"级别": "阻断", "项目": "方案终点解析", "说明": "已读取方案文件，但未从主次要终点或研究流程中识别出明确疗效指标。", "建议": "请确认方案文件内容可读，或手工指定需纳入的疗效指标。"})

    if not TARGET_CENTERS:
        issues.append({"级别": "阻断", "项目": "中心范围", "说明": "无法从主listing自动识别中心。", "建议": "请手动提供目标中心列表。"})
    elif len(TARGET_CENTERS) > 1 and not runtime.get("centers_from_user"):
        issues.append({"级别": "需确认", "项目": "中心范围", "说明": f"自动识别到 {len(TARGET_CENTERS)} 个中心，当前将默认纳入全部中心。", "建议": "请确认是否纳入全部中心；如仅需重点中心，请显式提供中心列表。"})

    questions: list[str] = []
    for issue in issues:
        if issue["级别"] in {"阻断", "需确认"}:
            questions.append(f"{issue['项目']}：{issue['建议']}")

    return {
        "runtime": runtime,
        "workbook_inventory": workbook_inventory,
        "suggested_config": suggested_config,
        "issues": issues,
        "questions": questions,
        "ready_to_build": not any(issue["级别"] == "阻断" for issue in issues),
    }


def write_precheck_outputs(precheck: dict[str, Any]) -> None:
    OUTPUT_DIR.mkdir(parents=True, exist_ok=True)
    (OUTPUT_DIR / "input_precheck.json").write_text(json.dumps(precheck, ensure_ascii=False, indent=2), encoding="utf-8")
    (OUTPUT_DIR / "suggested_project_config.json").write_text(json.dumps(precheck["suggested_config"], ensure_ascii=False, indent=2), encoding="utf-8")
    (OUTPUT_DIR / "protocol_endpoint_summary.json").write_text(json.dumps(precheck["runtime"].get("protocol_summary", {}), ensure_ascii=False, indent=2), encoding="utf-8")
    (OUTPUT_DIR / "protocol_endpoint_summary.md").write_text(build_protocol_summary_markdown(precheck["runtime"].get("protocol_summary", {})), encoding="utf-8")
    lines = [
        "# Patient Profile 输入预检查",
        "",
        f"- 可继续构建：{'是' if precheck['ready_to_build'] else '否'}",
        f"- 已识别中心：{'；'.join(precheck['runtime'].get('target_centers', [])) or MISSING_TEXT}",
        f"- 主listing：{precheck['runtime'].get('listing_xlsx') or MISSING_TEXT}",
        f"- Finding台账：{precheck['runtime'].get('finding_xlsx') or MISSING_TEXT}",
        f"- 方案偏离分类表：{precheck['runtime'].get('pd_def_xlsx') or MISSING_TEXT}",
        f"- 受试者报表：{precheck['runtime'].get('subject_report_xls') or MISSING_TEXT}",
        f"- 建议配置文件：{OUTPUT_DIR / 'suggested_project_config.json'}",
        f"- 方案终点解构：{OUTPUT_DIR / 'protocol_endpoint_summary.md'}",
        "",
        "## 问题清单",
    ]
    if precheck["issues"]:
        for issue in precheck["issues"]:
            lines.append(f"- [{issue['级别']}] {issue['项目']}：{issue['说明']} 建议：{issue['建议']}")
    else:
        lines.append("- 无阻断性问题。")
    lines.extend(["", "## 需向用户确认的问题"])
    if precheck["questions"]:
        for question in precheck["questions"]:
            lines.append(f"- {question}")
    else:
        lines.append("- 无。")
    (OUTPUT_DIR / "input_precheck.md").write_text("\n".join(lines), encoding="utf-8")


def build_field_mapping_rows(sheet_specs: list[dict[str, Any]]) -> list[dict[str, Any]]:
    rows: list[dict[str, Any]] = []
    for sheet in sheet_specs:
        category = infer_sheet_category(sheet["sheet_name"])
        for header in sheet["headers"]:
            if not clean_text(header):
                continue
            rows.append(
                {
                    "文件名": sheet["file_name"],
                    "sheet名": sheet["sheet_name"],
                    "数据类别": category,
                    "原始字段名": header,
                    "标准化字段名": normalize_field_name(header),
                    "数据类型": infer_data_type(header),
                    "是否用于HTML展示": "是" if should_display_field(header) else "否",
                    "是否需要人工复核": "是" if needs_manual_review(header) else "否",
                    "备注": "字段标准化自动推断" if not needs_manual_review(header) else "需人工确认映射",
                }
            )
    return rows


def parse_reference_html_groups() -> dict[str, str]:
    subject_group: dict[str, str] = {}
    pattern = re.compile(
        r'<section class="subject-card" data-subject="([^"]+)" data-center="[^"]*">.*?<div class="subject-sub">(.*?)</div>',
        re.S,
    )
    for path in REF_HTMLS:
        if not path.exists():
            continue
        text = path.read_text(encoding="utf-8", errors="ignore")
        for subject, group in pattern.findall(text):
            subject_group[clean_text(subject)] = clean_text(group)
    return subject_group


def build_visit_index() -> dict[tuple[str, str], dict[str, str]]:
    sheet = read_listing_alias("SV")
    visit_index: dict[tuple[str, str], dict[str, str]] = {}
    for row in sheet["records"]:
        center = standardize_center_name(row.get("研究中心"))
        subject = clean_text(row.get("受试者"))
        if center not in TARGET_CENTERS or not subject:
            continue
        visit = clean_text(row.get("访视OID"))
        visit_index[(subject, visit)] = {
            "访视名称": clean_text(row.get("访视名称")),
            "访视编号": visit,
            "访视日期": parse_date(row.get("访视日期")),
            "计划最早访视日期": parse_date(row.get("计划最早访视日期（系统生成）")),
            "计划最晚访视日期": parse_date(row.get("计划最晚访视日期（系统生成）")),
        }
    return visit_index


def parse_subject_profiles(reference_groups: dict[str, str]) -> list[dict[str, Any]]:
    subj_sheet = read_listing_alias("SUBJ")
    dm_sheet = read_listing_alias("DM")
    rand_sheet = read_listing_alias("RAND")
    ic_sheet = read_listing_alias("IC")
    dseot_sheet = read_listing_alias("DSEOT", optional=True)
    dseos_sheet = read_listing_alias("DSEOS", optional=True)
    ae_sheet = read_listing_alias("AE")
    pc_sheet = read_listing_alias("PC", optional=True)
    visit_index = build_visit_index()
    subject_report_index = load_subject_report_index()

    dm_map = {clean_text(r.get("受试者")): r for r in dm_sheet["records"] if clean_text(r.get("受试者"))}
    rand_map = {clean_text(r.get("受试者")): r for r in rand_sheet["records"] if clean_text(r.get("受试者"))}
    ic_map = {clean_text(r.get("受试者")): r for r in ic_sheet["records"] if clean_text(r.get("受试者"))}
    dseot_map = {clean_text(r.get("受试者")): r for r in dseot_sheet["records"] if clean_text(r.get("受试者"))}
    dseos_map = {clean_text(r.get("受试者")): r for r in dseos_sheet["records"] if clean_text(r.get("受试者"))}

    ae_flags = defaultdict(lambda: {"ae": False, "sae": False})
    for row in ae_sheet["records"]:
        subject = clean_text(row.get("受试者"))
        if not subject:
            continue
        if clean_text(row.get("是否发生不良事件？")) == "是":
            ae_flags[subject]["ae"] = True
        if clean_text(row.get("是否为严重不良事件（SAE）？")) == "是":
            ae_flags[subject]["sae"] = True

    pk_flags = defaultdict(bool)
    for row in pc_sheet["records"]:
        subject = clean_text(row.get("受试者"))
        if clean_text(row.get("是否进行药代动力学（PK）血样采集？")) == "是":
            pk_flags[subject] = True

    visit_presence = defaultdict(set)
    for (subject, visit), meta in visit_index.items():
        if meta.get("访视日期"):
            visit_presence[subject].add(visit)

    profiles: list[dict[str, Any]] = []
    for row in subj_sheet["records"]:
        center = standardize_center_name(row.get("研究中心"))
        subject = clean_text(row.get("受试者"))
        if center not in TARGET_CENTERS or not subject:
            continue
        dm = dm_map.get(subject, {})
        rand = rand_map.get(subject, {})
        ic = ic_map.get(subject, {})
        dseot = dseot_map.get(subject, {})
        dseos = dseos_map.get(subject, {})
        report_row = subject_report_index.get(subject, {})

        age = clean_text(dm.get("年龄（系统生成）"))
        age_group = MISSING_TEXT
        age_num = to_float(age)
        if age_num is not None:
            age_group = "青少年" if age_num < 18 else "成人"
        if report_row.get("年龄组"):
            age_group = report_row["年龄组"]

        stage = clean_text(dseot.get("治疗结束阶段"))
        completed_blind = "是" if "D57" in visit_presence[subject] else MISSING_TEXT
        completed_open = "是" if stage == "开放治疗期" else MISSING_TEXT
        completed_study = "是" if clean_text(row.get("受试者状态")) == "完成试验" else "否"
        screen_fail_reason = clean_text(rand.get("未入组原因")) or clean_text(rand.get("其他原因，请说明")) or MISSING_TEXT
        blind_enter = parse_date(rand.get("入组日期")) or parse_date(rand.get("随机日期和时间（系统生成）")) or MISSING_TEXT
        blind_complete = visit_index.get((subject, "D57"), {}).get("访视日期", MISSING_TEXT)
        open_enter = visit_index.get((subject, "D85"), {}).get("访视日期", MISSING_TEXT)
        open_complete = parse_date(dseot.get("末次给药日期")) if stage == "开放治疗期" else MISSING_TEXT

        profiles.append(
            {
                "中心": center,
                "中心编号": clean_text(row.get("中心编号")) or MISSING_TEXT,
                "受试者编号": subject,
                "受试者键": f"{center}|{subject}",
                "受试者状态": clean_text(row.get("受试者状态")) or MISSING_TEXT,
                "年龄": age or MISSING_TEXT,
                "年龄组": age_group,
                "性别": report_row.get("性别") or clean_text(dm.get("性别")) or MISSING_TEXT,
                "随机组别": report_row.get("随机组别") or reference_groups.get(subject, MISSING_TEXT),
                "筛选日期": visit_index.get((subject, "SCR"), {}).get("访视日期", MISSING_TEXT),
                "知情同意日期": parse_date(ic.get("首次签署知情同意书日期")) or MISSING_TEXT,
                "基线/随机日期": blind_enter,
                "筛选失败原因": screen_fail_reason if clean_text(row.get("受试者状态")) == "筛选失败" else "",
                "双盲期进入日期": blind_enter,
                "双盲期完成日期": blind_complete,
                "开放期进入日期": open_enter,
                "开放期完成日期": open_complete or MISSING_TEXT,
                "是否完成双盲期": completed_blind,
                "是否完成开放期": completed_open,
                "是否完成研究": completed_study,
                "是否进入ITT": "是" if clean_text(rand.get("受试者是否入组本研究？")) == "是" else MISSING_TEXT,
                "是否进入PKCS": "是" if clean_text(rand.get("是否进行PK血样采集？")) == "是" or pk_flags[subject] else "否",
                "是否存在AE/SAE/SUSAR": "AE" if ae_flags[subject]["ae"] and not ae_flags[subject]["sae"] else ("SAE" if ae_flags[subject]["sae"] else "否"),
                "是否存在AE": "是" if ae_flags[subject]["ae"] else "否",
                "是否存在PD": MISSING_TEXT,
                "是否存在方案偏离": MISSING_TEXT,
                "是否存在核查Finding": "否",
                "是否为重点核查受试者": MISSING_TEXT,
                "数据完整性风险等级": "待定",
                "研究结束日期": parse_date(dseos.get("研究结束日期")) or MISSING_TEXT,
                "治疗结束日期": parse_date(dseot.get("末次给药日期")) or MISSING_TEXT,
                "受试者报表随机号": report_row.get("随机号", ""),
            }
        )
    profiles.sort(key=lambda x: (x["中心"], x["受试者编号"]))
    return profiles


def build_ae_rows() -> list[dict[str, Any]]:
    ae_sheet = read_listing_alias("AE")
    rows: list[dict[str, Any]] = []
    for row in ae_sheet["records"]:
        center = standardize_center_name(row.get("研究中心"))
        subject = clean_text(row.get("受试者"))
        if center not in TARGET_CENTERS or not subject:
            continue
        if clean_text(row.get("是否发生不良事件？")) != "是":
            continue
        rows.append(
            {
                "中心": center,
                "受试者编号": subject,
                "受试者键": f"{center}|{subject}",
                "AE名称": clean_text(row.get("不良事件名称")) or MISSING_TEXT,
                "最早开始日期": parse_date(row.get("最早开始日期")) or MISSING_TEXT,
                "最严重程度（CTCAE V5.0）": clean_text(row.get("最严重程度（CTCAE V5.0）")) or MISSING_TEXT,
                "与研究用药的关系": clean_text(row.get("与研究用药的关系")) or MISSING_TEXT,
                "转归": clean_text(row.get("转归")) or MISSING_TEXT,
                "结束日期": parse_date(row.get("结束日期")) or MISSING_TEXT,
                "是否为严重不良事件（SAE）": clean_text(row.get("是否为严重不良事件（SAE）？")) or "否",
            }
        )
    return rows


def numeric_response_label(metric: str, baseline: float | None, current: float | None) -> str:
    if baseline is None or current is None:
        return MISSING_TEXT
    selected_labels = {item["label"] for item in selected_response_rules_for_metric(metric)}
    labels: list[str] = []
    if metric == "EASI":
        reduction = (baseline - current) / baseline * 100 if baseline else None
        if reduction is None:
            return MISSING_TEXT
        threshold_map = [(90, "EASI-90"), (75, "EASI-75"), (50, "EASI-50")]
        for threshold, label in threshold_map:
            if label in selected_labels and reduction >= threshold:
                labels.append(label)
        return " / ".join(labels) if labels else MISSING_TEXT
    if metric == "IGA":
        if current in {0, 1}:
            if "IGA-TS" in selected_labels and (baseline - current) >= 2:
                labels.append("IGA-TS")
            if "IGA 0/1分" in selected_labels:
                labels.append("IGA 0/1分")
        return "；".join(labels) if labels else MISSING_TEXT
    if metric == "NRS":
        diff = baseline - current
        if "Itch NRS改善≥4分" in selected_labels and diff >= 4:
            return "Itch NRS改善≥4分"
        return MISSING_TEXT
    if metric in {"QSI-PROMIS睡眠相关影响8a", "QSI-PROMIS睡眠困扰8b"}:
        diff = baseline - current
        label = "PROMIS-8a临床意义改善" if metric.endswith("8a") else "PROMIS-8b临床意义改善"
        if label in selected_labels and diff >= 6:
            return label
        return MISSING_TEXT
    return MISSING_TEXT


def build_efficacy_rows(visit_index: dict[tuple[str, str], dict[str, str]]) -> list[dict[str, Any]]:
    all_rows: list[dict[str, Any]] = []
    grouped_for_baseline: defaultdict[tuple[str, str], list[dict[str, Any]]] = defaultdict(list)
    for config in EFFICACY_CONFIG:
        sheet = read_listing_sheet(LISTING_XLSX, config["sheet"])
        grouped = defaultdict(list)
        for row in sheet["records"]:
            center = standardize_center_name(row.get("研究中心"))
            subject = clean_text(row.get("受试者"))
            if center not in TARGET_CENTERS or not subject:
                continue
            visit_oid = clean_text(row.get("访视OID"))
            grouped[(subject, visit_oid)].append(row)

        for (subject, visit_oid), rows in grouped.items():
            first = rows[0]
            metric = config["metric"]
            raw_value = first.get(config["value_col"])
            numeric_value = parse_iga_score(raw_value) if metric == "IGA" else to_float(raw_value)
            date_value = parse_date(first.get(config["date_col"])) or visit_index.get((subject, visit_oid), {}).get("访视日期", "")
            duplicate_count = len(rows)
            note_parts = []
            if duplicate_count > 1 and config["aggregate"] == "total_with_details":
                note_parts.append(f"同一访视按部位展开 {duplicate_count} 条记录，已汇总为单个总分数据点")
            elif duplicate_count > 1:
                note_parts.append("同一指标同一访视存在多条记录，需人工复核")
            if not date_value:
                note_parts.append("日期缺失")
            planned = visit_index.get((subject, visit_oid), {})
            early = planned.get("计划最早访视日期", "")
            late = planned.get("计划最晚访视日期", "")
            if date_value and early and late and (date_value < early or date_value > late):
                note_parts.append("访视日期疑似超出访视窗")

            details = []
            for item in rows:
                details.append({field: clean_text(item.get(field)) for field in config["detail_fields"] if clean_text(item.get(field))})

            merged = {
                "中心": standardize_center_name(first.get("研究中心")),
                "受试者编号": subject,
                "受试者键": f"{standardize_center_name(first.get('研究中心'))}|{subject}",
                "指标名称": metric,
                "指标分组": config["group"],
                "访视编号": visit_oid or MISSING_TEXT,
                "访视名称": clean_text(first.get("访视名称")) or planned.get("访视名称", MISSING_TEXT),
                "日期": date_value or MISSING_TEXT,
                "原始值": clean_text(raw_value) or MISSING_TEXT,
                "数值结果": numeric_value,
                "较基线变化": MISSING_TEXT,
                "较基线百分比变化": MISSING_TEXT,
                "关键应答判定": MISSING_TEXT,
                "数据来源sheet": config["sheet"],
                "原始字段名": config["value_col"],
                "是否存在缺失、重复、异常日期或访视窗问题": "是" if note_parts else "否",
                "问题标记": "；".join(note_parts) if note_parts else "",
                "与Finding是否有关联": "否",
                "关联Finding编号": "",
                "部位/单项详情": json.dumps(details, ensure_ascii=False),
                "重复记录数": str(duplicate_count),
            }
            grouped_for_baseline[(subject, metric)].append(merged)
            all_rows.append(merged)

    for (_, metric), rows in grouped_for_baseline.items():
        baseline_row = None
        for row in rows:
            if row["访视编号"] == "D1":
                baseline_row = row
                break
        if baseline_row is None:
            rows_sorted = sorted(rows, key=lambda x: (x["日期"] == MISSING_TEXT, x["日期"], x["访视编号"]))
            baseline_row = rows_sorted[0] if rows_sorted else None
        baseline = baseline_row["数值结果"] if baseline_row else None
        for row in rows:
            current = row["数值结果"]
            if baseline is not None and current is not None:
                change = current - baseline
                row["较基线变化"] = f"{change:.4g}"
                if baseline != 0:
                    row["较基线百分比变化"] = f"{change / baseline * 100:.2f}%"
                if should_assign_response_label(row["访视编号"], row["访视名称"], metric):
                    row["关键应答判定"] = numeric_response_label(metric, baseline, current)
    return sorted(all_rows, key=lambda x: (x["中心"], x["受试者编号"], efficacy_sort_key(x["指标名称"]), x["日期"], x["访视编号"]))


def standardize_lab_test_name(raw_name: str) -> str:
    text = clean_text(raw_name)
    if not text:
        return MISSING_TEXT
    match = re.search(r"（([A-Za-z0-9\-]+)）", text)
    if match:
        code = match.group(1)
        return code if code.isupper() else text
    if text in {"尿PH", "PH"}:
        return "尿pH"
    return text


def abnormal_direction(result: float | None, low: float | None, high: float | None) -> str:
    if result is None:
        return "无法判断"
    if low is None and high is None:
        return "无法判断"
    if low is not None and result < low:
        return "低"
    if high is not None and result > high:
        return "高"
    if low is not None or high is not None:
        return "正常"
    return "正常"


def normalize_lab_row(row: dict[str, Any]) -> dict[str, Any]:
    result_numeric = to_float(row.get("结果值"))
    low = to_float(row.get("正常值下限"))
    high = to_float(row.get("正常值上限"))
    direction = abnormal_direction(result_numeric, low, high)
    note = clean_text(row.get("备注"))
    if direction == "无法判断" and (not clean_text(row.get("正常值下限")) and not clean_text(row.get("正常值上限"))):
        note = f"{note}；正常范围缺失，无法判断异常".strip("；")
    output = dict(row)
    output["数值结果"] = result_numeric
    output["异常方向"] = direction
    output["是否低于正常下限"] = "是" if direction == "低" else "否"
    output["是否高于正常上限"] = "是" if direction == "高" else "否"
    output["备注"] = note or ""
    return output


def apply_longitudinal_attention_flags(
    metric_rows: list[dict[str, Any]],
    metric_label_key: str,
    category_key: str,
) -> None:
    metric_rows.sort(key=lambda x: (x["日期"] == MISSING_TEXT, x["日期"], x["访视编号"]))
    later_normals = [r for r in metric_rows if r["异常方向"] == "正常"]
    baseline_row = next((r for r in metric_rows if r["访视编号"] == "D1"), metric_rows[0] if metric_rows else None)
    baseline_value = baseline_row["数值结果"] if baseline_row else None
    baseline_normal = baseline_row["异常方向"] == "正常" if baseline_row else False
    deviations: list[tuple[dict[str, Any], float | None]] = []
    for row in metric_rows:
        if row["异常方向"] in {"高", "低"}:
            row["是否转归正常"] = "是" if any(
                r["日期"] > row["日期"]
                for r in later_normals
                if r["日期"] != MISSING_TEXT and row["日期"] != MISSING_TEXT
            ) else "未在现有文件中确认"
        if row.get(category_key) == "尿常规" and row.get("图表模式") == "categorical":
            row["备注"] = (row["备注"] + "；定性结果使用分类展示，不强制数值化").strip("；")
        flags: list[str] = []
        high = to_float(row.get("正常值上限"))
        low = to_float(row.get("正常值下限"))
        value = row.get("数值结果")
        deviation_ratio = None
        if row["异常方向"] == "高" and value is not None and high not in (None, 0):
            deviation_ratio = (value - high) / high
        elif row["异常方向"] == "低" and value is not None and low not in (None, 0):
            deviation_ratio = (low - value) / low
        if deviation_ratio is not None and deviation_ratio >= 0.5:
            flags.append("偏离正常值范围≥50%")
        if baseline_value not in (None, 0) and value is not None and row is not baseline_row:
            if abs(value - baseline_value) / abs(baseline_value) >= 0.5:
                flags.append("较基线变化幅度≥50%")
        if baseline_normal and row is not baseline_row and row["异常方向"] in {"高", "低"}:
            flags.append("基线正常后转异常")
        metric_name = clean_text(row.get(metric_label_key))
        if metric_name == "收缩压" and value is not None:
            if value >= 180:
                flags.append("收缩压达到180mmHg档位")
            elif value >= 160:
                flags.append("收缩压达到160mmHg档位")
        if metric_name == "舒张压" and value is not None:
            if value >= 110:
                flags.append("舒张压达到110mmHg档位")
            elif value >= 100:
                flags.append("舒张压达到100mmHg档位")
        row["额外关注标记"] = "；".join(unique(flags))
        deviations.append((row, deviation_ratio))

    cs_rows = [item for item in deviations if item[0]["临床意义判断"] == "异常有临床意义" and item[1] is not None]
    ncs_rows = [item for item in deviations if item[0]["临床意义判断"] == "异常无临床意义" and item[1] is not None]
    if cs_rows and ncs_rows:
        min_cs = min(item[1] for item in cs_rows)
        max_ncs = max(item[1] for item in ncs_rows)
        if max_ncs > min_cs:
            for row, ratio in ncs_rows:
                if ratio == max_ncs:
                    row["额外关注标记"] = "；".join(
                        unique([item for item in clean_text(row["额外关注标记"]).split("；") if item] + ["偏离更多但评估为异常无临床意义"])
                    )
        for row_cs, ratio_cs in cs_rows:
            for row_ncs, ratio_ncs in ncs_rows:
                if abs(ratio_cs - ratio_ncs) <= 0.1:
                    row_cs["额外关注标记"] = "；".join(
                        unique([item for item in clean_text(row_cs["额外关注标记"]).split("；") if item] + ["偏离程度相近但临床意义评价不一致"])
                    )
                    row_ncs["额外关注标记"] = "；".join(
                        unique([item for item in clean_text(row_ncs["额外关注标记"]).split("；") if item] + ["偏离程度相近但临床意义评价不一致"])
                    )


def build_lab_rows(visit_index: dict[tuple[str, str], dict[str, str]]) -> list[dict[str, Any]]:
    rows: list[dict[str, Any]] = []
    by_subject_metric: defaultdict[tuple[str, str], list[dict[str, Any]]] = defaultdict(list)
    for lab_config in LAB_CONFIG:
        sheet_name = lab_config["sheet"]
        lab_category = lab_config["lab_category"]
        sheet = read_listing_sheet(LISTING_XLSX, sheet_name)
        for row in sheet["records"]:
            center = standardize_center_name(row.get("研究中心"))
            subject = clean_text(row.get("受试者"))
            if center not in TARGET_CENTERS or not subject:
                continue
            visit_oid = clean_text(row.get("访视OID"))
            test_name = clean_text(row.get("检查项目"))
            if not should_include_lab_test(test_name, lab_category):
                continue
            metric_name = standardize_lab_test_name(test_name)
            value_info = classify_lab_value_for_plot(test_name, row.get("检查结果"), row.get("单位"))
            clinical_significance = clean_text(row.get("临床意义评价")) or MISSING_TEXT
            clinical_comment = clean_text(row.get("异常有临床意义，请说明"))
            lab_row = normalize_lab_row(
                {
                    "中心": center,
                    "受试者编号": subject,
                    "受试者键": f"{center}|{subject}",
                    "实验室类别": lab_category,
                    "实验室显示分组": infer_lab_group(metric_name, lab_category),
                    "访视编号": visit_oid or MISSING_TEXT,
                    "访视名称": clean_text(row.get("访视名称")) or visit_index.get((subject, visit_oid), {}).get("访视名称", MISSING_TEXT),
                    "日期": parse_date(row.get("采样日期")) or visit_index.get((subject, visit_oid), {}).get("访视日期", MISSING_TEXT),
                    "检验项目标准化名称": metric_name,
                    "原始检验项目名称": test_name or MISSING_TEXT,
                    "结果值": clean_text(row.get("检查结果")) or MISSING_TEXT,
                    "单位": clean_text(row.get("单位")) or MISSING_TEXT,
                    "正常值下限": clean_text(row.get("下限")),
                    "正常值上限": clean_text(row.get("上限")),
                    "正常范围原始文本": clean_text(row.get("标准值")) or "",
                    "CTCAE分级": MISSING_TEXT,
                    "临床意义判断": clinical_significance,
                    "研究者临床评估": "" if clinical_significance == "正常" else clinical_comment,
                    "备注": "",
                    "是否转归正常": MISSING_TEXT,
                    "是否对应AE或Finding": "否",
                    "数据来源sheet": sheet_name,
                    "原始字段名": "检查结果",
                    "是否需人工复核": "是" if value_info["plot_mode"] == "categorical" and lab_category != "尿常规" else "否",
                    "图表模式": value_info["plot_mode"],
                    "正常值范围标记": clean_text(row.get("正常值范围标记")),
                    "额外关注标记": "",
                }
            )
            rows.append(lab_row)
            by_subject_metric[(subject, lab_row["检验项目标准化名称"])].append(lab_row)

    eg_sheet = read_listing_alias("EG", optional=True)
    for row in eg_sheet["records"]:
        center = standardize_center_name(row.get("研究中心"))
        subject = clean_text(row.get("受试者"))
        if center not in TARGET_CENTERS or not subject:
            continue
        if clean_text(row.get("是否进行12导联心电图检查？")) != "是":
            continue
        visit_oid = clean_text(row.get("访视OID"))
        metric_name = "QTc间期"
        lab_row = normalize_lab_row(
            {
                "中心": center,
                "受试者编号": subject,
                "受试者键": f"{center}|{subject}",
                "实验室类别": "心电图",
                "实验室显示分组": "心电图",
                "访视编号": visit_oid or MISSING_TEXT,
                "访视名称": clean_text(row.get("访视名称")) or visit_index.get((subject, visit_oid), {}).get("访视名称", MISSING_TEXT),
                "日期": parse_date(row.get("检查日期")) or visit_index.get((subject, visit_oid), {}).get("访视日期", MISSING_TEXT),
                "检验项目标准化名称": metric_name,
                "原始检验项目名称": metric_name,
                "结果值": clean_text(row.get("QTc间期")) or MISSING_TEXT,
                "单位": clean_text(row.get("QTc间期_UNIT")) or MISSING_TEXT,
                "正常值下限": "",
                "正常值上限": "",
                "正常范围原始文本": "",
                "CTCAE分级": MISSING_TEXT,
                "临床意义判断": clean_text(row.get("临床意义评价")) or MISSING_TEXT,
                "研究者临床评估": "" if clean_text(row.get("临床意义评价")) == "正常" else clean_text(row.get("异常有临床意义，请说明")),
                "备注": "",
                "是否转归正常": MISSING_TEXT,
                "是否对应AE或Finding": "否",
                "数据来源sheet": eg_sheet["sheet_name"] or MISSING_TEXT,
                "原始字段名": "QTc间期",
                "是否需人工复核": "否",
                "图表模式": "numeric",
                "正常值范围标记": "",
                "额外关注标记": "",
            }
        )
        rows.append(lab_row)
        by_subject_metric[(subject, metric_name)].append(lab_row)

    for (_, _), metric_rows in by_subject_metric.items():
        apply_longitudinal_attention_flags(metric_rows, "检验项目标准化名称", "实验室类别")

    return sorted(rows, key=lab_sort_key)


def build_vital_rows(visit_index: dict[tuple[str, str], dict[str, str]]) -> list[dict[str, Any]]:
    rows: list[dict[str, Any]] = []
    by_subject_metric: defaultdict[tuple[str, str], list[dict[str, Any]]] = defaultdict(list)
    sheet = read_listing_alias("VS")
    metric_columns = [
        ("体温", "体温", "体温_UNIT"),
        ("心率", "脉搏", "脉搏_UNIT"),
        ("收缩压", "收缩压", "收缩压_UNIT"),
        ("舒张压", "舒张压", "舒张压_UNIT"),
    ]
    for row in sheet["records"]:
        center = standardize_center_name(row.get("研究中心"))
        subject = clean_text(row.get("受试者"))
        if center not in TARGET_CENTERS or not subject:
            continue
        if clean_text(row.get("是否进行生命体征检查？")) != "是":
            continue
        visit_oid = clean_text(row.get("访视OID"))
        clinical_significance = clean_text(row.get("临床意义评价")) or MISSING_TEXT
        clinical_comment = clean_text(row.get("异常有临床意义，请说明"))
        for metric_name, raw_col, unit_col in metric_columns:
            value = clean_text(row.get(raw_col))
            if not value or value == MISSING_TEXT:
                continue
            low, high = vital_reference_range(metric_name)
            vital_row = normalize_lab_row(
                {
                    "中心": center,
                    "受试者编号": subject,
                    "受试者键": f"{center}|{subject}",
                    "生命体征类别": "生命体征",
                    "访视编号": visit_oid or MISSING_TEXT,
                    "访视名称": clean_text(row.get("访视名称")) or visit_index.get((subject, visit_oid), {}).get("访视名称", MISSING_TEXT),
                    "日期": parse_date(row.get("检查日期")) or visit_index.get((subject, visit_oid), {}).get("访视日期", MISSING_TEXT),
                    "生命体征指标": metric_name,
                    "原始指标名称": raw_col,
                    "结果值": value,
                    "单位": clean_text(row.get(unit_col)) or MISSING_TEXT,
                    "正常值下限": "" if low is None else f"{low:g}",
                    "正常值上限": "" if high is None else f"{high:g}",
                    "正常范围原始文本": "",
                    "临床意义判断": clinical_significance,
                    "研究者临床评估": "" if clinical_significance == "正常" else clinical_comment,
                    "备注": "",
                    "是否对应AE或Finding": "否",
                    "数据来源sheet": sheet["sheet_name"] or MISSING_TEXT,
                    "原始字段名": raw_col,
                    "图表模式": "numeric",
                    "额外关注标记": "",
                }
            )
            rows.append(vital_row)
            by_subject_metric[(subject, metric_name)].append(vital_row)

    for (_, _), metric_rows in by_subject_metric.items():
        apply_longitudinal_attention_flags(metric_rows, "生命体征指标", "生命体征类别")

    return sorted(rows, key=vital_sort_row_key)


def standardize_problem_category(raw: str, desc: str) -> str:
    text = clean_text(raw) or clean_text(desc)
    mapping = [
        ("知情", "知情同意"),
        ("入排", "入排标准"),
        ("排除", "入排标准"),
        ("疗效", "疗效评价"),
        ("量表", "量表一致性"),
        ("eCOA", "eCOA"),
        ("日记", "日记卡"),
        ("药物", "研究药物管理"),
        ("合并用药", "合并用药"),
        ("AE", "AE/SAE/SUSAR"),
        ("实验室", "实验室检查"),
        ("PK", "PK采样"),
        ("授权", "授权分工"),
        ("伦理", "伦理文件"),
        ("PD", "方案偏离"),
    ]
    for key, label in mapping:
        if key.lower() in text.lower():
            return label
    return clean_text(raw) or "其他"


def standardize_problem_severity(raw: str) -> str:
    text = clean_text(raw)
    if any(k in text for k in ["重要", "重大", "高"]):
        return "高"
    if any(k in text for k in ["一般", "中"]):
        return "中"
    if any(k in text for k in ["轻微", "低"]):
        return "低"
    return "待定"


def impact_from_text(text: str, keywords: tuple[str, ...]) -> str:
    low = clean_text(text).lower()
    return "可能" if any(k.lower() in low for k in keywords) else "未见直接证据"


def build_medical_reply(record: dict[str, Any]) -> str:
    category = record["问题分类"]
    description = record["原始问题描述"]
    if category in {"疗效评价", "量表一致性", "eCOA", "日记卡"}:
        return "现有资料提示问题主要影响对应疗效/PRO记录的一致性或时间逻辑解释，尚未见足以直接改变已录入评分结果的明确证据；建议现场准备原始量表、eCOA后台时间戳、EDC截图和研究者说明。"
    if category in {"实验室检查", "AE/SAE/SUSAR"}:
        return "该问题需结合实验室原始报告、临床评估、AE记录及后续复查共同判断；若当前证据不足，建议以“需进一步核查”作答，避免直接下绝对结论。"
    if category in {"研究药物管理", "方案偏离", "PK采样"}:
        return "该问题更偏执行和依从性层面，需聚焦其对暴露、访视窗、样本可解释性和关键终点的实际影响；建议准备日志卡、发药回收、用药记录和方案条款。"
    if "知情" in description or category == "知情同意":
        return "该问题首先关系到受试者权益与GCP合规性，建议现场明确事实顺序、可追溯证据及补救措施，并避免仅用“笔误”作笼统解释。"
    return "当前资料显示该问题主要涉及记录完整性或流程留痕，是否影响关键结论仍需结合源文件进一步核查；建议现场准备原始记录、中心解释和整改依据。"


def build_reply_example(record: dict[str, Any]) -> str:
    category = record["问题分类"]
    desc = record["原始问题描述"]
    if category in {"疗效评价", "量表一致性", "eCOA", "日记卡"}:
        return "现场回复示例：先准备对应访视的原始量表、eCOA后台时间戳、病历记录和EDC截图，并把同一访视的时间顺序整理成一页对照。依据这些材料的评分原始记录和时间戳，回复老师“我们已逐项核对，该问题主要涉及留痕/时间逻辑一致性复核，现有证据未见直接改变该访视评分结果；如您需要，我们可以按原件顺序逐页说明。”"
    if category in {"实验室检查", "AE/SAE/SUSAR"}:
        return "现场回复示例：先准备原始化验单、研究者临床意义判断、AE记录、合并用药记录和后续复查结果，并把异常发生前后访视串联展示。依据这些材料的原始结果和临床评估，回复老师“该异常已定位到具体记录，我们先如实说明检测结果、研究者判断及后续复查情况；目前按现有证据解读其临床意义，如需进一步判断，我们也已备好对应源文件供现场核对。”"
    if category in {"研究药物管理", "方案偏离", "PK采样"}:
        return "现场回复示例：先准备药物发放回收记录、用药日志卡、样本处理记录、访视安排和整改留痕，并把偏差发生经过按时间线列清。依据这些材料的时间记录和执行证据，回复老师“该问题属于执行/样本管理偏差，我们已核实事实经过，当前重点说明其对依从性、访视窗或样本可解释性的具体影响范围，并同步提供整改措施和后续防范安排。”"
    if category == "知情同意":
        return "现场回复示例：先准备知情同意书各版本、病历知情记录、签署日期相关页和伦理批准文件，并按时间顺序摆放。依据这些材料的签署时间和版本信息，回复老师“我们优先从受试者权益保护角度说明该问题，现有证据显示的事实顺序是这样的；对于已确认的缺口和补救措施，我们也会如实逐项说明。”"
    return "现场回复示例：先准备原始记录、中心解释、沟通留痕和整改依据，并把涉及数据点定位到具体访视和表单。依据这些材料的可追溯内容，回复老师“该问题我们已经完成受试者级定位，下面按事实、影响范围和已采取措施逐项说明；对仍需补充核实的部分，我们会明确标注并现场补充说明。”"


def normalize_finding_row(row: dict[str, Any]) -> dict[str, Any]:
    category = standardize_problem_category(clean_text(row.get("问题分类")), clean_text(row.get("原始问题描述")))
    severity = standardize_problem_severity(clean_text(row.get("问题严重程度")))
    desc = clean_text(row.get("原始问题描述"))
    impact_primary = impact_from_text(desc, ("IGA", "EASI", "NRS", "主要终点", "第8周", "疗效"))
    impact_key_secondary = impact_from_text(desc, ("EASI75", "NRS", "关键次要", "SCORAD", "DLQI", "CDLQI", "PROMIS"))
    impact_safety = impact_from_text(desc, ("实验室", "AE", "SAE", "SUSAR", "安全", "胆红素", "中性粒", "血"))
    impact_eligibility = impact_from_text(desc, ("入选", "排除", "筛选", "知情", "合格性"))
    impact_data = impact_from_text(desc, ("不一致", "漏记", "漏填", "逻辑", "原始记录", "EDC", "回签", "无法定位"))
    needs_regulatory = "是" if severity == "高" or "可能" in {impact_primary, impact_safety, impact_eligibility} else "需进一步核查"
    risk_level = severity if severity in {"高", "中", "低"} else "待定"
    return {
        "中心": clean_text(row.get("中心")) or MISSING_TEXT,
        "受试者编号": clean_text(row.get("受试者编号")) or MISSING_TEXT,
        "受试者键": build_subject_key({"中心": row.get("中心"), "受试者编号": row.get("受试者编号")}) if clean_text(row.get("受试者编号")) else "",
        "Finding编号": clean_text(row.get("Finding编号")) or MISSING_TEXT,
        "Finding来源文件": clean_text(row.get("Finding来源文件")) or MISSING_TEXT,
        "Finding来源sheet": clean_text(row.get("Finding来源sheet")) or MISSING_TEXT,
        "发现日期": parse_date(row.get("发现日期")) or MISSING_TEXT,
        "问题分类": category,
        "问题严重程度": severity,
        "原始问题描述": desc or MISSING_TEXT,
        "涉及访视": clean_text(row.get("涉及访视")) or MISSING_TEXT,
        "涉及日期": clean_text(row.get("涉及日期")) or MISSING_TEXT,
        "涉及数据点或文件": clean_text(row.get("涉及数据点或文件")) or MISSING_TEXT,
        "问题状态": clean_text(row.get("问题状态")) or MISSING_TEXT,
        "申办方或CRO回复": clean_text(row.get("申办方或CRO回复")) or MISSING_TEXT,
        "中心回复": clean_text(row.get("中心回复")) or MISSING_TEXT,
        "QA意见": clean_text(row.get("QA意见")) or MISSING_TEXT,
        "医学评估意见": clean_text(row.get("医学评估意见")) or build_medical_reply({"问题分类": category, "原始问题描述": desc}),
        "CAPA": clean_text(row.get("CAPA")) or MISSING_TEXT,
        "是否影响受试者权益": impact_from_text(desc, ("知情", "权益", "隐私", "安全")) or MISSING_TEXT,
        "是否影响安全性": impact_safety,
        "是否影响入组合格性": impact_eligibility,
        "是否影响主要终点": impact_primary,
        "是否影响关键次要终点": impact_key_secondary,
        "是否影响实验室安全性判断": impact_from_text(desc, ("实验室", "胆红素", "血", "尿", "CTCAE")),
        "是否影响用药依从性/暴露": impact_from_text(desc, ("用药", "依从性", "间隔", "发药", "回收", "PK")),
        "是否影响数据完整性": impact_data,
        "是否构成重大方案偏离": "可能" if category == "方案偏离" and severity == "高" else "未见直接证据",
        "是否需监管备答": needs_regulatory,
        "建议现场核查回复口径": build_medical_reply({"问题分类": category, "原始问题描述": desc}),
        "现场回复示例": build_reply_example({"问题分类": category, "原始问题描述": desc}),
        "残余风险等级": risk_level,
    }


def build_finding_rows() -> list[dict[str, Any]]:
    all_rows: list[dict[str, Any]] = []
    for spec in FINDING_SHEET_SPECS:
        sheet_name = spec["sheet_name"]
        center_name = spec["center"]
        subject_col = spec["subject_col"]
        if center_name in {"", MISSING_TEXT} or subject_col in {"", MISSING_TEXT}:
            continue
        sheet = read_generic_sheet(FINDING_XLSX, sheet_name)
        counter = 0
        for row in sheet["records"]:
            description = clean_text(row.get("自查问题描述"))
            if not description:
                continue
            subject = extract_subject_id(row.get(subject_col), description)
            counter += 1
            raw = {
                "中心": center_name,
                "受试者编号": subject,
                "Finding编号": f"{'NP' if center_name.startswith('南方') else 'HZ'}-{counter:04d}",
                "Finding来源文件": FINDING_XLSX.name,
                "Finding来源sheet": sheet_name,
                "发现日期": row.get("发现日期"),
                "问题分类": row.get("分类") or row.get("二类分级") or row.get("三类分级"),
                "问题严重程度": row.get("一类分级") or row.get("分类"),
                "原始问题描述": description,
                "涉及访视": row.get("访视"),
                "涉及日期": row.get("发现日期"),
                "涉及数据点或文件": row.get("分类") or row.get("三类分级"),
                "问题状态": row.get("问题状态（关闭/未关闭/无法整改，解释）") or row.get("问题状态（关闭/未关闭/无法整改，解释），CRA审核后回复"),
                "申办方或CRO回复": row.get("PM审核后回复") or row.get("问题确认"),
                "中心回复": row.get("研究者确认回复") or row.get("研究者回复"),
                "QA意见": row.get("自查问题整改措施"),
                "医学评估意见": "",
                "CAPA": row.get("自查问题整改措施"),
            }
            normalized = normalize_finding_row(raw)
            if normalized["受试者编号"] == MISSING_TEXT:
                normalized["受试者编号"] = ""
                normalized["受试者键"] = ""
            all_rows.append(normalized)
    return all_rows


def attach_finding_links(
    subject_profiles: list[dict[str, Any]],
    efficacy_rows: list[dict[str, Any]],
    lab_rows: list[dict[str, Any]],
    finding_rows: list[dict[str, Any]],
    ae_rows: list[dict[str, Any]] | None = None,
    vital_rows: list[dict[str, Any]] | None = None,
) -> None:
    finding_by_subject = defaultdict(list)
    for row in finding_rows:
        if row["受试者编号"]:
            finding_by_subject[row["受试者键"]].append(row)

    for subject in subject_profiles:
        key = subject["受试者键"]
        related = finding_by_subject.get(key, [])
        subject["是否存在核查Finding"] = "是" if related else "否"
        subject["是否存在方案偏离"] = "是" if any(r["问题分类"] == "方案偏离" for r in related) else "否"
        subject["是否存在PD"] = "是" if any(r["问题分类"] == "方案偏离" for r in related) else "否"
        subject["是否存在AE/PD/Finding"] = "；".join(
            [label for label, flag in [("AE", subject.get("是否存在AE") == "是"), ("PD", subject.get("是否存在PD") == "是"), ("Finding", bool(related))] if flag]
        ) or "否"
        subject["是否为重点核查受试者"] = "是" if any(r["残余风险等级"] in {"高", "中"} for r in related) else MISSING_TEXT
        if any(r["残余风险等级"] == "高" for r in related):
            subject["数据完整性风险等级"] = "高"
        elif any(r["残余风险等级"] == "中" for r in related):
            subject["数据完整性风险等级"] = "中"
        elif related:
            subject["数据完整性风险等级"] = "低"

    for row in efficacy_rows:
        related = []
        for finding in finding_by_subject.get(row["受试者键"], []):
            if finding_matches_efficacy(finding, row):
                related.append(finding["Finding编号"])
        if related:
            row["与Finding是否有关联"] = "是"
            row["关联Finding编号"] = "；".join(sorted(set(related)))

    for row in lab_rows:
        related = []
        for finding in finding_by_subject.get(row["受试者键"], []):
            if finding_matches_lab(finding, row):
                related.append(finding["Finding编号"])
        if related:
            row["是否对应AE或Finding"] = "Finding"
            row["备注"] = (row["备注"] + f"；关联Finding: {'；'.join(sorted(set(related)))}").strip("；")
    for row in vital_rows or []:
        related = []
        for finding in finding_by_subject.get(row["受试者键"], []):
            if finding_matches_vital(finding, row):
                related.append(finding["Finding编号"])
        if related:
            row["是否对应AE或Finding"] = "Finding"
            row["备注"] = (row["备注"] + f"；关联Finding: {'；'.join(sorted(set(related)))}").strip("；")


def summarize_lab_attention_by_subject(subject_profiles: list[dict[str, Any]], lab_rows: list[dict[str, Any]]) -> None:
    flags_by_subject: defaultdict[str, set[str]] = defaultdict(set)
    for row in lab_rows:
        for flag in [item for item in clean_text(row.get("额外关注标记")).split("；") if item]:
            flags_by_subject[row["受试者键"]].add(f"{row['检验项目标准化名称']}：{flag}")
    for subject in subject_profiles:
        items = sorted(flags_by_subject.get(subject["受试者键"], set()))
        subject["实验室特殊关注"] = "无" if not items else "；".join(items[:12])
        subject["是否存在实验室特殊关注"] = "否" if not items else "是"


def summarize_vital_attention_by_subject(subject_profiles: list[dict[str, Any]], vital_rows: list[dict[str, Any]]) -> None:
    flags_by_subject: defaultdict[str, set[str]] = defaultdict(set)
    for row in vital_rows:
        for flag in [item for item in clean_text(row.get("额外关注标记")).split("；") if item]:
            flags_by_subject[row["受试者键"]].add(f"{row['生命体征指标']}：{flag}")
    for subject in subject_profiles:
        items = sorted(flags_by_subject.get(subject["受试者键"], set()))
        subject["生命体征特殊关注"] = "无" if not items else "；".join(items[:12])
        subject["是否存在生命体征特殊关注"] = "否" if not items else "是"


def build_center_aggregate_payload(
    subject_profiles: list[dict[str, Any]],
    efficacy_rows: list[dict[str, Any]],
    lab_rows: list[dict[str, Any]],
    vital_rows: list[dict[str, Any]],
) -> dict[str, Any]:
    payload: dict[str, Any] = {}
    subject_group_map = {clean_text(row.get("受试者键")): clean_text(row.get("随机组别")) for row in subject_profiles}
    binary_configs = [
        (
            item["metric"],
            item["label"],
            lambda row, response_name=item["label"]: response_label_matches(row.get("关键应答判定"), response_name),
        )
        for item in PROTOCOL_SUMMARY.get("response_rules", [])
    ]
    for center in TARGET_CENTERS:
        denominator_by_group = {
            "试验组": sum(1 for row in subject_profiles if clean_text(row.get("中心")) == center and clean_text(row.get("受试者状态")) != "筛选失败" and summary_group_label(row.get("随机组别")) == "试验组"),
            "对照组": sum(1 for row in subject_profiles if clean_text(row.get("中心")) == center and clean_text(row.get("受试者状态")) != "筛选失败" and summary_group_label(row.get("随机组别")) == "对照组"),
        }
        efficacy_metrics = sorted({clean_text(row.get("指标名称")) for row in efficacy_rows if clean_text(row.get("中心")) == center}, key=efficacy_sort_key)
        continuous_efficacy_metrics = [metric for metric in efficacy_metrics if selected_metric_variable_type(metric) != "binary"]
        lab_metrics = sorted(
            {
                clean_text(row.get("检验项目标准化名称"))
                for row in lab_rows
                if clean_text(row.get("中心")) == center and isinstance(row.get("数值结果"), (int, float))
            },
            key=lambda name: (LAB_METRIC_ORDER.get(name, 999), name),
        )
        vital_metrics = sorted({clean_text(row.get("生命体征指标")) for row in vital_rows if clean_text(row.get("中心")) == center}, key=vital_sort_key)
        payload[center] = {
            "profile_summary": build_center_profile_summary(subject_profiles, center),
            "efficacy_continuous": [
                {
                    "metric": metric,
                    "rows": aggregate_continuous_by_visit_group(
                        efficacy_rows,
                        center=center,
                        metric_key="指标名称",
                        metric_value=metric,
                        subject_group_map=subject_group_map,
                        denominator_by_group=denominator_by_group,
                    ),
                }
                for metric in continuous_efficacy_metrics
            ],
            "efficacy_binary": [
                {
                    "metric": metric,
                    "response_name": response_name,
                    "rows": aggregate_binary_by_visit_group(
                        efficacy_rows,
                        center=center,
                        metric_key="指标名称",
                        metric_value=metric,
                        response_name=response_name,
                        matcher=matcher,
                        subject_group_map=subject_group_map,
                        denominator_by_group=denominator_by_group,
                    ),
                }
                for metric, response_name, matcher in binary_configs
                if metric in efficacy_metrics
            ],
            "lab_continuous": [
                {
                    "metric": metric,
                    "rows": aggregate_continuous_by_visit_group(
                        lab_rows,
                        center=center,
                        metric_key="检验项目标准化名称",
                        metric_value=metric,
                        subject_group_map=subject_group_map,
                        denominator_by_group=denominator_by_group,
                    ),
                }
                for metric in lab_metrics
            ],
            "vital_continuous": [
                {
                    "metric": metric,
                    "rows": aggregate_continuous_by_visit_group(
                        vital_rows,
                        center=center,
                        metric_key="生命体征指标",
                        metric_value=metric,
                        subject_group_map=subject_group_map,
                        denominator_by_group=denominator_by_group,
                    ),
                }
                for metric in vital_metrics
            ],
        }
    return payload


def run_qc_checks(
    subject_rows: list[dict[str, Any]],
    efficacy_rows: list[dict[str, Any]],
    lab_rows: list[dict[str, Any]],
    finding_rows: list[dict[str, Any]],
) -> dict[str, list[dict[str, Any]]]:
    summary: list[dict[str, Any]] = []
    issues: list[dict[str, Any]] = []
    completeness: list[dict[str, Any]] = []

    subject_ids = {r["受试者编号"] for r in subject_rows if r["受试者编号"]}
    efficacy_subjects = {r["受试者编号"] for r in efficacy_rows if r["受试者编号"]}
    lab_subjects = {r["受试者编号"] for r in lab_rows if r["受试者编号"]}
    finding_subjects = {r["受试者编号"] for r in finding_rows if r["受试者编号"]}

    def add_summary(name: str, passed: bool, count: int, note: str) -> None:
        summary.append({"QC项目": name, "结果": "通过" if passed else "需关注", "问题数": count, "说明": note})

    # 1 consistency
    missing_subject_match = sorted((efficacy_subjects | lab_subjects | finding_subjects) - subject_ids)
    for subject in missing_subject_match:
        issues.append({"issue_id": f"QC-{len(issues)+1:04d}", "issue_type": "受试者缺失", "severity": "高", "center": "", "subject": subject, "sheet": "", "detail": "相关数据集中存在，但受试者主表不存在"})
    add_summary("受试者编号是否一致", len(missing_subject_match) == 0, len(missing_subject_match), "检查主表、疗效、实验室和Finding之间的受试者编号一致性")

    # 2 center consistency
    invalid_centers = [r for r in subject_rows if r["中心"] not in TARGET_CENTERS]
    add_summary("中心名称是否一致", len(invalid_centers) == 0, len(invalid_centers), "限定为两家目标中心")

    # 3 duplicate cross-center
    subject_center_count = defaultdict(set)
    for row in subject_rows:
        subject_center_count[row["受试者编号"]].add(row["中心"])
    cross_center = [s for s, centers in subject_center_count.items() if len(centers) > 1]
    add_summary("同一受试者是否跨中心重复", len(cross_center) == 0, len(cross_center), "受试者编号不应跨中心重复")

    # 4 visit standardization
    unstd_visits = [r for r in efficacy_rows if r.get("访视编号", MISSING_TEXT) == MISSING_TEXT or r.get("访视名称", MISSING_TEXT) == MISSING_TEXT]
    add_summary("访视名称是否标准化", len(unstd_visits) == 0, len(unstd_visits), "检查疗效数据访视名称与访视编号")

    # 5 dates parse
    bad_dates = [r for r in efficacy_rows + lab_rows if r.get("日期") and not re.match(r"\d{4}-\d{2}-\d{2}", r.get("日期", ""))]
    add_summary("日期是否可解析", len(bad_dates) == 0, len(bad_dates), "非标准日期保留并提示人工复核")

    # 6 duplicates
    dup_counter = Counter((r.get("受试者键", ""), r.get("指标名称", ""), r.get("访视编号", "")) for r in efficacy_rows)
    duplicate_metrics = [k for k, v in dup_counter.items() if v > 1]
    add_summary("同一指标同一访视是否重复", len(duplicate_metrics) == 0, len(duplicate_metrics), "疗效指标按受试者/指标/访视去重检查")

    # 7 missing results
    missing_results = [r for r in efficacy_rows if r.get("原始值", MISSING_TEXT) == MISSING_TEXT] + [r for r in lab_rows if r.get("结果值", MISSING_TEXT) == MISSING_TEXT]
    for row in missing_results:
        issues.append({
            "issue_id": f"QC-{len(issues)+1:04d}",
            "issue_type": "结果值缺失",
            "severity": "中",
            "center": row.get("中心", ""),
            "subject": row.get("受试者编号", ""),
            "sheet": row.get("数据来源sheet", ""),
            "detail": f"{row.get('指标名称') or row.get('检验项目标准化名称') or '未识别字段'} 在 {row.get('访视名称','')} 缺失结果值",
        })
    add_summary("是否存在结果值缺失", len(missing_results) == 0, len(missing_results), "缺失值不填0")

    # 8 units inconsistent
    unit_counter = defaultdict(set)
    for row in lab_rows:
        unit_counter[row.get("检验项目标准化名称", "")].add(row.get("单位", ""))
    unit_issues = [metric for metric, units in unit_counter.items() if len({u for u in units if u and u != MISSING_TEXT}) > 1]
    for metric in unit_issues:
        issues.append({
            "issue_id": f"QC-{len(issues)+1:04d}",
            "issue_type": "单位不一致",
            "severity": "中",
            "center": "",
            "subject": "",
            "sheet": "实验室检查",
            "detail": f"{metric} 存在多个单位：{' / '.join(sorted(unit_counter[metric]))}",
        })
    add_summary("是否存在单位不一致", len(unit_issues) == 0, len(unit_issues), "按单项实验室指标汇总单位")

    # 9/10 normal range
    missing_ranges = [r for r in lab_rows if not clean_text(r.get("正常值下限")) or not clean_text(r.get("正常值上限"))]
    for row in missing_ranges:
        issues.append({
            "issue_id": f"QC-{len(issues)+1:04d}",
            "issue_type": "正常范围缺失",
            "severity": "中",
            "center": row.get("中心", ""),
            "subject": row.get("受试者编号", ""),
            "sheet": row.get("数据来源sheet", ""),
            "detail": f"{row.get('检验项目标准化名称','')} 在 {row.get('访视名称','')} 缺少正常值上下限",
        })
    add_summary("是否存在正常值范围缺失", len(missing_ranges) == 0, len(missing_ranges), "正常范围缺失时不能判断异常")
    add_summary("是否存在正常值上下限无法解析", all(to_float(r.get("正常值下限")) is not None and to_float(r.get("正常值上限")) is not None for r in lab_rows if clean_text(r.get("正常值下限")) and clean_text(r.get("正常值上限"))), sum(1 for r in lab_rows if clean_text(r.get("正常值下限")) and clean_text(r.get("正常值上限")) and (to_float(r.get("正常值下限")) is None or to_float(r.get("正常值上限")) is None)), "数值上下限解析失败时保留文本")

    # 11 abnormal date
    subject_by_key = {}
    for r in subject_rows:
        key = r.get("受试者键") or build_subject_key({"中心": r.get("中心", ""), "受试者编号": r.get("受试者编号", "")})
        if key:
            subject_by_key[key] = r
    abnormal_chronology = []
    for row in efficacy_rows + lab_rows:
        row_key = row.get("受试者键") or build_subject_key({"中心": row.get("中心", ""), "受试者编号": row.get("受试者编号", "")})
        subject = subject_by_key.get(row_key)
        if not subject:
            continue
        icf = subject.get("知情同意日期", "")
        current_date = row.get("日期", "")
        if re.match(r"\d{4}-\d{2}-\d{2}", icf or "") and re.match(r"\d{4}-\d{2}-\d{2}", current_date or "") and current_date < icf:
            abnormal_chronology.append(row)
    add_summary("是否存在明显异常日期，例如访视日期早于知情同意日期", len(abnormal_chronology) == 0, len(abnormal_chronology), "仅做可直接比较的日期逻辑检查")

    # 12 finding visit match
    unmatched_finding_visit = [r for r in finding_rows if r.get("涉及访视", MISSING_TEXT) != MISSING_TEXT and not any(r.get("涉及访视", "") in e.get("访视名称", "") or r.get("涉及访视", "") in e.get("访视编号", "") for e in efficacy_rows if e.get("受试者键", "") == r.get("受试者键", ""))]
    for row in unmatched_finding_visit:
        issues.append({
            "issue_id": f"QC-{len(issues)+1:04d}",
            "issue_type": "Finding访视无法匹配",
            "severity": "中",
            "center": row.get("中心", ""),
            "subject": row.get("受试者编号", ""),
            "sheet": row.get("Finding来源sheet", ""),
            "detail": f"Finding {row.get('Finding编号','')} 的涉及访视“{row.get('涉及访视','')}”未在疗效访视中匹配",
        })
    add_summary("是否存在疗效数据与Finding中涉及访视无法匹配", len(unmatched_finding_visit) == 0, len(unmatched_finding_visit), "Finding访视文本与疗效访视比对")

    # 13 abnormal lab without assessment
    abnormal_without_assessment = [r for r in lab_rows if r.get("异常方向") in {"高", "低"} and r.get("研究者临床评估", MISSING_TEXT) == MISSING_TEXT]
    for row in abnormal_without_assessment:
        issues.append({
            "issue_id": f"QC-{len(issues)+1:04d}",
            "issue_type": "实验室异常缺少临床评估",
            "severity": "中",
            "center": row.get("中心", ""),
            "subject": row.get("受试者编号", ""),
            "sheet": row.get("数据来源sheet", ""),
            "detail": f"{row.get('检验项目标准化名称','')} {row.get('结果值','')} {row.get('单位','')} 为{row.get('异常方向','')}，但未见研究者临床评估",
        })
    add_summary("是否存在实验室异常但无临床评估", len(abnormal_without_assessment) == 0, len(abnormal_without_assessment), "当前源文件中多数未直接提供研究者评估字段")

    # 14 finding subjects not in listing
    missing_finding_subjects = [r for r in finding_rows if r.get("受试者编号") and r.get("受试者编号") not in subject_ids]
    add_summary("是否存在Finding提到的受试者编号在listing中不存在", len(missing_finding_subjects) == 0, len(missing_finding_subjects), "Finding中受试者编号与主表对照")

    # 15 finding not located to data point
    unlocated_findings = [r for r in finding_rows if r.get("受试者编号") and r.get("涉及访视", MISSING_TEXT) == MISSING_TEXT and r.get("涉及数据点或文件", MISSING_TEXT) == MISSING_TEXT]
    for row in unlocated_findings:
        issues.append({
            "issue_id": f"QC-{len(issues)+1:04d}",
            "issue_type": "Finding无法定位数据点",
            "severity": "中",
            "center": row.get("中心", ""),
            "subject": row.get("受试者编号", ""),
            "sheet": row.get("Finding来源sheet", ""),
            "detail": f"Finding {row.get('Finding编号','')} 缺少访视或数据点定位信息",
        })
    add_summary("是否存在Finding提到的问题无法定位到对应数据点", len(unlocated_findings) == 0, len(unlocated_findings), "缺少访视或数据点定位信息")

    # 16 adult/child scale mismatch
    age_lookup = {
        (r.get("受试者键") or build_subject_key({"中心": r.get("中心", ""), "受试者编号": r.get("受试者编号", "")})): r.get("年龄组", "")
        for r in subject_rows
    }
    scale_mismatch = []
    for row in efficacy_rows:
        age_group = age_lookup.get(row.get("受试者键", ""), "")
        if age_group == "成人" and row.get("指标名称") == "CDLQI":
            scale_mismatch.append(row)
        if age_group == "青少年" and row.get("指标名称") == "DLQI":
            scale_mismatch.append(row)
    for row in scale_mismatch:
        issues.append({
            "issue_id": f"QC-{len(issues)+1:04d}",
            "issue_type": "年龄组与量表不匹配",
            "severity": "中",
            "center": row.get("中心", ""),
            "subject": row.get("受试者编号", ""),
            "sheet": row.get("数据来源sheet", ""),
            "detail": f"{row.get('年龄组','未识别年龄组')} 使用了 {row.get('指标名称','')}，需人工确认量表适用性",
        })
    add_summary("是否存在儿童/成人量表使用不一致，例如成人使用CDLQI或儿童使用DLQI", len(scale_mismatch) == 0, len(scale_mismatch), "按年龄组和量表名称比对")

    # 17 SCORAD source
    preferred_scorad_sheet = next((cfg["sheet"] for cfg in EFFICACY_CONFIG if cfg.get("metric") == "SCORAD"), "")
    scorad_source_bad = [r for r in efficacy_rows if r.get("指标名称") == "SCORAD" and preferred_scorad_sheet and r.get("数据来源sheet") != preferred_scorad_sheet]
    add_summary("是否存在SCORAD总分来源不符合要求", len(scorad_source_bad) == 0, len(scorad_source_bad), "SCORAD总分优先来源固定为SR sheet")

    # 18 qualitative urine numeric
    urine_numeric = [r for r in lab_rows if r.get("实验室类别") == "尿常规" and any(k in r.get("结果值", "") for k in ["+", "阴性", "阳性"]) and r.get("图表模式") == "numeric"]
    for row in urine_numeric:
        issues.append({
            "issue_id": f"QC-{len(issues)+1:04d}",
            "issue_type": "定性尿常规疑似被数值化",
            "severity": "中",
            "center": row.get("中心", ""),
            "subject": row.get("受试者编号", ""),
            "sheet": row.get("数据来源sheet", ""),
            "detail": f"{row.get('检验项目标准化名称','')} 结果为 {row.get('结果值','')}，但图表模式为 numeric",
        })
    add_summary("定性尿常规被错误数值化", len(urine_numeric) == 0, len(urine_numeric), "定性尿常规应使用分类展示")

    for subject in subject_rows:
        key = subject.get("受试者键") or build_subject_key({"中心": subject.get("中心", ""), "受试者编号": subject.get("受试者编号", "")})
        eff_count = sum(1 for r in efficacy_rows if r.get("受试者键", "") == key)
        lab_count = sum(1 for r in lab_rows if r.get("受试者键", "") == key)
        finding_count = sum(1 for r in finding_rows if r.get("受试者键", "") == key)
        completeness.append(
            {
                "中心": subject["中心"],
                "受试者编号": subject["受试者编号"],
                "基本信息完整度": "高" if all(subject.get(field) and subject.get(field) != MISSING_TEXT for field in ["年龄", "性别", "筛选日期", "知情同意日期", "基线/随机日期"]) else "中/低",
                "疗效记录数": eff_count,
                "实验室记录数": lab_count,
                "Finding数": finding_count,
                "缺口提示": "无明显缺口" if eff_count and lab_count else "需人工复核资料完整性",
            }
        )

    return {"summary": summary, "issues": issues, "completeness": completeness}


def write_csv(path: Path, rows: list[dict[str, Any]]) -> None:
    path.parent.mkdir(parents=True, exist_ok=True)
    if not rows:
        rows = [{}]
    fieldnames: list[str] = []
    for row in rows:
        for key in row.keys():
            if key not in fieldnames:
                fieldnames.append(key)
    with path.open("w", encoding="utf-8-sig", newline="") as f:
        writer = csv.DictWriter(f, fieldnames=fieldnames)
        writer.writeheader()
        for row in rows:
            writer.writerow({key: clean_text(value) for key, value in row.items()})


def build_protocol_summary_markdown(protocol_summary: dict[str, Any]) -> str:
    lines = ["# 方案终点与研究流程解构", ""]
    files = protocol_summary.get("files", [])
    lines.append("## 已读取方案文件")
    if files:
        for item in files:
            lines.append(f"- {item}")
    else:
        lines.append("- 未读取到方案文件。")
    lines.extend(["", "## 识别到的疗效指标"])
    metrics = protocol_summary.get("selected_metrics", [])
    if metrics:
        for item in metrics:
            lines.append(f"- {item['metric']}：{item.get('endpoint_role', '未分类')}；变量类型 {item.get('variable_type', '需确认')}。证据：{item.get('source_excerpt', '')}")
    else:
        lines.append("- 未识别到明确疗效指标。")
    lines.extend(["", "## 识别到的二分类应答规则"])
    response_rules = protocol_summary.get("response_rules", [])
    if response_rules:
        for item in response_rules:
            lines.append(f"- {item['label']}（来源指标 {item['metric']}）：{item.get('endpoint_role', '未分类')}。证据：{item.get('source_excerpt', '')}")
    else:
        lines.append("- 未识别到明确的二分类应答规则。")
    lines.extend(["", "## 研究流程表/访视提及"])
    visit_mentions = protocol_summary.get("visit_mentions", [])
    if visit_mentions:
        lines.append(f"- {'；'.join(visit_mentions[:30])}")
    else:
        lines.append("- 未识别到明确访视提及。")
    return "\n".join(lines) + "\n"


def selected_response_rules_for_metric(metric: str) -> list[dict[str, Any]]:
    return [item for item in PROTOCOL_SUMMARY.get("response_rules", []) if item.get("metric") == metric]


def selected_metric_variable_type(metric: str) -> str:
    for item in PROTOCOL_SUMMARY.get("selected_metrics", []):
        if item.get("metric") == metric:
            return item.get("variable_type", "continuous")
    return "continuous"


def build_unresolved_questions(field_mapping: list[dict[str, Any]], qc: dict[str, Any], subjects: list[dict[str, Any]], findings: list[dict[str, Any]]) -> str:
    unresolved = []
    unresolved.append("# 未解决数据问题\n")
    unresolved.append("以下项目需人工复核或补充源文件后再形成最终核查口径。\n")
    unresolved.append("## 字段映射\n")
    unknown_fields = [r for r in field_mapping if r["是否需要人工复核"] == "是"]
    for row in unknown_fields[:30]:
        unresolved.append(f"- [{row['sheet名']}] `{row['原始字段名']}` 的标准化映射需人工确认。")
    unresolved.append("\n## 质控待关注项\n")
    for row in qc["summary"]:
        if row["结果"] != "通过":
            unresolved.append(f"- {row['QC项目']}：{row['说明']}（问题数 {row['问题数']}）。")
    unresolved.append("\n## 受试者级关注点\n")
    for subject in subjects:
        if subject["数据完整性风险等级"] in {"高", "中"}:
            unresolved.append(f"- {subject['中心']} {subject['受试者编号']}：当前数据完整性风险等级为{subject['数据完整性风险等级']}。")
    unresolved.append("\n## Finding 需补充证据\n")
    for row in findings[:20]:
        if row["是否需监管备答"] != "否":
            unresolved.append(f"- {row['受试者编号'] or '未定位受试者'} / {row['Finding编号']}：{row['建议现场核查回复口径']}")
    return "\n".join(unresolved).strip() + "\n"


def render_html(payload: dict[str, Any]) -> str:
    data_json = json.dumps(payload, ensure_ascii=False)
    return f"""<!doctype html>
<html lang="zh-CN">
<head>
<meta charset="utf-8">
<meta name="viewport" content="width=device-width, initial-scale=1">
<title>{HTML_TITLE or "受试者Patient Profile"}</title>
<style>
:root {{
  --brand: #FF9900;
  --brand-2: #FFCC00;
  --bg: #f7f5ef;
  --card: #ffffff;
  --text: #232323;
  --muted: #6d6d6d;
  --border: #e6ddcc;
  --danger: #d4380d;
  --info: #3b6ea8;
}}
* {{ box-sizing: border-box; }}
body {{ margin: 0; font-family: -apple-system, BlinkMacSystemFont, "PingFang SC", "Microsoft YaHei", sans-serif; background: var(--bg); color: var(--text); }}
header {{ position: sticky; top: 0; z-index: 20; background: #fff; border-bottom: 1px solid var(--border); padding: 14px 18px; }}
.title {{ font-size: 22px; font-weight: 700; }}
.subtitle {{ color: var(--muted); font-size: 13px; margin-top: 4px; }}
.toolbar.compact {{ display: grid; grid-template-columns: 1.1fr 1fr 1fr 1fr 1fr 1fr 1.1fr; gap: 10px; margin-top: 14px; align-items: end; }}
.toolbar .control {{ display: grid; gap: 4px; min-width: 0; }}
.toolbar label {{ font-size: 12px; color: #8b5c00; font-weight: 700; }}
select, input {{ width: 100%; border: 1px solid var(--border); border-radius: 10px; padding: 8px 10px; background: #fff; font-size: 13px; }}
main {{ padding: 18px; display: grid; gap: 18px; }}
.panel {{ background: var(--card); border: 1px solid var(--border); border-radius: 16px; padding: 16px; box-shadow: 0 4px 14px rgba(0,0,0,0.04); }}
.panel h2 {{ margin: 0 0 10px; font-size: 18px; }}
.summary-line {{ color: var(--muted); font-size: 13px; }}
.subject-card {{ border: 1px solid var(--border); border-radius: 14px; padding: 14px; margin-top: 16px; background: #fffdf9; }}
.subject-head {{ display: flex; justify-content: space-between; gap: 12px; align-items: flex-start; flex-wrap: wrap; }}
.subject-head h3 {{ margin: 0; font-size: 21px; }}
.pill {{ display: inline-block; padding: 4px 9px; border-radius: 999px; font-size: 12px; font-weight: 700; background: #eef4ff; color: var(--info); margin-right: 6px; margin-bottom: 6px; }}
.pill.high {{ background: #ffe7e1; color: var(--danger); }}
.pill.mid {{ background: #fff3cf; color: #8a5a00; }}
.pill.low {{ background: #eef4ff; color: var(--info); }}
.pill.adolescent {{ background: #fff0c4; color: #8a5a00; }}
.flag-badge {{ display: inline-block; margin: 2px 6px 2px 0; padding: 3px 8px; border-radius: 999px; font-size: 11px; font-weight: 700; line-height: 1.4; }}
.flag-badge.attn {{ background: #ffe3d7; color: #a61d00; border: 1px solid #ffb59e; }}
.flag-badge.watch {{ background: #fff3cf; color: #8a5a00; border: 1px solid #f0d27a; }}
.flag-badge.info {{ background: #eaf1fb; color: #245b97; border: 1px solid #bdd3ef; }}
.flag-badge.good {{ background: #eef8ef; color: #1f6b30; border: 1px solid #b8dfc1; }}
.grid {{ display: grid; grid-template-columns: repeat(4, minmax(0, 1fr)); gap: 10px; margin-top: 12px; }}
.field {{ border: 1px solid var(--border); border-radius: 12px; padding: 10px; background: #fff; }}
.field .k {{ font-size: 12px; color: #8a5a00; margin-bottom: 4px; }}
.field .v {{ font-size: 14px; font-weight: 700; }}
.module-title {{ font-size: 16px; font-weight: 700; margin: 16px 0 8px; color: #8a5a00; }}
.chart-box {{ border: 1px solid var(--border); border-radius: 12px; background: #fff; padding: 10px; margin-bottom: 12px; overflow: auto; }}
.chart-box svg {{ width: 100%; min-width: 860px; height: 320px; display: block; }}
.chart-caption {{ font-size: 12px; color: var(--muted); margin-top: 6px; }}
.flow {{ display: grid; grid-template-columns: repeat(3, minmax(0, 1fr)); gap: 12px; margin-top: 12px; }}
.flow-step {{ border: 1px solid var(--border); border-radius: 14px; padding: 14px; background: linear-gradient(180deg, #fff8ea, #fff); text-align: center; }}
.flow-step .num {{ font-size: 28px; font-weight: 800; color: #8a5a00; }}
.flow-step .label {{ font-size: 13px; color: var(--muted); margin-top: 4px; }}
.table-wrap {{ overflow: auto; border: 1px solid var(--border); border-radius: 12px; background: #fff; }}
table {{ width: 100%; border-collapse: collapse; }}
th, td {{ border: 1px solid #efe8da; padding: 8px 10px; font-size: 12px; text-align: left; vertical-align: top; word-break: break-word; }}
th {{ background: #fff7e7; cursor: pointer; position: sticky; top: 0; }}
.finding-fixed {{ border-top: 2px solid var(--brand); padding-top: 8px; }}
details {{ border: 1px solid var(--border); border-radius: 12px; padding: 10px; background: #fff; margin-top: 10px; }}
summary {{ cursor: pointer; font-weight: 700; }}
pre {{ white-space: pre-wrap; word-break: break-word; background: #fcfbf7; border: 1px solid var(--border); border-radius: 12px; padding: 12px; }}
@media (max-width: 1500px) {{ .toolbar.compact {{ grid-template-columns: repeat(4, minmax(0, 1fr)); }} }}
@media (max-width: 900px) {{ .toolbar.compact, .grid {{ grid-template-columns: 1fr 1fr; }} }}
@media (max-width: 640px) {{ .toolbar.compact, .grid {{ grid-template-columns: 1fr; }} }}
</style>
</head>
<body>
<header>
  <div class="title">{HTML_TITLE or "受试者Patient Profile"}</div>
  <div class="subtitle">疗效、实验室、生命体征与 Finding 的中文交互式核查视图</div>
  <div class="toolbar compact">
    <div class="control"><label>中心</label><select id="centerSelect"></select></div>
    <div class="control"><label>受试者</label><select id="subjectSelect"></select></div>
    <div class="control"><label>疗效指标</label><select id="efficacyMetricSelect"></select></div>
    <div class="control"><label>实验室指标</label><select id="labMetricSelect"></select></div>
    <div class="control"><label>生命体征</label><select id="vitalMetricSelect"></select></div>
    <div class="control"><label>异常值</label><select id="abnormalFilter"><option value="all">全部</option><option value="low">仅低值</option><option value="high">仅高值</option><option value="comment">仅有评估/备注</option><option value="finding">仅Finding关联</option></select></div>
    <div class="control"><label>Finding筛选</label><select id="findingFilter"><option value="all">全部Finding</option><option value="high">高风险</option><option value="mid">中风险</option><option value="primary">影响主要终点</option><option value="safety">影响安全性</option><option value="data">影响数据完整性</option><option value="regulatory">需监管备答</option></select></div>
    <div class="control"><label>关键词</label><input id="globalSearch" placeholder="受试者/指标/Finding"></div>
  </div>
</header>
<main>
  <section class="panel">
    <h2>当前筛选摘要</h2>
    <div id="currentSummary" class="summary-line"></div>
  </section>
  <section class="panel" id="profilesPanel"></section>
  <section class="panel">
    <details>
      <summary>数据源与字段映射附录</summary>
      <div id="mappingTable"></div>
    </details>
  </section>
</main>
<script>
const DATA = {data_json};

function escapeHtml(value) {{
  return String(value ?? "").replace(/&/g, "&amp;").replace(/</g, "&lt;").replace(/>/g, "&gt;").replace(/"/g, "&quot;");
}}
function unique(arr) {{ return Array.from(new Set(arr.filter(Boolean))); }}
function setOptions(id, options, selected) {{
  const el = document.getElementById(id);
  el.innerHTML = options.map(opt => `<option value="${{escapeHtml(opt.value)}}">${{escapeHtml(opt.label)}}</option>`).join("");
  el.value = selected;
}}
function getValue(id) {{ return document.getElementById(id).value; }}
function cellBadgeClass(text) {{
  if (/^未达到/.test(text)) return "watch";
  if (/≥50%|达到180|达到160|达到110|达到100|高风险|异常有临床意义|高/.test(text)) return "attn";
  if (/相近|更多但评估为异常无临床意义|中风险|低|Finding|AE/.test(text)) return "watch";
  if (/EASI75|EASI90|IGA-TS|正常/.test(text)) return "good";
  return "info";
}}
function renderCell(header, value) {{
  const raw = value ?? "";
  const text = String(raw);
  if (!text) return "";
  if (["额外关注标记","关键应答判定","异常方向","是否对应AE或Finding","是否存在AE/PD/Finding"].includes(header)) {{
    return text.split("；").filter(Boolean).map(part => `<span class="flag-badge ${{cellBadgeClass(part)}}">${{escapeHtml(part)}}</span>`).join("");
  }}
  return escapeHtml(text);
}}
function efficacyHeaders(metric) {{
  const headers = ["访视编号","访视名称","日期","原始值","较基线变化","较基线百分比变化"];
  if (["EASI","IGA","NRS","QSI-PROMIS睡眠相关影响8a","QSI-PROMIS睡眠困扰8b"].includes(metric)) {{
    headers.push("关键应答判定");
  }}
  headers.push("问题标记","关联Finding编号","数据来源sheet");
  return headers;
}}
function makeTable(headers, rows, tableId) {{
  const head = headers.map(h => `<th data-key="${{escapeHtml(h)}}">${{escapeHtml(h)}}</th>`).join("");
  const body = rows.map(row => `<tr>${{headers.map(h => `<td>${{renderCell(h, row[h])}}</td>`).join("")}}</tr>`).join("");
  return `<div class="table-wrap"><table id="${{tableId}}"><thead><tr>${{head}}</tr></thead><tbody>${{body}}</tbody></table></div>`;
}}
function makeGroupedTable(headerRows, leafHeaders, rows, tableId) {{
  const head = headerRows.map(row => `<tr>${{row.map(cell => {{
    const attrs = [];
    if (cell.colspan) attrs.push(`colspan="${{cell.colspan}}"`);
    if (cell.rowspan) attrs.push(`rowspan="${{cell.rowspan}}"`);
    if (cell.key) attrs.push(`data-key="${{escapeHtml(cell.key)}}"`);
    if (cell.colIndex !== undefined) attrs.push(`data-col-index="${{cell.colIndex}}"`);
    return `<th ${{attrs.join(" ")}}>${{escapeHtml(cell.label)}}</th>`;
  }}).join("")}}</tr>`).join("");
  const body = rows.map(row => `<tr>${{leafHeaders.map(col => {{
    const value = typeof col.render === "function" ? col.render(row) : row[col.key];
    return `<td>${{renderCell(col.outputKey || col.key, value)}}</td>`;
  }}).join("")}}</tr>`).join("");
  return `<div class="table-wrap"><table id="${{tableId}}"><thead>${{head}}</thead><tbody>${{body}}</tbody></table></div>`;
}}
function filterSubjectsByCenter(center) {{
  return center === "__ALL__" ? DATA.subjects : DATA.subjects.filter(s => s["中心"] === center);
}}
function initSelectors() {{
  const firstSubject = DATA.subjects[0];
  const firstCenter = firstSubject ? firstSubject["中心"] : "__ALL__";
  setOptions("centerSelect", [{{value:"__ALL__", label:"全部中心"}}, ...DATA.centers.map(c => ({{value:c,label:c}}))], firstCenter);
  syncSubjectOptions("__ALL__");
  setOptions("efficacyMetricSelect", [{{value:"__ALL__", label:"全部疗效指标"}}, {{value:"__NONE__", label:"不显示疗效"}}, ...unique(DATA.efficacy.map(r => r["指标名称"])).map(v => ({{value:v,label:v}}))], "__ALL__");
  setOptions("labMetricSelect", [{{value:"__ALL__", label:"全部实验室指标"}}, {{value:"__NONE__", label:"不显示实验室"}}, ...unique(DATA.labs.map(r => r["检验项目标准化名称"])).map(v => ({{value:v,label:v}}))], "__ALL__");
  setOptions("vitalMetricSelect", [{{value:"__ALL__", label:"全部生命体征"}}, {{value:"__NONE__", label:"不显示生命体征"}}, ...unique(DATA.vitals.map(r => r["生命体征指标"])).map(v => ({{value:v,label:v}}))], "__ALL__");
}}
function syncSubjectOptions(preferred) {{
  const center = getValue("centerSelect");
  const subjects = filterSubjectsByCenter(center);
  const options = [{{value:"__ALL__", label:"全部受试者"}}].concat(subjects.map(s => ({{value:s["受试者编号"], label:`${{s["受试者编号"]}} / ${{s["中心"]}}`}})));
  const selected = preferred && options.some(o => o.value === preferred) ? preferred : (subjects[0] ? subjects[0]["受试者编号"] : "__ALL__");
  setOptions("subjectSelect", options, selected);
}}
function currentFilters() {{
  return {{
    center: getValue("centerSelect"),
    subject: getValue("subjectSelect"),
    efficacyMetric: getValue("efficacyMetricSelect"),
    labMetric: getValue("labMetricSelect"),
    vitalMetric: getValue("vitalMetricSelect"),
    abnormal: getValue("abnormalFilter"),
    finding: getValue("findingFilter"),
    search: document.getElementById("globalSearch").value.trim().toLowerCase()
  }};
}}
function findingPass(row, filterValue) {{
  if (filterValue === "all") return true;
  if (filterValue === "high") return row["残余风险等级"] === "高";
  if (filterValue === "mid") return row["残余风险等级"] === "中";
  if (filterValue === "primary") return row["是否影响主要终点"] === "可能";
  if (filterValue === "safety") return row["是否影响安全性"] === "可能";
  if (filterValue === "data") return row["是否影响数据完整性"] === "可能";
  if (filterValue === "regulatory") return row["是否需监管备答"] !== "否";
  return true;
}}
function labPass(row, abnormal) {{
  if (abnormal === "all") return true;
  if (abnormal === "low") return row["异常方向"] === "低";
  if (abnormal === "high") return row["异常方向"] === "高";
  if (abnormal === "comment") return !!((row["研究者临床评估"] || "") || (row["备注"] || ""));
  if (abnormal === "finding") return row["是否对应AE或Finding"] !== "否";
  return true;
}}
function shouldRenderSeriesChart(rows, seriesType, metricName) {{
  if (!rows.length) return false;
  const numericRows = rows.filter(r => r["图表模式"] !== "categorical" && r["数值结果"] !== null && r["数值结果"] !== "" && !Number.isNaN(Number(r["数值结果"])));
  if (!numericRows.length) return false;
  if ((seriesType === "lab" || seriesType === "vital") && (rows[0]["访视编号"] === "UNS" || (rows[0]["访视名称"] || "").includes("计划外"))) {{
    return false;
  }}
  if (seriesType === "lab" && rows[0]["实验室类别"] === "尿常规") {{
    return rows.some(r => r["访视编号"] === "UNS" || (r["访视名称"] || "").includes("计划外"));
  }}
  if (seriesType === "lab" && metricName === "QTc间期" && numericRows.length <= 1) {{
    return false;
  }}
  return true;
}}
function buildYAxis(left, right, top, bottom, width, height, yMin, yMax, tickCount) {{
  let svg = "";
  for (let i = 0; i <= tickCount; i += 1) {{
    const value = yMin + (yMax - yMin) * ((tickCount - i) / tickCount);
    const y = top + (i / tickCount) * (height - top - bottom);
    svg += `<line x1="${{left}}" y1="${{y.toFixed(1)}}" x2="${{width-right}}" y2="${{y.toFixed(1)}}" stroke="#f0e6d2"></line>`;
    svg += `<text x="${{(left-8).toFixed(1)}}" y="${{(y+4).toFixed(1)}}" text-anchor="end" font-size="10" fill="#7a6b4d">${{escapeHtml(value.toFixed(1))}}</text>`;
  }}
  return svg;
}}
function renderTransferMarker(x, top, height, bottom, label) {{
  return `<line x1="${{x}}" y1="${{top}}" x2="${{x}}" y2="${{height-bottom}}" stroke="#b7b7b7" stroke-width="1.5" stroke-dasharray="6 5"></line>
    <text x="${{x}}" y="${{(top + 12).toFixed(1)}}" text-anchor="middle" font-size="10" fill="#8a8a8a">${{escapeHtml(label)}}</text>`;
}}
function renderChart(rows, metricName, seriesType, chartOptions = {{}}) {{
  if (!shouldRenderSeriesChart(rows, seriesType, metricName) && rows.length && (seriesType === "lab" || seriesType === "vital")) {{
    return `<div class="summary-line">该指标当前按表格展示，不额外绘制历时变化图。</div>`;
  }}
  const numericRows = rows.filter(r => r["图表模式"] !== "categorical" && r["数值结果"] !== null && r["数值结果"] !== "" && !Number.isNaN(Number(r["数值结果"])));
  if (!numericRows.length) {{
    return `<div class="summary-line">该指标无可直接绘制的连续数值，已以下方表格展示。</div>`;
  }}
  const width = 980, height = 320, left = 70, right = 30, top = 28, bottom = 82;
  const plotW = width - left - right, plotH = height - top - bottom;
  const values = numericRows.map(r => Number(r["数值结果"]));
  const lows = numericRows.map(r => Number(r["正常值下限"])).filter(v => !Number.isNaN(v));
  const highs = numericRows.map(r => Number(r["正常值上限"])).filter(v => !Number.isNaN(v));
  let yMin = Math.min(...values, ...(lows.length ? lows : [Math.min(...values)]));
  let yMax = Math.max(...values, ...(highs.length ? highs : [Math.max(...values)]));
  if (yMin === yMax) yMax += 1;
  const yAxis = buildYAxis(left, right, top, bottom, width, height, yMin, yMax, 4);
  const xPos = i => left + (numericRows.length === 1 ? plotW / 2 : (i / (numericRows.length - 1)) * plotW);
  const yPos = v => top + ((yMax - v) / (yMax - yMin)) * plotH;
  const path = numericRows.map((r, i) => `${{i === 0 ? "M" : "L"}}${{xPos(i).toFixed(1)}},${{yPos(Number(r["数值结果"])).toFixed(1)}}`).join(" ");
  const transferIndex = chartOptions.showTransferLine ? numericRows.findIndex(r => r["访视编号"] === "D57") : -1;
  const transferLine = transferIndex >= 0 ? renderTransferMarker(xPos(transferIndex).toFixed(1), top, height, bottom, "转组") : "";
  let marks = "";
  numericRows.forEach((r, i) => {{
    const abnormal = r["异常方向"] === "高" || r["异常方向"] === "低";
    const keyResponse = String(r["关键应答判定"] || "");
    const isIgaTs = keyResponse.includes("IGA-TS");
    const isEasi75 = keyResponse.includes("EASI-75");
    const color = abnormal ? "#d4380d" : (isEasi75 || isIgaTs) ? "#1f6b30" : "#FF9900";
    const cx = xPos(i).toFixed(1);
    const cy = yPos(Number(r["数值结果"])).toFixed(1);
    const tip = seriesType === "efficacy"
      ? `${{r["受试者编号"]}} | ${{r["访视名称"]}} | ${{r["日期"]}} | ${{metricName}} | 原始值：${{r["原始值"]}} | 应答：${{r["关键应答判定"] || ""}}`
      : `${{r["受试者编号"]}} | ${{r["访视名称"]}} | ${{r["日期"]}} | ${{metricName}} | ${{r["结果值"]}} ${{r["单位"]}} | 临床意义：${{r["临床意义判断"] || ""}} | 评估：${{r["研究者临床评估"] || ""}}`;
    marks += `<circle cx="${{cx}}" cy="${{cy}}" r="5" fill="${{color}}"><title>${{escapeHtml(tip)}}</title></circle>`;
    const labelValue = seriesType === "efficacy"
      ? (metricName === "IGA" ? String(r["数值结果"] ?? "") : String(r["数值结果"] ?? r["原始值"] ?? ""))
      : r["结果值"];
    marks += `<text x="${{cx}}" y="${{(Number(cy)-10).toFixed(1)}}" text-anchor="middle" font-size="11" fill="${{color}}">${{escapeHtml(labelValue)}}</text>`;
    if (seriesType === "efficacy" && (isEasi75 || isIgaTs)) {{
      const responseLabel = isIgaTs ? "IGA-TS" : "EASI-75";
      marks += `<text x="${{cx}}" y="${{(Number(cy)+16).toFixed(1)}}" text-anchor="middle" font-size="10" fill="#1f6b30">${{responseLabel}}</text>`;
    }}
    marks += `<text x="${{cx}}" y="${{height-34}}" text-anchor="middle" font-size="10" fill="#555">${{escapeHtml(r["访视编号"])}}</text>`;
    marks += `<text x="${{cx}}" y="${{height-20}}" text-anchor="middle" font-size="10" fill="#777">${{escapeHtml(r["日期"])}}</text>`;
  }});
  let rangeLines = "";
  if (seriesType !== "efficacy" && lows.length) {{
    const low = Math.min(...lows);
    rangeLines += `<line x1="${{left}}" y1="${{yPos(low).toFixed(1)}}" x2="${{width-right}}" y2="${{yPos(low).toFixed(1)}}" stroke="#9c7d18" stroke-dasharray="5 4"></line>`;
    rangeLines += `<text x="${{left-6}}" y="${{(yPos(low)-4).toFixed(1)}}" text-anchor="end" font-size="11" fill="#9c7d18">LLN ${{low}}</text>`;
  }}
  if (seriesType !== "efficacy" && highs.length) {{
    const high = Math.max(...highs);
    rangeLines += `<line x1="${{left}}" y1="${{yPos(high).toFixed(1)}}" x2="${{width-right}}" y2="${{yPos(high).toFixed(1)}}" stroke="#9c7d18" stroke-dasharray="5 4"></line>`;
    rangeLines += `<text x="${{left-6}}" y="${{(yPos(high)-4).toFixed(1)}}" text-anchor="end" font-size="11" fill="#9c7d18">ULN ${{high}}</text>`;
  }}
  const caption = seriesType === "efficacy"
    ? "数据点已标注具体数值；如该指标存在方案定义的关键应答，达到应答的点会额外标识。"
    : "数据点标注具体数值，X轴标注访视编号和对应日期；若仅有正常值下限或上限，则只绘制存在的一条参考线。";
  return `<div class="chart-box"><div><strong>${{escapeHtml(metricName)}}</strong></div>
    <svg viewBox="0 0 ${{width}} ${{height}}">
      ${{yAxis}}
      <line x1="${{left}}" y1="${{height-bottom}}" x2="${{width-right}}" y2="${{height-bottom}}" stroke="#cdbf9d"></line>
      <line x1="${{left}}" y1="${{top}}" x2="${{left}}" y2="${{height-bottom}}" stroke="#cdbf9d"></line>
      ${{rangeLines}}
      ${{transferLine}}
      <path d="${{path}}" fill="none" stroke="#FF9900" stroke-width="3"></path>
      ${{marks}}
    </svg>
    <div class="chart-caption">${{caption}}</div>
  </div>`;
}}
function renderAggregateContinuousChart(rows, metricName) {{
  if (!rows.length) return `<div class="summary-line">该指标当前无可汇总的连续数据。</div>`;
  const trialColor = "#FF9900";
  const trialTextColor = "#8a5a00";
  const controlColor = "#8A8A8A";
  const width = 980, height = 320, left = 70, right = 30, top = 28, bottom = 82;
  const plotW = width - left - right, plotH = height - top - bottom;
  const trialMeans = rows.map(r => Number(r["试验组均值数值"])).filter(v => !Number.isNaN(v));
  const controlMeans = rows.map(r => Number(r["对照组均值数值"])).filter(v => !Number.isNaN(v));
  const trialSds = rows.map(r => Number(r["试验组标准差数值"] || 0));
  const controlSds = rows.map(r => Number(r["对照组标准差数值"] || 0));
  const mins = rows.flatMap((r, i) => [
    Number.isNaN(Number(r["试验组均值数值"])) ? null : Number(r["试验组均值数值"]) - trialSds[i],
    Number.isNaN(Number(r["对照组均值数值"])) ? null : Number(r["对照组均值数值"]) - controlSds[i],
  ]).filter(v => v !== null);
  const maxs = rows.flatMap((r, i) => [
    Number.isNaN(Number(r["试验组均值数值"])) ? null : Number(r["试验组均值数值"]) + trialSds[i],
    Number.isNaN(Number(r["对照组均值数值"])) ? null : Number(r["对照组均值数值"]) + controlSds[i],
  ]).filter(v => v !== null);
  if (!mins.length || !maxs.length) return `<div class="summary-line">该指标当前无可汇总的连续数据。</div>`;
  let yMin = Math.min(...mins);
  let yMax = Math.max(...maxs);
  if (yMin === yMax) yMax += 1;
  const yAxis = buildYAxis(left, right, top, bottom, width, height, yMin, yMax, 4);
  const xPos = i => left + (rows.length === 1 ? plotW / 2 : (i / (rows.length - 1)) * plotW);
  const yPos = v => top + ((yMax - v) / (yMax - yMin)) * plotH;
  const trialPath = rows.map((r, i) => Number.isNaN(Number(r["试验组均值数值"])) ? null : `${{i === 0 ? "M" : "L"}}${{xPos(i).toFixed(1)}},${{yPos(Number(r["试验组均值数值"])).toFixed(1)}}`).filter(Boolean).join(" ");
  const controlPath = rows.map((r, i) => Number.isNaN(Number(r["对照组均值数值"])) ? null : `${{i === 0 ? "M" : "L"}}${{xPos(i).toFixed(1)}},${{yPos(Number(r["对照组均值数值"])).toFixed(1)}}`).filter(Boolean).join(" ");
  const transferIndex = rows.findIndex(r => r["访视编号"] === "D57");
  const transferLine = transferIndex >= 0 ? renderTransferMarker(xPos(transferIndex).toFixed(1), top, height, bottom, "转组") : "";
  let marks = "";
  rows.forEach((r, i) => {{
    const cx = xPos(i);
    marks += `<text x="${{cx.toFixed(1)}}" y="${{height-34}}" text-anchor="middle" font-size="10" fill="#555">${{escapeHtml(r["访视编号"])}}</text>`;
    if (!Number.isNaN(Number(r["试验组均值数值"]))) {{
      const mean = Number(r["试验组均值数值"]);
      const sd = Number(r["试验组标准差数值"] || 0);
      const cy = yPos(mean).toFixed(1);
      const yLow = yPos(mean - sd).toFixed(1);
      const yHigh = yPos(mean + sd).toFixed(1);
      const tip = `试验组 | ${{metricName}} | ${{r["访视名称"]}} | 均值 ${{r["试验组均值"]}} | 中位数 ${{r["试验组中位数"]}} | SD ${{r["试验组标准差"]}} | 最小-最大 ${{r["试验组最小值"]}}-${{r["试验组最大值"]}}`;
      marks += `<line x1="${{(cx-8).toFixed(1)}}" y1="${{yLow}}" x2="${{(cx-8).toFixed(1)}}" y2="${{yHigh}}" stroke="${{trialColor}}" stroke-width="2"></line>`;
      marks += `<line x1="${{(cx-12).toFixed(1)}}" y1="${{yLow}}" x2="${{(cx-4).toFixed(1)}}" y2="${{yLow}}" stroke="${{trialColor}}" stroke-width="2"></line>`;
      marks += `<line x1="${{(cx-12).toFixed(1)}}" y1="${{yHigh}}" x2="${{(cx-4).toFixed(1)}}" y2="${{yHigh}}" stroke="${{trialColor}}" stroke-width="2"></line>`;
      marks += `<circle cx="${{(cx-8).toFixed(1)}}" cy="${{cy}}" r="5" fill="${{trialColor}}"><title>${{escapeHtml(tip)}}</title></circle>`;
      marks += `<text x="${{(cx-8).toFixed(1)}}" y="${{(Number(cy)-10).toFixed(1)}}" text-anchor="middle" font-size="11" fill="${{trialTextColor}}">${{escapeHtml(r["试验组均值"])}}</text>`;
    }}
    if (!Number.isNaN(Number(r["对照组均值数值"]))) {{
      const mean = Number(r["对照组均值数值"]);
      const sd = Number(r["对照组标准差数值"] || 0);
      const cy = yPos(mean).toFixed(1);
      const yLow = yPos(mean - sd).toFixed(1);
      const yHigh = yPos(mean + sd).toFixed(1);
      const tip = `对照组 | ${{metricName}} | ${{r["访视名称"]}} | 均值 ${{r["对照组均值"]}} | 中位数 ${{r["对照组中位数"]}} | SD ${{r["对照组标准差"]}} | 最小-最大 ${{r["对照组最小值"]}}-${{r["对照组最大值"]}}`;
      marks += `<line x1="${{(cx+8).toFixed(1)}}" y1="${{yLow}}" x2="${{(cx+8).toFixed(1)}}" y2="${{yHigh}}" stroke="${{controlColor}}" stroke-width="2"></line>`;
      marks += `<line x1="${{(cx+4).toFixed(1)}}" y1="${{yLow}}" x2="${{(cx+12).toFixed(1)}}" y2="${{yLow}}" stroke="${{controlColor}}" stroke-width="2"></line>`;
      marks += `<line x1="${{(cx+4).toFixed(1)}}" y1="${{yHigh}}" x2="${{(cx+12).toFixed(1)}}" y2="${{yHigh}}" stroke="${{controlColor}}" stroke-width="2"></line>`;
      marks += `<circle cx="${{(cx+8).toFixed(1)}}" cy="${{cy}}" r="5" fill="${{controlColor}}"><title>${{escapeHtml(tip)}}</title></circle>`;
      marks += `<text x="${{(cx+8).toFixed(1)}}" y="${{(Number(cy)-10).toFixed(1)}}" text-anchor="middle" font-size="11" fill="${{controlColor}}">${{escapeHtml(r["对照组均值"])}}</text>`;
    }}
  }});
  return `<div class="chart-box"><div><strong>${{escapeHtml(metricName)}}</strong></div>
    <svg viewBox="0 0 ${{width}} ${{height}}">
      ${{yAxis}}
      <line x1="${{left}}" y1="${{height-bottom}}" x2="${{width-right}}" y2="${{height-bottom}}" stroke="#cdbf9d"></line>
      <line x1="${{left}}" y1="${{top}}" x2="${{left}}" y2="${{height-bottom}}" stroke="#cdbf9d"></line>
      ${{transferLine}}
      <path d="${{trialPath}}" fill="none" stroke="${{trialColor}}" stroke-width="3"></path>
      <path d="${{controlPath}}" fill="none" stroke="${{controlColor}}" stroke-width="3"></path>
      ${{marks}}
    </svg>
    <div class="chart-caption">历时图展示试验组与对照组各访视均值±标准差；鼠标停留可查看均值、标准差、中位数和最大最小值。</div>
  </div>`;
}}
function renderAggregateBinaryChart(rows, responseName) {{
  if (!rows.length) return `<div class="summary-line">该应答当前无可汇总数据。</div>`;
  const trialFill = "#FF9900";
  const controlFill = "#8A8A8A";
  const width = 980, height = 320, left = 70, right = 30, top = 28, bottom = 82;
  const plotW = width - left - right, plotH = height - top - bottom;
  const yAxis = buildYAxis(left, right, top, bottom, width, height, 0, 100, 4);
  const xPos = i => left + (i + 0.5) * (plotW / rows.length);
  const barW = Math.max(24, plotW / Math.max(rows.length * 2.2, 1));
  const yPos = v => top + ((100 - v) / 100) * plotH;
  const transferIndex = rows.findIndex(r => r["访视编号"] === "D57");
  const transferLine = transferIndex >= 0 ? renderTransferMarker(xPos(transferIndex).toFixed(1), top, height, bottom, "转组") : "";
  let bars = "";
  rows.forEach((r, i) => {{
    const cx = xPos(i);
    const trialPct = Number(r["试验组应答比例数值"] || 0);
    const controlPct = Number(r["对照组应答比例数值"] || 0);
    const yTrial = yPos(trialPct);
    const yControl = yPos(controlPct);
    const hTrial = height - bottom - yTrial;
    const hControl = height - bottom - yControl;
    const tipTrial = `试验组 | ${{responseName}} | ${{r["访视名称"]}} | 应答 ${{r["试验组应答例数"]}} | 例次 ${{r["试验组例次（缺失）"]}} | 比例 ${{r["试验组应答比例"]}}`;
    const tipControl = `对照组 | ${{responseName}} | ${{r["访视名称"]}} | 应答 ${{r["对照组应答例数"]}} | 例次 ${{r["对照组例次（缺失）"]}} | 比例 ${{r["对照组应答比例"]}}`;
    bars += `<rect x="${{(cx - barW).toFixed(1)}}" y="${{yTrial.toFixed(1)}}" width="${{(barW-4).toFixed(1)}}" height="${{hTrial.toFixed(1)}}" fill="${{trialFill}}" stroke="#d97f00"><title>${{escapeHtml(tipTrial)}}</title></rect>`;
    bars += `<rect x="${{(cx + 4).toFixed(1)}}" y="${{yControl.toFixed(1)}}" width="${{(barW-4).toFixed(1)}}" height="${{hControl.toFixed(1)}}" fill="${{controlFill}}" stroke="#2f2f2f"><title>${{escapeHtml(tipControl)}}</title></rect>`;
    bars += `<text x="${{(cx - barW/2).toFixed(1)}}" y="${{(yTrial - 10).toFixed(1)}}" text-anchor="middle" font-size="11" fill="#8a5a00">${{escapeHtml(r["试验组应答比例"])}}</text>`;
    bars += `<text x="${{(cx + barW/2).toFixed(1)}}" y="${{(yControl - 10).toFixed(1)}}" text-anchor="middle" font-size="11" fill="${{controlFill}}">${{escapeHtml(r["对照组应答比例"])}}</text>`;
    bars += `<text x="${{cx.toFixed(1)}}" y="${{height-34}}" text-anchor="middle" font-size="10" fill="#555">${{escapeHtml(r["访视编号"])}}</text>`;
  }});
  return `<div class="chart-box"><div><strong>${{escapeHtml(responseName)}}</strong></div>
    <svg viewBox="0 0 ${{width}} ${{height}}">
      ${{yAxis}}
      <line x1="${{left}}" y1="${{height-bottom}}" x2="${{width-right}}" y2="${{height-bottom}}" stroke="#cdbf9d"></line>
      <line x1="${{left}}" y1="${{top}}" x2="${{left}}" y2="${{height-bottom}}" stroke="#cdbf9d"></line>
      ${{transferLine}}
      ${{bars}}
    </svg>
    <div class="chart-caption">柱状图展示试验组与对照组各访视应答比例；鼠标停留可查看应答例数和比例。</div>
  </div>`;
}}
function centerAggregateMetricRows(center, bucket, metricName) {{
  const group = ((DATA.center_aggregate || {{}})[center] || {{}})[bucket] || [];
  return group.filter(item => item.metric === metricName);
}}
function renderCenterSummary(center, filters) {{
  const data = (DATA.center_aggregate || {{}})[center];
  if (!data) return `<div class="summary-line">当前未找到该中心的汇总数据。</div>`;
  const p = data.profile_summary || {{}};
  const flow = `<div class="flow">
    <div class="flow-step"><div class="num">${{escapeHtml(p["筛选失败例数"] ?? "")}}</div><div class="label">筛选失败</div></div>
    <div class="flow-step"><div class="num">${{escapeHtml(p["进入治疗例数"] ?? "")}}</div><div class="label">进入治疗</div></div>
    <div class="flow-step"><div class="num">${{escapeHtml(p["研究结束例数"] ?? "")}}</div><div class="label">研究结束</div></div>
  </div>`;
  const profileCards = [
    ["年龄均值", `全部受试者：${{escapeHtml(p["年龄均值"] ?? "")}}<br>剔除筛败后（n=${{escapeHtml(p["剔除筛败后受试者数"] ?? "")}}）：${{escapeHtml(p["年龄均值_剔除筛败后"] ?? "")}}`],
    ["性别比例", `全部受试者：${{escapeHtml(p["性别分布"] ?? "")}}<br>剔除筛败后（n=${{escapeHtml(p["剔除筛败后受试者数"] ?? "")}}）：${{escapeHtml(p["性别分布_剔除筛败后"] ?? "")}}`],
    ["年龄组比例", `全部受试者：${{escapeHtml(p["年龄组分布"] ?? "")}}<br>剔除筛败后（n=${{escapeHtml(p["剔除筛败后受试者数"] ?? "")}}）：${{escapeHtml(p["年龄组分布_剔除筛败后"] ?? "")}}`],
    ["受试者总数", p["受试者总数"]],
  ].map(([k,v]) => `<div class="field"><div class="k">${{escapeHtml(k)}}</div><div class="v">${{v ?? ""}}</div></div>`).join("");

  const renderContinuousBlocks = (items, selectedMetric) => {{
    const chosen = selectedMetric === "__ALL__" ? items : items.filter(item => item.metric === selectedMetric);
    return chosen.map(item => {{
      const headerRows = [
        [
          {{ label: "访视编号", rowspan: 2, key: "访视编号", colIndex: 0 }},
          {{ label: "访视名称", rowspan: 2, key: "访视名称", colIndex: 1 }},
          {{ label: "例次（缺失）", colspan: 2 }},
          {{ label: "均值（标准差）", colspan: 2 }},
          {{ label: "中位数（最大值, 最小值）", colspan: 2 }},
        ],
        [
          {{ label: "试验组", key: "试验组例次（缺失）", colIndex: 2 }},
          {{ label: "对照组", key: "对照组例次（缺失）", colIndex: 3 }},
          {{ label: "试验组", key: "试验组均值标准差组合", colIndex: 4 }},
          {{ label: "对照组", key: "对照组均值标准差组合", colIndex: 5 }},
          {{ label: "试验组", key: "试验组中位数极值组合", colIndex: 6 }},
          {{ label: "对照组", key: "对照组中位数极值组合", colIndex: 7 }},
        ],
      ];
      const leafHeaders = [
        {{ key: "访视编号" }},
        {{ key: "访视名称" }},
        {{ key: "试验组例次（缺失）" }},
        {{ key: "对照组例次（缺失）" }},
        {{ key: "试验组均值标准差组合", render: row => row["试验组均值"] ? `${{row["试验组均值"]}} (${{row["试验组标准差"] || ""}})` : "" }},
        {{ key: "对照组均值标准差组合", render: row => row["对照组均值"] ? `${{row["对照组均值"]}} (${{row["对照组标准差"] || ""}})` : "" }},
        {{ key: "试验组中位数极值组合", render: row => row["试验组中位数"] ? `${{row["试验组中位数"]}} (${{row["试验组最大值"] || ""}}, ${{row["试验组最小值"] || ""}})` : "" }},
        {{ key: "对照组中位数极值组合", render: row => row["对照组中位数"] ? `${{row["对照组中位数"]}} (${{row["对照组最大值"] || ""}}, ${{row["对照组最小值"] || ""}})` : "" }},
      ];
      return `<div class="module-title">${{escapeHtml(item.metric)}}</div>${{renderAggregateContinuousChart(item.rows, item.metric)}}${{makeGroupedTable(headerRows, leafHeaders, item.rows, `agg-${{center}}-${{item.metric}}`)}}`;
    }}).join("");
  }};
  const renderBinaryBlocks = (items, selectedMetric) => {{
    const chosen = selectedMetric === "__ALL__" ? items : items.filter(item => item.metric === selectedMetric);
    return chosen.map(item => {{
      const headerRows = [
        [
          {{ label: "访视编号", rowspan: 2, key: "访视编号", colIndex: 0 }},
          {{ label: "访视名称", rowspan: 2, key: "访视名称", colIndex: 1 }},
          {{ label: "应答例数", colspan: 2 }},
          {{ label: "例次（缺失）", colspan: 2 }},
          {{ label: "比例", colspan: 2 }},
        ],
        [
          {{ label: "试验组", key: "试验组应答例数", colIndex: 2 }},
          {{ label: "对照组", key: "对照组应答例数", colIndex: 3 }},
          {{ label: "试验组", key: "试验组例次（缺失）", colIndex: 4 }},
          {{ label: "对照组", key: "对照组例次（缺失）", colIndex: 5 }},
          {{ label: "试验组", key: "试验组应答比例", colIndex: 6 }},
          {{ label: "对照组", key: "对照组应答比例", colIndex: 7 }},
        ],
      ];
      const leafHeaders = [
        {{ key: "访视编号" }},
        {{ key: "访视名称" }},
        {{ key: "试验组应答例数" }},
        {{ key: "对照组应答例数" }},
        {{ key: "试验组例次（缺失）" }},
        {{ key: "对照组例次（缺失）" }},
        {{ key: "试验组应答比例" }},
        {{ key: "对照组应答比例" }},
      ];
      return `<div class="module-title">${{escapeHtml(item.response_name)}}</div>${{renderAggregateBinaryChart(item.rows, item.response_name)}}${{makeGroupedTable(headerRows, leafHeaders, item.rows, `agg-bin-${{center}}-${{item.metric}}-${{item.response_name}}`)}}`;
    }}).join("");
  }};

  let efficacySection = "";
  if (filters.efficacyMetric !== "__NONE__") {{
    efficacySection = `<section><div class="module-title">疗效汇总</div>${{renderContinuousBlocks(data.efficacy_continuous || [], filters.efficacyMetric)}}${{renderBinaryBlocks(data.efficacy_binary || [], filters.efficacyMetric)}}</section>`;
  }}
  let labSection = "";
  if (filters.labMetric !== "__NONE__") {{
    labSection = `<section><div class="module-title">安全性汇总</div>${{renderContinuousBlocks(data.lab_continuous || [], filters.labMetric) || "<div class='summary-line'>当前筛选下无连续安全性指标汇总。</div>"}}</section>`;
  }}
  let vitalSection = "";
  if (filters.vitalMetric !== "__NONE__") {{
    vitalSection = `<section><div class="module-title">生命体征汇总</div>${{renderContinuousBlocks(data.vital_continuous || [], filters.vitalMetric) || "<div class='summary-line'>当前筛选下无生命体征汇总。</div>"}}</section>`;
  }}

  return `<article class="subject-card">
    <div class="subject-head">
      <div><h3>${{escapeHtml(center)}} / 全部受试者汇总</h3></div>
    </div>
    <section><div class="module-title">Profiles汇总</div>${{flow}}<div class="grid">${{profileCards}}</div></section>
    ${{efficacySection}}
    ${{labSection}}
    ${{vitalSection}}
  </article>`;
}}
function basicFieldsForSubject(subject) {{
  if (subject["受试者状态"] === "筛选失败") {{
    return ["中心","受试者编号","年龄","年龄组","性别","知情同意日期","筛选失败原因","是否存在AE/PD/Finding"];
  }}
  return ["中心","受试者编号","年龄","年龄组","性别","随机组别","筛选日期","知情同意日期","基线/随机日期","双盲期进入日期","双盲期完成日期","开放期进入日期","开放期完成日期","是否完成双盲期","是否完成开放期","是否完成研究","是否进入ITT","是否进入PKCS","是否存在AE/PD/Finding","实验室特殊关注","生命体征特殊关注"];
}}
function subjectTags(subject) {{
  const tags = [];
  if (subject["随机组别"]) tags.push(`<span class="pill">${{escapeHtml(subject["随机组别"])}}</span>`);
  if (subject["年龄组"] === "青少年") tags.push(`<span class="pill adolescent">青少年受试者</span>`);
  if (subject["受试者状态"] === "筛选失败") tags.push(`<span class="pill high">筛选失败</span>`);
  if (subject["是否存在核查Finding"] === "是") tags.push(`<span class="pill mid">有Finding</span>`);
  if ((subject["是否存在AE/SAE/SUSAR"] || "否") !== "否") tags.push(`<span class="pill low">${{escapeHtml(subject["是否存在AE/SAE/SUSAR"])}}</span>`);
  return tags.join("");
}}
function renderSubjectCard(subject, subjectFindings, subjectEfficacy, subjectLabs, subjectVitals, subjectAEs, filters) {{
  const basicGrid = basicFieldsForSubject(subject).map(key => `<div class="field"><div class="k">${{escapeHtml(key)}}</div><div class="v">${{escapeHtml(subject[key] ?? "")}}</div></div>`).join("");
  let efficacySection = "";
  if (subject["受试者状态"] !== "筛选失败" && filters.efficacyMetric !== "__NONE__") {{
    const metrics = unique(subjectEfficacy.map(r => r["指标名称"]));
    const chosen = filters.efficacyMetric === "__ALL__" ? metrics : metrics.filter(m => m === filters.efficacyMetric);
    const blocks = chosen.map(metric => {{
      const rows = subjectEfficacy.filter(r => r["指标名称"] === metric);
      const headers = efficacyHeaders(metric);
      return `<div class="module-title">${{escapeHtml(metric)}}</div>${{renderChart(rows, metric, "efficacy", {{ showTransferLine: String(subject["随机组别"] || "").includes("安慰剂") }})}}${{makeTable(headers, rows, `eff-${{subject["受试者编号"]}}-${{metric}}`)}}`;
    }}).join("");
    efficacySection = `<section><div class="module-title">疗效数据模块</div>${{blocks || "<div class='summary-line'>当前筛选下无疗效数据。</div>"}}</section>`;
  }}
  let labSection = "";
  if (subject["受试者状态"] !== "筛选失败" && filters.labMetric !== "__NONE__") {{
    const metrics = unique(subjectLabs.map(r => r["检验项目标准化名称"]));
    const chosen = filters.labMetric === "__ALL__" ? metrics : metrics.filter(m => m === filters.labMetric);
    const blocks = chosen.map(metric => {{
      const rows = subjectLabs.filter(r => r["检验项目标准化名称"] === metric);
      const headers = ["实验室显示分组","访视编号","访视名称","日期","检验项目标准化名称","原始检验项目名称","结果值","单位","正常值下限","正常值上限","异常方向","临床意义判断","研究者临床评估","额外关注标记","备注","是否对应AE或Finding","数据来源sheet"];
      return `<div class="module-title">${{escapeHtml(metric)}}</div>${{renderChart(rows, metric, "lab", {{ showTransferLine: String(subject["随机组别"] || "").includes("安慰剂") }})}}${{makeTable(headers, rows, `lab-${{subject["受试者编号"]}}-${{metric}}`)}}`;
    }}).join("");
    labSection = `<section><div class="module-title">实验室检查数据模块</div>${{blocks || "<div class='summary-line'>当前筛选下无实验室数据。</div>"}}</section>`;
  }}
  let vitalSection = "";
  if (subject["受试者状态"] !== "筛选失败" && filters.vitalMetric !== "__NONE__") {{
    const metrics = unique(subjectVitals.map(r => r["生命体征指标"]));
    const chosen = filters.vitalMetric === "__ALL__" ? metrics : metrics.filter(m => m === filters.vitalMetric);
    const blocks = chosen.map(metric => {{
      const rows = subjectVitals.filter(r => r["生命体征指标"] === metric);
      const headers = ["访视编号","访视名称","日期","生命体征指标","原始指标名称","结果值","单位","正常值下限","正常值上限","异常方向","临床意义判断","研究者临床评估","额外关注标记","备注","是否对应AE或Finding","数据来源sheet"];
      return `<div class="module-title">${{escapeHtml(metric)}}</div>${{renderChart(rows, metric, "vital", {{ showTransferLine: String(subject["随机组别"] || "").includes("安慰剂") }})}}${{makeTable(headers, rows, `vital-${{subject["受试者编号"]}}-${{metric}}`)}}`;
    }}).join("");
    vitalSection = `<section><div class="module-title">生命体征模块</div>${{blocks || "<div class='summary-line'>当前筛选下无生命体征数据。</div>"}}</section>`;
  }}
  let aeSection = "";
  if (subject["受试者状态"] === "筛选失败") {{
    const headers = ["AE名称","最早开始日期","最严重程度（CTCAE V5.0）","与研究用药的关系","转归","结束日期","是否为严重不良事件（SAE）"];
    aeSection = `<section><div class="module-title">AE信息</div>${{subjectAEs.length ? makeTable(headers, subjectAEs, `ae-${{subject["受试者编号"]}}`) : "<div class='summary-line'>该筛选失败受试者未匹配到AE记录。</div>"}}</section>`;
  }}
  const findingHeaders = ["Finding编号","问题分类","问题严重程度","原始问题描述","涉及访视","问题状态","是否影响受试者权益","是否影响安全性","是否影响入组合格性","是否影响主要终点","是否影响关键次要终点","是否影响实验室安全性判断","是否影响用药依从性/暴露","是否影响数据完整性","是否需监管备答","建议现场核查回复口径","现场回复示例","残余风险等级"];
  const findingSection = `<section class="finding-fixed" data-section="findings-fixed"><div class="module-title">Finding固定展示模块</div>${{subjectFindings.length ? makeTable(findingHeaders, subjectFindings, `find-${{subject["受试者编号"]}}`) : "<div class='summary-line'>该受试者当前未匹配到 subject-level Finding。</div>"}}</section>`;
  return `<article class="subject-card">
    <div class="subject-head">
      <div><h3>${{escapeHtml(subject["受试者编号"])}}</h3><div class="summary-line">${{escapeHtml(subject["中心"])}}</div></div>
      <div>${{subjectTags(subject)}}</div>
    </div>
    <div class="grid">${{basicGrid}}</div>
    ${{aeSection}}
    ${{efficacySection}}
    ${{labSection}}
    ${{vitalSection}}
    ${{findingSection}}
  </article>`;
}}
function bindTableSorting() {{
  document.querySelectorAll("th[data-key]").forEach(th => {{
    th.onclick = () => {{
      const table = th.closest("table");
      const tbody = table.querySelector("tbody");
      const index = th.dataset.colIndex !== undefined && th.dataset.colIndex !== ""
        ? Number(th.dataset.colIndex)
        : Array.from(th.parentNode.children).indexOf(th);
      const rows = Array.from(tbody.querySelectorAll("tr"));
      const asc = th.dataset.asc !== "true";
      rows.sort((a, b) => {{
        const av = a.children[index].innerText.trim();
        const bv = b.children[index].innerText.trim();
        return asc ? av.localeCompare(bv, "zh-CN", {{numeric:true}}) : bv.localeCompare(av, "zh-CN", {{numeric:true}});
      }});
      th.dataset.asc = asc ? "true" : "false";
      tbody.innerHTML = "";
      rows.forEach(r => tbody.appendChild(r));
    }};
  }});
}}
function renderStaticAppendices() {{
  document.getElementById("mappingTable").innerHTML = makeTable(["文件名","sheet名","数据类别","原始字段名","标准化字段名","数据类型","是否用于HTML展示","是否需要人工复核","备注"], DATA.field_mapping.slice(0, 500), "mapping");
}}
function render() {{
  const filters = currentFilters();
  let subjects = filterSubjectsByCenter(filters.center);
  if (filters.subject !== "__ALL__") subjects = subjects.filter(s => s["受试者编号"] === filters.subject);
  if (filters.search) subjects = subjects.filter(s => JSON.stringify(s).toLowerCase().includes(filters.search));
  const isCenterSummary = filters.subject === "__ALL__" && filters.center !== "__ALL__";
  document.getElementById("currentSummary").textContent = isCenterSummary
    ? `中心 ${{filters.center}}；当前为全部受试者汇总视图。`
    : `中心 ${{filters.center === "__ALL__" ? "全部中心" : filters.center}}；当前展示 ${{subjects.length}} 例受试者。`;
  if (filters.subject === "__ALL__") {{
    if (filters.center === "__ALL__") {{
      document.getElementById("profilesPanel").innerHTML = `<h2>中心汇总</h2><div class="summary-line">请选择具体中心后，再查看“全部受试者”的中心级汇总。</div>`;
      return;
    }}
    document.getElementById("profilesPanel").innerHTML = `<h2>中心汇总</h2>${{renderCenterSummary(filters.center, filters)}}`;
    bindTableSorting();
    return;
  }}
  const cards = subjects.map(subject => {{
    const key = `${{subject["中心"]}}|${{subject["受试者编号"]}}`;
    const subjectFindings = DATA.findings.filter(r => r["受试者键"] === key && findingPass(r, filters.finding) && JSON.stringify(r).toLowerCase().includes(filters.search || ""));
    const subjectEfficacy = DATA.efficacy.filter(r => r["受试者键"] === key && (filters.efficacyMetric === "__ALL__" || r["指标名称"] === filters.efficacyMetric) && JSON.stringify(r).toLowerCase().includes(filters.search || ""));
    const subjectLabs = DATA.labs.filter(r => r["受试者键"] === key && (filters.labMetric === "__ALL__" || r["检验项目标准化名称"] === filters.labMetric) && labPass(r, filters.abnormal) && JSON.stringify(r).toLowerCase().includes(filters.search || ""));
    const subjectVitals = DATA.vitals.filter(r => r["受试者键"] === key && (filters.vitalMetric === "__ALL__" || r["生命体征指标"] === filters.vitalMetric) && labPass(r, filters.abnormal) && JSON.stringify(r).toLowerCase().includes(filters.search || ""));
    const subjectAEs = DATA.ae_events.filter(r => r["受试者键"] === key && JSON.stringify(r).toLowerCase().includes(filters.search || ""));
    return renderSubjectCard(subject, subjectFindings, subjectEfficacy, subjectLabs, subjectVitals, subjectAEs, filters);
  }}).join("");
  document.getElementById("profilesPanel").innerHTML = `<h2>受试者 Profile</h2>${{cards || "<div class='summary-line'>当前筛选下无可展示的受试者。</div>"}}`;
  bindTableSorting();
}}
document.getElementById("centerSelect").addEventListener("change", () => {{ syncSubjectOptions(); render(); }});
["subjectSelect","efficacyMetricSelect","labMetricSelect","vitalMetricSelect","abnormalFilter","findingFilter"].forEach(id => document.getElementById(id).addEventListener("change", render));
document.getElementById("globalSearch").addEventListener("input", render);
initSelectors();
renderStaticAppendices();
render();
</script>
</body>
</html>
"""


def build_payload() -> dict[str, Any]:
    reference_groups = parse_reference_html_groups()
    subject_profiles = parse_subject_profiles(reference_groups)
    visit_index = build_visit_index()
    efficacy_rows = build_efficacy_rows(visit_index)
    lab_rows = build_lab_rows(visit_index)
    vital_rows = build_vital_rows(visit_index)
    finding_rows = build_finding_rows()
    ae_rows = build_ae_rows()
    attach_finding_links(subject_profiles, efficacy_rows, lab_rows, finding_rows, ae_rows, vital_rows)
    summarize_lab_attention_by_subject(subject_profiles, lab_rows)
    summarize_vital_attention_by_subject(subject_profiles, vital_rows)
    center_aggregate = build_center_aggregate_payload(subject_profiles, efficacy_rows, lab_rows, vital_rows)
    qc = run_qc_checks(subject_profiles, efficacy_rows, lab_rows, finding_rows)
    field_mapping_specs = []
    if LISTING_XLSX.exists():
        field_mapping_specs.extend(scan_workbook_structure(LISTING_XLSX, listing_mode=True))
    if FINDING_XLSX.exists():
        field_mapping_specs.extend(scan_workbook_structure(FINDING_XLSX, listing_mode=False))
    if PD_DEF_XLSX and PD_DEF_XLSX.exists():
        field_mapping_specs.extend(scan_workbook_structure(PD_DEF_XLSX, listing_mode=False))
    field_mapping = build_field_mapping_rows(field_mapping_specs)
    unresolved_markdown = build_unresolved_questions(field_mapping, qc, subject_profiles, finding_rows)
    efficacy_metric_count = len(unique([r["指标名称"] for r in efficacy_rows]))
    lab_metric_count = len(unique([r["检验项目标准化名称"] for r in lab_rows]))
    final_check_statement = {
        "已读取的数据源": "主listing + Finding台账 + 方案偏离分类表 + 受试者报表 + 旧版HTML只读参照",
        "已纳入的中心": "；".join(TARGET_CENTERS),
        "已纳入的受试者数量": str(len(subject_profiles)),
        "已纳入的疗效指标数量": str(efficacy_metric_count),
        "已纳入的实验室指标数量": str(lab_metric_count),
        "已纳入的Finding数量": str(len(finding_rows)),
        "已完成的QC项目": str(len(qc["summary"])),
        "仍需人工复核的问题": str(sum(1 for row in qc["summary"] if row["结果"] != "通过")),
        "核对声明": "已对字段结构、受试者主键、疗效与实验室纵表、Finding固定展示逻辑及核心QC项进行实际核对；未将结果笼统表述为全部准确。",
    }
    return {
        "generated_at": datetime.now().strftime("%Y-%m-%d %H:%M:%S"),
        "centers": TARGET_CENTERS,
        "protocol_summary": deepcopy(PROTOCOL_SUMMARY),
        "subjects": subject_profiles,
        "efficacy": efficacy_rows,
        "labs": lab_rows,
        "vitals": vital_rows,
        "center_aggregate": center_aggregate,
        "ae_events": ae_rows,
        "findings": finding_rows,
        "field_mapping": field_mapping,
        "qc_summary": qc["summary"],
        "issue_log": qc["issues"],
        "completeness": qc["completeness"],
        "unresolved_markdown": unresolved_markdown,
        "final_check_statement": final_check_statement,
    }


def main() -> None:
    args = parse_args()
    runtime = configure_runtime(args)
    precheck = build_precheck_summary(runtime)
    OUTPUT_DIR.mkdir(parents=True, exist_ok=True)
    write_precheck_outputs(precheck)
    if args.precheck_only:
        return
    if not precheck["ready_to_build"]:
        raise SystemExit("Input precheck failed. Review output/input_precheck.md and resolve blocking issues before building.")
    payload = build_payload()
    write_csv(OUTPUT_DIR / "cleaned_subject_profile_dataset.csv", payload["subjects"])
    write_csv(OUTPUT_DIR / "efficacy_longitudinal_dataset.csv", payload["efficacy"])
    write_csv(OUTPUT_DIR / "lab_longitudinal_dataset.csv", payload["labs"])
    write_csv(OUTPUT_DIR / "vital_signs_longitudinal_dataset.csv", payload["vitals"])
    write_csv(OUTPUT_DIR / "finding_subject_level_dataset.csv", payload["findings"])
    write_csv(OUTPUT_DIR / "data_source_field_mapping.csv", payload["field_mapping"])
    write_csv(OUTPUT_DIR / "data_qc_summary.csv", payload["qc_summary"])
    write_csv(OUTPUT_DIR / "data_issue_log.csv", payload["issue_log"])
    write_csv(OUTPUT_DIR / "subject_data_completeness.csv", payload["completeness"])
    (OUTPUT_DIR / "protocol_endpoint_summary.json").write_text(json.dumps(payload["protocol_summary"], ensure_ascii=False, indent=2), encoding="utf-8")
    (OUTPUT_DIR / "protocol_endpoint_summary.md").write_text(build_protocol_summary_markdown(payload["protocol_summary"]), encoding="utf-8")
    (OUTPUT_DIR / "unresolved_data_questions.md").write_text(payload["unresolved_markdown"], encoding="utf-8")
    (OUTPUT_DIR / "patient_profile.html").write_text(render_html(payload), encoding="utf-8")


if __name__ == "__main__":
    main()
